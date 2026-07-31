"""Finalize EdgeBench evidence and write comparison workbooks."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from . import io
from .context import current_paths
from .evidence import summarize_cell
from .profiles import (
    LEGACY_PAPER_PROTOCOL_ISSUES,
    PAPER_LARGE_GAP_THRESHOLD_PP,
)


def load_paper_reference(path: Path | None = None) -> dict[str, Any]:
    selected_path = path or current_paths().paper_reference_path
    payload = io.read_json(selected_path)
    reference = payload.get("reference", {})
    tasks = payload.get("tasks", {})
    if (
        payload.get("schema_version") != 1
        or reference.get("agent") != "Codex"
        or reference.get("model") != "GPT-5.5"
        or reference.get("budget_hours") != 12
        or not isinstance(tasks, dict)
        or not tasks
    ):
        raise ValueError(f"invalid EdgeBench paper reference: {selected_path}")
    for task, score in tasks.items():
        if not isinstance(score, dict) or not isinstance(
            score.get("mean"), (int, float)
        ):
            raise ValueError(f"invalid paper score for {task}: {selected_path}")
    return payload


def load_local_fast_reference(path: Path) -> dict[str, Any]:
    payload = io.read_json(path)
    reference = payload.get("reference", {})
    checkpoints = payload.get("checkpoints", {})
    if (
        payload.get("schema_version") != 2
        or reference.get("official_comparison") is not False
        or not isinstance(checkpoints, dict)
        or not checkpoints
    ):
        raise ValueError(f"invalid EdgeBench local fast reference: {path}")
    for label, checkpoint in checkpoints.items():
        tasks = checkpoint.get("tasks", {}) if isinstance(checkpoint, dict) else {}
        boundary = checkpoint.get("boundary_seconds") if isinstance(checkpoint, dict) else None
        if not isinstance(boundary, int) or not isinstance(tasks, dict):
            raise ValueError(f"invalid local fast checkpoint {label}: {path}")
        for task_id, record in tasks.items():
            if (
                not isinstance(record, dict)
                or record.get("task_id") != task_id
                or not isinstance(record.get("edgebench_score"), (int, float))
                or not 0 < int(record.get("checkpoint_seconds") or 0) <= boundary
            ):
                raise ValueError(
                    f"invalid local fast score for {task_id} at {label}: {path}"
                )
    return payload


def comparison_record(
    cell: dict[str, Any],
    paper_tasks: dict[str, Any],
    local_fast_checkpoints: dict[str, Any] | None = None,
) -> dict[str, Any]:
    best = cell.get("best") or {}
    observations = cell.get("observations", [])
    evaluator_calls = sum(
        int(item.get("evaluator_calls") or 0) for item in observations
    )
    runtime = sum(float(item.get("runtime_seconds") or 0) for item in observations)
    input_tokens = 0
    output_tokens = 0
    coverage: set[str] = set()
    for item in observations:
        usage = item.get("codex_usage") or {}
        tokens = usage.get("tokens") or {}
        input_tokens += int(tokens.get("input_tokens") or 0)
        output_tokens += int(tokens.get("output_tokens") or 0)
        if usage.get("coverage"):
            coverage.add(str(usage["coverage"]))
        annotator = (item.get("goal_plus") or {}).get(
            "evidence_annotator_usage"
        ) or {}
        input_tokens += int(annotator.get("input_tokens") or 0)
        output_tokens += int(annotator.get("output_tokens") or 0)
        if annotator.get("tasks"):
            coverage.add(str(annotator.get("coverage")))

    normalized = best.get("edgebench_score")
    official = best.get("official_comparison") or {}
    checkpoint_hours = official.get("checkpoint_hours")
    same_budget_score = (official.get("references") or {}).get("GPT-5.5")
    same_budget_delta = (
        float(normalized) - float(same_budget_score)
        if normalized is not None and same_budget_score is not None
        else None
    )
    paper_score = paper_tasks[cell["task_id"]]
    paper_delta = (
        float(normalized) - float(paper_score["mean"])
        if normalized is not None
        else None
    )
    local_fast_checkpoints = local_fast_checkpoints or {}
    local_half = (
        local_fast_checkpoints.get("0.5h", {})
        .get("tasks", {})
        .get(cell["task_id"], {})
    )
    local_one = (
        local_fast_checkpoints.get("1h", {})
        .get("tasks", {})
        .get(cell["task_id"], {})
    )
    local_half_score = local_half.get("edgebench_score")
    local_one_score = local_one.get("edgebench_score")
    local_half_delta = (
        float(normalized) - float(local_half_score)
        if normalized is not None and local_half_score is not None
        else None
    )
    local_one_delta = (
        float(normalized) - float(local_one_score)
        if normalized is not None and local_one_score is not None
        else None
    )
    completion = cell.get("completion_evidence") or {}
    completion_checks = completion.get("checks") or {}
    worker_check = completion_checks.get("actual_worker_launches") or (
        completion_checks.get("agent_sessions") or {}
    )
    if normalized is None:
        issue_marker = "MISSING_CURRENT"
    else:
        known_issue = cell.get("known_protocol_issue")
        if known_issue is None and cell.get("protocol_classification") is None:
            known_issue = LEGACY_PAPER_PROTOCOL_ISSUES.get(cell["task_id"])
        if known_issue:
            issue_marker = f"KNOWN_PROTOCOL: {known_issue}"
        elif same_budget_delta is None:
            issue_marker = "MISSING_SAME_BUDGET_REFERENCE"
        elif same_budget_delta >= PAPER_LARGE_GAP_THRESHOLD_PP:
            issue_marker = "REVIEW_HIGH"
        elif same_budget_delta <= -PAPER_LARGE_GAP_THRESHOLD_PP:
            issue_marker = "REVIEW_LOW"
        else:
            issue_marker = None
    return {
        "Task": cell["task_id"],
        "Method": cell["method"],
        "Model": cell.get("model"),
        "Reasoning": cell.get("reasoning_effort"),
        "Current budget (h)": float(cell["wall_time_seconds"]) / 3600.0,
        "T (s)": cell["wall_time_seconds"],
        "K": cell["live_search_concurrency"],
        "Outer trajectories": cell["completed_trajectories"],
        "Valid trajectories": cell["valid_trajectories"],
        "Best raw": best.get("raw_score"),
        "Current EdgeBench 0-100": normalized,
        "Local <=0.5h best": local_half_score,
        "Delta vs local <=0.5h (pp)": local_half_delta,
        "Local <=1h best": local_one_score,
        "Delta vs local <=1h (pp)": local_one_delta,
        "GPT-5.5 checkpoint (h)": checkpoint_hours,
        "GPT-5.5 same-budget": same_budget_score,
        "Delta vs same-budget (pp)": same_budget_delta,
        "Paper Codex + GPT-5.5 @12h mean": float(paper_score["mean"]),
        "Paper sample stddev": (
            float(paper_score["sample_stddev"])
            if paper_score.get("sample_stddev") is not None
            else None
        ),
        "Delta vs paper 12h (pp)": paper_delta,
        "Evaluator calls": evaluator_calls,
        "Runtime (s)": runtime,
        "Input tokens": input_tokens,
        "Output tokens": output_tokens,
        "Usage coverage": ", ".join(sorted(coverage)) or "unavailable",
        "Completion evidence": bool(completion.get("passed")),
        "Actual Goal Plus workers": worker_check.get("actual"),
        "Goal Plus candidates": (completion_checks.get("candidates") or {}).get(
            "actual"
        ),
        "Goal Plus verifier runs": (
            completion_checks.get("worker_verifier_runs") or {}
        ).get("actual"),
        "Incomplete reason": cell.get("incomplete_reason"),
        "Protocol classification": cell.get("protocol_classification"),
        "Official comparable": cell.get("official_edgebench_comparable", False),
        "Issue marker": issue_marker,
    }


def style_header(row: Iterable[Any]) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in row:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def write_comparison_workbook(payload: dict[str, Any], destination: Path) -> None:
    paper = payload["paper_reference"]
    paper_source = paper["source"]
    paper_contract = paper["reference"]
    paper_tasks = paper["tasks"]
    local_fast_reference = payload.get("local_fast_reference") or {}
    local_fast_contract = local_fast_reference.get("reference") or {}
    local_fast_checkpoints = local_fast_reference.get("checkpoints") or {}
    cells = payload["cells"]
    models = sorted({str(cell.get("model") or "unknown") for cell in cells})
    reasoning_levels = sorted(
        {str(cell.get("reasoning_effort") or "unspecified") for cell in cells}
    )
    wall_times = sorted({int(cell["wall_time_seconds"]) for cell in cells})
    valid_cells = sum(
        1 for cell in cells if int(cell.get("valid_trajectories") or 0) > 0
    )
    protocol_evidence = sum(
        1 for cell in cells if cell.get("protocol_classification") is not None
    )
    records = [
        comparison_record(cell, paper_tasks, local_fast_checkpoints) for cell in cells
    ]
    if not records:
        raise ValueError("comparison workbook requires at least one cell")
    workbook = Workbook()
    overview = workbook.active
    overview.title = "Overview"
    overview_rows = [
        ("Field", "Value"),
        ("Campaign", payload["campaign_id"]),
        ("Finalized at", payload.get("finalized_at")),
        ("Matched protocol", payload["matched_protocol"]),
        ("Models", ", ".join(models)),
        ("Reasoning", ", ".join(reasoning_levels)),
        ("Wall budgets (s)", ", ".join(str(value) for value in wall_times)),
        ("Cells with valid score", f"{valid_cells}/{len(cells)}"),
        ("Cells with protocol evidence", f"{protocol_evidence}/{len(cells)}"),
        ("EdgeBench commit", payload.get("edgebench_commit")),
        ("Goal Plus commit", payload.get("goal_plus_commit")),
        ("Dataset revision", payload.get("dataset_revision")),
        (
            "Same-budget reference",
            "Official EdgeBench GPT-5.5 checkpoint from each result",
        ),
        (
            "Issue rule",
            "Known protocol issue first; otherwise |same-budget delta| >= "
            f"{PAPER_LARGE_GAP_THRESHOLD_PP:g} pp",
        ),
        ("Local fast reference", local_fast_contract.get("label") or "Not included"),
        (
            "Local fast coverage",
            (
                "; ".join(
                    f"<={checkpoint.get('boundary_hours'):g}h: "
                    f"{checkpoint.get('available_count')}/"
                    f"{local_fast_reference.get('task_count')}"
                    for checkpoint in local_fast_checkpoints.values()
                )
                if local_fast_reference
                else "Not included"
            ),
        ),
        ("Local fast selection", local_fast_contract.get("selection") or "Not included"),
        (
            "Paper reference role",
            "12h diagnostic reference; not an apples-to-apples leaderboard comparison",
        ),
        ("Paper agent + model", f"{paper_contract['agent']} + {paper_contract['model']}"),
        ("Paper budget (h)", paper_contract["budget_hours"]),
        ("Paper scheduled runs", paper_contract["scheduled_runs"]),
        ("Paper arXiv", paper_source["arxiv_id"]),
        ("Paper TeX source", paper_source["source_file"]),
        ("Paper TeX SHA256", paper_source["source_file_sha256"]),
        (
            "Token note",
            "Zero with non-complete coverage means unavailable telemetry, not free usage",
        ),
    ]
    for row in overview_rows:
        overview.append(row)
    style_header(overview[1])
    overview.freeze_panes = "A2"
    overview.column_dimensions["A"].width = 30
    overview.column_dimensions["B"].width = 100
    for row in overview.iter_rows(min_row=2):
        row[0].font = Font(bold=True)
        row[1].alignment = Alignment(vertical="top", wrap_text=True)

    results = workbook.create_sheet("Results")
    headers = list(records[0])
    results.append(headers)
    for record in records:
        results.append([record[header] for header in headers])
    style_header(results[1])
    results.freeze_panes = "A2"
    table = Table(
        displayName="EdgeBenchResults",
        ref=f"A1:{get_column_letter(len(headers))}{len(records) + 1}",
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    results.add_table(table)
    widths = {
        "Task": 38,
        "Method": 18,
        "Model": 20,
        "Reasoning": 12,
        "Usage coverage": 20,
        "Protocol classification": 24,
        "Issue marker": 58,
    }
    for index, header in enumerate(headers, start=1):
        results.column_dimensions[get_column_letter(index)].width = widths.get(
            header, 18
        )
    numeric_headers = {
        "Best raw",
        "Current EdgeBench 0-100",
        "Local <=0.5h best",
        "Delta vs local <=0.5h (pp)",
        "Local <=1h best",
        "Delta vs local <=1h (pp)",
        "GPT-5.5 same-budget",
        "Delta vs same-budget (pp)",
        "Paper Codex + GPT-5.5 @12h mean",
        "Paper sample stddev",
        "Delta vs paper 12h (pp)",
        "Runtime (s)",
    }
    for header in numeric_headers:
        column = headers.index(header) + 1
        for row in range(2, len(records) + 2):
            results.cell(row=row, column=column).number_format = "0.00"
    issue_column = headers.index("Issue marker") + 1
    issue_letter = get_column_letter(issue_column)
    issue_range = f"{issue_letter}2:{issue_letter}{len(records) + 1}"
    for prefix, color in (
        ("REVIEW", "FFF2CC"),
        ("KNOWN", "F4CCCC"),
        ("MISSING", "F4CCCC"),
    ):
        results.conditional_formatting.add(
            issue_range,
            FormulaRule(
                formula=[f'LEFT({issue_letter}2,{len(prefix)})="{prefix}"'],
                fill=PatternFill("solid", fgColor=color),
            ),
        )

    if local_fast_reference:
        local_fast = workbook.create_sheet("Local Fast")
        local_fast_headers = [
            "Boundary <= (h)",
            "Task",
            "Status",
            "Selected checkpoint (h)",
            "Raw score",
            "EdgeBench 0-100",
            "Model",
            "Reasoning",
            "Campaign",
            "Protocol classification",
            "Normalization",
            "Best round",
            "Evidence source",
        ]
        local_fast.append(local_fast_headers)
        for checkpoint in local_fast_checkpoints.values():
            boundary = checkpoint.get("boundary_hours")
            for task_id, record in sorted((checkpoint.get("tasks") or {}).items()):
                local_fast.append(
                    [
                        boundary,
                        task_id,
                        "available",
                        record.get("checkpoint_hours"),
                        record.get("raw_score"),
                        record.get("edgebench_score"),
                        record.get("model"),
                        record.get("reasoning_effort"),
                        record.get("campaign_id"),
                        record.get("protocol_classification")
                        or "legacy development evidence",
                        record.get("normalization_source"),
                        record.get("best_round"),
                        record.get("source"),
                    ]
                )
            for task_id, attempts in sorted(
                (checkpoint.get("missing_tasks") or {}).items()
            ):
                reasons = sorted(
                    {
                        str(item.get("reason") or item.get("status") or "unavailable")
                        for item in attempts
                    }
                )
                local_fast.append(
                    [
                        boundary,
                        task_id,
                        "missing",
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        "; ".join(reasons),
                    ]
                )
        style_header(local_fast[1])
        local_fast.freeze_panes = "A2"
        local_fast.auto_filter.ref = f"A1:M{local_fast.max_row}"
        local_fast_widths = (16, 38, 12, 22, 18, 18, 20, 12, 65, 38, 24, 18, 100)
        for column, width in enumerate(local_fast_widths, start=1):
            local_fast.column_dimensions[get_column_letter(column)].width = width
        for row in local_fast.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    protocol = workbook.create_sheet("Protocol")
    protocol.append(
        [
            "Task",
            "Method",
            "Classification",
            "Official comparable",
            "Known protocol issue",
            "Protocol diff",
        ]
    )
    for cell in cells:
        protocol.append(
            [
                cell["task_id"],
                cell["method"],
                cell.get("protocol_classification"),
                cell.get("official_edgebench_comparable", False),
                cell.get("known_protocol_issue"),
                json.dumps(
                    cell.get("protocol_diff") or [],
                    ensure_ascii=False,
                    sort_keys=True,
                ),
            ]
        )
    style_header(protocol[1])
    protocol.freeze_panes = "A2"
    protocol.auto_filter.ref = f"A1:F{len(cells) + 1}"
    for column, width in enumerate((38, 18, 24, 20, 58, 100), start=1):
        protocol.column_dimensions[get_column_letter(column)].width = width
    for row in protocol.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)


def finalize_campaign(
    destination: Path, local_fast_reference_path: Path | None = None
) -> dict[str, Any]:
    campaign = io.read_json(destination / "campaign.json")
    paper_reference = load_paper_reference()
    missing_paper_tasks = sorted(
        set(campaign["task_ids"]) - set(paper_reference["tasks"])
    )
    if missing_paper_tasks:
        raise ValueError(
            "paper GPT-5.5 reference is missing campaign tasks: "
            + ", ".join(missing_paper_tasks)
        )
    summaries: list[dict[str, Any]] = []
    for item in campaign["cells"]:
        cell_path = destination / "cells" / item["cell_id"]
        cell = io.read_json(cell_path / "cell.json")
        summaries.append(summarize_cell(destination, cell))
    protocol_fields = {
        (
            summary["task_id"],
            summary["model"],
            summary["reasoning_effort"],
            summary["wall_time_seconds"],
            summary["live_search_concurrency"],
        )
        for summary in summaries
    }
    payload = {
        "schema_version": 1,
        "campaign_id": campaign["campaign_id"],
        "matched_protocol": len(protocol_fields) == len(set(campaign["task_ids"])),
        "edgebench_commit": campaign.get("edgebench_commit"),
        "goal_plus_commit": campaign.get("goal_plus_commit"),
        "dataset_revision": campaign.get("dataset_revision"),
        "wall_time_seconds": campaign.get("wall_time_seconds"),
        "live_search_concurrency": campaign.get("concurrency"),
        "cell_concurrency": campaign.get("cell_concurrency"),
        "paper_reference": paper_reference,
        "cells": summaries,
        "completion_evidence_passed": all(
            bool(summary["completion_evidence"]["passed"])
            for summary in summaries
        ),
        "finalized_at": io.utc_now(),
    }
    if not payload["completion_evidence_passed"]:
        incomplete = {
            summary["cell_id"]: summary["incomplete_reason"]
            for summary in summaries
            if not summary["completion_evidence"]["passed"]
        }
        for item in campaign["cells"]:
            if item["cell_id"] not in incomplete:
                continue
            item["state"] = "partial"
            item["incomplete_reason"] = incomplete[item["cell_id"]]
            cell_path = destination / "cells" / item["cell_id"] / "cell.json"
            cell = io.read_json(cell_path)
            cell["state"] = "partial"
            cell["incomplete_reason"] = incomplete[item["cell_id"]]
            io.write_json(cell_path, cell)
        campaign["state"] = "partial"
        campaign["completion_evidence_passed"] = False
        campaign["incomplete_cells"] = incomplete
        campaign["updated_at"] = io.utc_now()
    else:
        campaign["completion_evidence_passed"] = True
        campaign["updated_at"] = io.utc_now()
    io.write_json(destination / "campaign.json", campaign)
    if local_fast_reference_path is not None:
        payload["local_fast_reference"] = load_local_fast_reference(
            local_fast_reference_path
        )
    io.write_json(destination / "comparison.json", payload)
    write_comparison_workbook(payload, destination / f"{payload['campaign_id']}.xlsx")
    return payload
