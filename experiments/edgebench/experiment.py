#!/usr/bin/env python3
"""Provision, launch, monitor, stop, and summarize EdgeBench campaigns."""

from __future__ import annotations

import argparse
import atexit
import hashlib
import json
import os
import re
import signal
import socket
import subprocess
import sys
import tarfile
import time
import urllib.error
import urllib.parse
import urllib.request
from collections import defaultdict, deque
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

import yaml
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))

from bench_runtime_paths import configure_temp_environment, ensure_temp_root  # noqa: E402


EDGE_ROOT = ROOT / "third_party" / "edgebench"
GOAL_PLUS_ROOT = ROOT / "third_party" / "goal-plus"
TASKS_DIR = EDGE_ROOT / "tasks"
PROFILE_DIR = ROOT / "experiments" / "edgebench" / "profiles"
OFFICIAL_CODEX_PROTOCOL_PATH = (
    EDGE_ROOT / "examples" / "all-tasks-k8s" / "experiment-codex.yaml"
)
PAPER_REFERENCE_PATH = (
    ROOT
    / "experiments"
    / "edgebench"
    / "references"
    / "paper-gpt-5.5-codex-12h.json"
)
RUNS_ROOT = ROOT / "runs" / "edgebench"
UPSTREAM_MANIFEST = ROOT / "environment" / "upstreams.json"
EVIDENCE_ANNOTATOR_PROVIDER_ID = "edgebench-evidence"
VENV = ROOT / ".bench-env" / "venv"
VENV_BIN = VENV / ("Scripts" if sys.platform == "win32" else "bin")
VENV_PYTHON = VENV_BIN / ("python.exe" if sys.platform == "win32" else "python")
SFORGE = VENV_BIN / ("sforge.exe" if sys.platform == "win32" else "sforge")

METHODS = {
    "plain-codex": {
        "agent": "codex",
        "outer_replicas": "concurrency",
        "inner_search": False,
        "api_protocol": "openai",
    },
    "goal-plus-codex": {
        "agent": "codex-goal-plus",
        "outer_replicas": 1,
        "inner_search": True,
        "api_protocol": "openai",
    },
    "plain-claude": {
        "agent": "claude-code",
        "outer_replicas": "concurrency",
        "inner_search": False,
        "api_protocol": "anthropic",
    },
}


def api_protocol_for_methods(methods: Iterable[str]) -> str:
    protocols = {str(METHODS[method]["api_protocol"]) for method in methods}
    if len(protocols) != 1:
        raise ValueError(
            "one EdgeBench campaign cannot mix agent API protocols: "
            + ", ".join(sorted(protocols))
        )
    return next(iter(protocols))


CLAUDE_REASONING_EFFORTS = frozenset(
    {"none", "minimal", "low", "medium", "high", "xhigh", "max"}
)


def validate_claude_thinking_contract(
    thinking: Any, reasoning_effort: Any
) -> None:
    if thinking == {"type": "adaptive"}:
        if reasoning_effort is not None:
            raise ValueError(
                "adaptive Claude EdgeBench profiles must not set reasoning effort"
            )
        return
    effort = str(reasoning_effort or "")
    if effort not in CLAUDE_REASONING_EFFORTS:
        raise ValueError(
            "Claude EdgeBench profiles must use adaptive thinking without effort "
            "or pin a supported reasoning effort"
        )
    expected_type = "disabled" if effort in {"none", "minimal"} else "enabled"
    if thinking != {"type": expected_type}:
        raise ValueError(
            "Claude EdgeBench profiles must pair "
            f"reasoning_effort={effort!r} with thinking.type={expected_type!r}"
        )

PAPER_LARGE_GAP_THRESHOLD_PP = 20.0
LEGACY_PAPER_PROTOCOL_ISSUES = {
    "borden_source_inversion": "no cooldown; unusually high evaluator-call frequency",
    "exchange_core_throughput": "Internet access and unbounded CPU/hardware-sensitive score",
    "schemathesis_config_modernization": "Internet access used by the agent; no official cooldown",
    "schemathesis_datagen_pipeline": "Internet access used by the agent; no official cooldown",
    "schemathesis_reporting_observability": "Internet access used by the agent; no official cooldown",
}
OFFICIAL_PROTOCOL_FIELDS = frozenset(
    {
        "agent",
        "backend",
        "disable_auto_eval",
        "disable_auto_resume",
        "disable_stop_hook",
        "eval_interval",
        "judge_cpu_limit",
        "judge_mem_limit",
        "max_submissions",
        "submission_cooldown",
        "timeout",
        "work_cpu_limit",
        "work_mem_limit",
    }
)
OFFICIAL_REQUIRED_DEFAULTS = frozenset(
    {
        "agent",
        "backend",
        "eval_interval",
        "judge_cpu_limit",
        "judge_mem_limit",
        "submission_cooldown",
        "timeout",
        "work_cpu_limit",
        "work_mem_limit",
    }
)
ALLOWED_PROTOCOL_OVERRIDE_FIELDS = frozenset(
    {
        "agent",
        "attempts_per_task",
        "backend",
        "cell_concurrency",
        "judge_concurrency",
        "model",
        "reasoning_effort",
        "timeout",
    }
)
PROFILE_PROTOCOL_OVERRIDE_FIELDS = frozenset({"eval_interval", "internet"})
OFFICIAL_TASK_COUNT = 51
OFFICIAL_SCHEDULED_RUNS = 3


def utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def campaign_stamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def load_paper_reference(path: Path = PAPER_REFERENCE_PATH) -> dict[str, Any]:
    payload = read_json(path)
    reference = payload.get("reference", {})
    tasks = payload.get("tasks", {})
    if (
        payload.get("schema_version") != 1
        or reference.get("agent") != "Codex"
        or reference.get("model") != "GPT-5.5"
        or reference.get("budget_hours") != 12
        or not isinstance(tasks, dict)
        or not tasks
    ):
        raise ValueError(f"invalid EdgeBench paper reference: {path}")
    for task, score in tasks.items():
        if not isinstance(score, dict) or not isinstance(score.get("mean"), (int, float)):
            raise ValueError(f"invalid paper score for {task}: {path}")
    return payload


def load_local_fast_reference(path: Path) -> dict[str, Any]:
    payload = read_json(path)
    reference = payload.get("reference", {})
    checkpoints = payload.get("checkpoints", {})
    if (
        payload.get("schema_version") != 2
        or reference.get("official_comparison") is not False
        or not isinstance(checkpoints, dict)
        or not checkpoints
    ):
        raise ValueError(f"invalid EdgeBench local fast reference: {path}")
    for label, checkpoint in checkpoints.items():
        tasks = checkpoint.get("tasks", {}) if isinstance(checkpoint, dict) else {}
        boundary = checkpoint.get("boundary_seconds") if isinstance(checkpoint, dict) else None
        if not isinstance(boundary, int) or not isinstance(tasks, dict):
            raise ValueError(f"invalid local fast checkpoint {label}: {path}")
        for task_id, record in tasks.items():
            if (
                not isinstance(record, dict)
                or record.get("task_id") != task_id
                or not isinstance(record.get("edgebench_score"), (int, float))
                or not 0 < int(record.get("checkpoint_seconds") or 0) <= boundary
            ):
                raise ValueError(
                    f"invalid local fast score for {task_id} at {label}: {path}"
                )
    return payload


def sha256_text(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def git_head(path: Path) -> str | None:
    completed = subprocess.run(
        ["git", "-C", str(path), "rev-parse", "HEAD"],
        capture_output=True,
        text=True,
        check=False,
    )
    return completed.stdout.strip() if completed.returncode == 0 else None


def git_branch(path: Path) -> str | None:
    completed = subprocess.run(
        ["git", "-C", str(path), "symbolic-ref", "--short", "-q", "HEAD"],
        capture_output=True,
        text=True,
        check=False,
    )
    return completed.stdout.strip() if completed.returncode == 0 else None


def git_dirty(path: Path) -> bool | None:
    completed = subprocess.run(
        ["git", "-C", str(path), "status", "--porcelain"],
        capture_output=True,
        text=True,
        check=False,
    )
    return bool(completed.stdout.strip()) if completed.returncode == 0 else None


def upstream_entry(name: str) -> dict[str, Any]:
    manifest = read_json(UPSTREAM_MANIFEST)
    return dict(manifest["upstreams"][name])


def load_profile(value: str | Path) -> tuple[Path, dict[str, Any]]:
    candidate = Path(value)
    if not candidate.suffix:
        candidate = PROFILE_DIR / f"{candidate.name}.json"
    elif not candidate.is_absolute():
        candidate = (ROOT / candidate).resolve()
    if not candidate.is_file():
        raise FileNotFoundError(f"EdgeBench profile not found: {candidate}")
    profile = read_json(candidate)
    if profile.get("schema_version") != 1:
        raise ValueError("unsupported EdgeBench profile schema")
    for key in (
        "id",
        "dataset_repository",
        "dataset_revision",
        "task_ids",
        "methods",
        "model",
        "wall_time_seconds",
        "concurrency",
    ):
        if key not in profile:
            raise ValueError(f"EdgeBench profile is missing {key!r}")
    unknown = set(profile["methods"]) - set(METHODS)
    if unknown:
        raise ValueError("unknown EdgeBench method(s): " + ", ".join(sorted(unknown)))
    api_protocol = api_protocol_for_methods(profile["methods"])
    if api_protocol == "anthropic":
        validate_claude_thinking_contract(
            profile.get("thinking"), profile.get("reasoning_effort")
        )
        context_window = profile.get("claude_context_window_tokens")
        compact_percent = profile.get("claude_autocompact_percent")
        if (context_window is None) != (compact_percent is None):
            raise ValueError(
                "Claude context window and autocompact percent must be set together"
            )
        if context_window is not None and (
            not isinstance(context_window, int)
            or isinstance(context_window, bool)
            or context_window < 1
        ):
            raise ValueError("claude_context_window_tokens must be positive")
        if compact_percent is not None and (
            not isinstance(compact_percent, int)
            or isinstance(compact_percent, bool)
            or not 1 <= compact_percent <= 100
        ):
            raise ValueError("claude_autocompact_percent must be between 1 and 100")
    if int(profile["wall_time_seconds"]) < 1 or int(profile["concurrency"]) < 1:
        raise ValueError("wall_time_seconds and concurrency must be positive")
    if int(profile.get("cell_concurrency", 1)) < 1:
        raise ValueError("cell_concurrency must be positive")
    if profile.get("protocol_source") != "edgebench-official-codex":
        raise ValueError("EdgeBench profile must use edgebench-official-codex")
    reasons = profile.get("protocol_override_reasons")
    if not isinstance(reasons, dict) or not reasons:
        raise ValueError("EdgeBench profile must record protocol_override_reasons")
    protocol_overrides = profile.get("protocol_overrides", {})
    if not isinstance(protocol_overrides, dict):
        raise ValueError("EdgeBench profile protocol_overrides must be an object")
    unknown_overrides = set(protocol_overrides) - PROFILE_PROTOCOL_OVERRIDE_FIELDS
    if unknown_overrides:
        raise ValueError(
            "EdgeBench profile has unsupported protocol overrides: "
            f"{sorted(unknown_overrides)}"
        )
    if "internet" in protocol_overrides and not isinstance(
        protocol_overrides["internet"], bool
    ):
        raise ValueError("EdgeBench profile internet override must be boolean")
    if "eval_interval" in protocol_overrides and (
        not isinstance(protocol_overrides["eval_interval"], int)
        or isinstance(protocol_overrides["eval_interval"], bool)
        or protocol_overrides["eval_interval"] < 1
    ):
        raise ValueError(
            "EdgeBench profile eval_interval override must be a positive integer"
        )
    missing_override_reasons = set(protocol_overrides) - set(reasons)
    if missing_override_reasons:
        raise ValueError(
            "EdgeBench profile protocol overrides are missing reasons: "
            f"{sorted(missing_override_reasons)}"
        )
    unknown_reasons = set(reasons) - (
        ALLOWED_PROTOCOL_OVERRIDE_FIELDS | PROFILE_PROTOCOL_OVERRIDE_FIELDS
    )
    if unknown_reasons:
        raise ValueError(
            "EdgeBench profile has unsupported protocol override reasons: "
            f"{sorted(unknown_reasons)}"
        )
    invalid_reasons = sorted(
        key
        for key, reason in reasons.items()
        if not isinstance(reason, str) or not reason.strip()
    )
    if invalid_reasons:
        raise ValueError(
            "EdgeBench profile has invalid protocol override reasons: "
            f"{invalid_reasons}"
        )
    return candidate, profile


def _normalize_protocol_fields(data: Any, *, context: str) -> dict[str, Any]:
    if data is None:
        return {}
    if not isinstance(data, dict):
        raise ValueError(f"{context} must be a mapping")
    unknown = set(data) - OFFICIAL_PROTOCOL_FIELDS
    if unknown:
        raise ValueError(f"{context} has unsupported fields: {sorted(unknown)}")
    result: dict[str, Any] = {}
    for key, value in data.items():
        if key in {
            "disable_auto_eval",
            "disable_auto_resume",
            "disable_stop_hook",
        }:
            if not isinstance(value, bool):
                raise ValueError(f"{context}.{key} must be boolean")
        elif key in {
            "eval_interval",
            "judge_cpu_limit",
            "max_submissions",
            "submission_cooldown",
            "timeout",
            "work_cpu_limit",
        }:
            if not isinstance(value, int) or isinstance(value, bool) or value < 0:
                raise ValueError(f"{context}.{key} must be a non-negative integer")
        elif key in {"judge_mem_limit", "work_mem_limit", "agent", "backend"}:
            if not isinstance(value, str) or not value:
                raise ValueError(f"{context}.{key} must be a non-empty string")
        result[key] = value
    return result


def load_official_codex_protocol(
    path: Path = OFFICIAL_CODEX_PROTOCOL_PATH,
) -> dict[str, Any]:
    raw = yaml.safe_load(path.read_text(encoding="utf-8"))
    if not isinstance(raw, dict):
        raise ValueError(f"official EdgeBench protocol must be a mapping: {path}")
    allowed_top = {"defaults", "env", "model", "stagger", "tasks"}
    unknown_top = {
        key for key in raw if key not in allowed_top and not str(key).startswith("x-")
    }
    if unknown_top:
        raise ValueError(
            f"official EdgeBench protocol has unsupported top-level fields: "
            f"{sorted(unknown_top)}"
        )
    defaults = _normalize_protocol_fields(
        raw.get("defaults"), context="official defaults"
    )
    missing_defaults = OFFICIAL_REQUIRED_DEFAULTS - set(defaults)
    if missing_defaults:
        raise ValueError(
            f"official EdgeBench defaults are missing: {sorted(missing_defaults)}"
        )
    tasks_raw = raw.get("tasks")
    if not isinstance(tasks_raw, dict) or not tasks_raw:
        raise ValueError("official EdgeBench protocol must define tasks")
    tasks = {
        str(task_id): _normalize_protocol_fields(
            overrides, context=f"official task {task_id}"
        )
        for task_id, overrides in tasks_raw.items()
    }
    if len(tasks) != OFFICIAL_TASK_COUNT:
        raise ValueError(
            "official EdgeBench protocol must define exactly "
            f"{OFFICIAL_TASK_COUNT} tasks, found {len(tasks)}"
        )
    model_raw = raw.get("model")
    if not isinstance(model_raw, dict) or not isinstance(model_raw.get("model"), str):
        raise ValueError("official EdgeBench protocol must define model.model")
    stagger = raw.get("stagger", 0)
    if not isinstance(stagger, int) or isinstance(stagger, bool) or stagger < 0:
        raise ValueError("official EdgeBench protocol stagger must be a non-negative integer")
    return {
        "schema_version": 1,
        "source": portable_path(path),
        "source_sha256": sha256_file(path),
        "official_model": str(model_raw["model"]),
        "stagger_seconds": stagger,
        "defaults": defaults,
        "tasks": tasks,
    }


def official_task_protocol(
    protocol: dict[str, Any], task_id: str, config: dict[str, Any]
) -> dict[str, Any]:
    if task_id not in protocol["tasks"]:
        raise ValueError(f"official EdgeBench protocol is missing task {task_id}")
    internet = config.get("internet")
    if not isinstance(internet, bool):
        raise ValueError(f"task {task_id} must define boolean internet")
    resolved = {**protocol["defaults"], **protocol["tasks"][task_id]}
    resolved.setdefault("disable_auto_eval", False)
    resolved.setdefault("disable_auto_resume", False)
    resolved.setdefault("disable_stop_hook", False)
    resolved.setdefault("max_submissions", None)
    resolved["internet"] = internet
    return resolved


def profile_task_protocol(
    profile: dict[str, Any],
    protocol: dict[str, Any],
    task_id: str,
    config: dict[str, Any],
) -> dict[str, Any]:
    resolved = official_task_protocol(protocol, task_id, config)
    return {
        **resolved,
        **dict(profile.get("protocol_overrides") or {}),
    }


def _protocol_diff(
    *,
    official: dict[str, Any],
    effective: dict[str, Any],
    reasons: dict[str, Any],
    allowed_fields: frozenset[str] | set[str] | None = None,
) -> list[dict[str, Any]]:
    permitted = (
        ALLOWED_PROTOCOL_OVERRIDE_FIELDS
        if allowed_fields is None
        else frozenset(allowed_fields)
    )
    fields = sorted(set(official) | set(effective))
    result: list[dict[str, Any]] = []
    for field in fields:
        before = official.get(field)
        after = effective.get(field)
        if before == after:
            continue
        if field not in permitted:
            raise ValueError(f"unsupported EdgeBench protocol override: {field}")
        reason = reasons.get(field)
        if not isinstance(reason, str) or not reason.strip():
            raise ValueError(f"protocol override {field!r} is missing a reason")
        result.append(
            {
                "field": field,
                "official": before,
                "effective": after,
                "reason": reason,
            }
        )
    return result


def campaign_dir(value: str | Path) -> Path:
    candidate = Path(value).expanduser()
    if not candidate.is_absolute():
        direct = (ROOT / candidate).resolve()
        candidate = direct if direct.is_dir() else (RUNS_ROOT / candidate).resolve()
    else:
        candidate = candidate.resolve()
    try:
        candidate.relative_to(RUNS_ROOT.resolve())
    except ValueError as exc:
        raise ValueError(f"campaign must be under {RUNS_ROOT}") from exc
    if not (candidate / "campaign.json").is_file():
        raise FileNotFoundError(f"campaign.json not found in {candidate}")
    return candidate


def portable_path(path: Path) -> str:
    resolved = path.resolve()
    try:
        return str(resolved.relative_to(ROOT))
    except ValueError:
        return resolved.name


def portable_command(command: Iterable[str]) -> list[str]:
    replacements = (
        (str(ROOT.resolve()), "<bench-goal-plus>"),
        (str(Path.home().resolve()), "<home>"),
    )
    result: list[str] = []
    for argument in command:
        clean = str(argument)
        for source, replacement in replacements:
            clean = clean.replace(source, replacement)
        result.append(clean)
    return result


def run_capture(command: list[str], *, env: dict[str, str] | None = None) -> dict[str, Any]:
    try:
        completed = subprocess.run(
            command,
            cwd=ROOT,
            env=env,
            capture_output=True,
            text=True,
            check=False,
        )
    except FileNotFoundError as exc:
        return {"returncode": 127, "stdout": "", "stderr": str(exc)}
    return {
        "returncode": completed.returncode,
        "stdout": completed.stdout.strip(),
        "stderr": completed.stderr.strip(),
    }


def resolve_agent_api_config(
    env: dict[str, str] | None = None,
    *,
    protocol: str = "openai",
) -> dict[str, str | None]:
    source = os.environ if env is None else env

    def first(names: tuple[str, ...]) -> tuple[str | None, str | None]:
        for name in names:
            value = source.get(name)
            if value:
                return value, name
        return None, None

    if protocol == "anthropic":
        key_names = (
            "SFORGE_AGENT_API_KEY",
            "ANTHROPIC_AUTH_TOKEN",
            "ANTHROPIC_API_KEY",
        )
        base_names = ("SFORGE_AGENT_API_BASE_URL", "ANTHROPIC_BASE_URL")
    elif protocol == "openai":
        key_names = (
            "SFORGE_AGENT_API_KEY",
            "OPENAI_API_KEY",
            "CODEX_API_KEY",
        )
        base_names = ("SFORGE_AGENT_API_BASE_URL", "OPENAI_BASE_URL")
    else:
        raise ValueError(f"unsupported agent API protocol: {protocol!r}")
    api_key, key_source = first(key_names)
    base_url, base_source = first(base_names)
    return {
        "api_key": api_key,
        "api_key_source": key_source,
        "api_base_url": base_url,
        "api_base_url_source": base_source,
    }


def loopback_api_target(base_url: str) -> tuple[str, int] | None:
    parsed = urllib.parse.urlsplit(base_url)
    if parsed.scheme not in {"http", "https"} or not parsed.hostname:
        raise ValueError(f"invalid agent API base URL: {base_url!r}")
    if parsed.hostname not in {"127.0.0.1", "localhost", "::1"}:
        return None
    port = parsed.port or (443 if parsed.scheme == "https" else 80)
    return parsed.hostname, port


def bridged_base_url(base_url: str, host: str, port: int) -> str:
    parsed = urllib.parse.urlsplit(base_url)
    netloc = f"[{host}]:{port}" if ":" in host else f"{host}:{port}"
    return urllib.parse.urlunsplit(
        (parsed.scheme, netloc, parsed.path, parsed.query, parsed.fragment)
    )


def default_route_ipv4() -> str:
    route = run_capture(["ip", "-j", "-4", "route", "get", "1.1.1.1"])
    if route["returncode"] == 0:
        try:
            payload = json.loads(route["stdout"])
            if payload and payload[0].get("prefsrc"):
                return str(payload[0]["prefsrc"])
        except (json.JSONDecodeError, TypeError, KeyError):
            pass
    raise RuntimeError("could not determine the host default-route IPv4 address")


def reserve_tcp_port(host: str) -> int:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as listener:
        listener.bind((host, 0))
        return int(listener.getsockname()[1])


def start_socket_bridge(
    destination: Path,
    *,
    name: str,
    listen_host: str,
    target_host: str,
    target_port: int,
) -> tuple[subprocess.Popen[str], dict[str, Any], Any]:
    socket_activate = Path("/usr/bin/systemd-socket-activate")
    socket_proxyd = Path("/lib/systemd/systemd-socket-proxyd")
    if not socket_activate.is_file() or not socket_proxyd.is_file():
        raise RuntimeError(
            "rootless Docker loopback bridging requires systemd-socket-activate "
            "and systemd-socket-proxyd"
        )

    listen_port = reserve_tcp_port(listen_host)
    target = (
        f"[{target_host}]:{target_port}"
        if ":" in target_host
        else f"{target_host}:{target_port}"
    )
    command = [
        str(socket_activate),
        f"--listen={listen_host}:{listen_port}",
        str(socket_proxyd),
        target,
    ]
    log_path = destination / "bridges" / f"{name}.log"
    log_path.parent.mkdir(parents=True, exist_ok=True)
    log = log_path.open("a", encoding="utf-8")
    process = subprocess.Popen(
        command,
        cwd=ROOT,
        env=dict(configure_temp_environment(dict(os.environ))),
        stdout=log,
        stderr=subprocess.STDOUT,
        text=True,
        start_new_session=True,
    )
    log.close()
    closed = False

    def close_bridge() -> None:
        nonlocal closed
        if closed:
            return
        closed = True
        if process.poll() is not None:
            return
        try:
            os.killpg(process.pid, signal.SIGTERM)
            process.wait(timeout=5)
        except (ProcessLookupError, subprocess.TimeoutExpired):
            if process.poll() is None:
                os.killpg(process.pid, signal.SIGKILL)
                process.wait(timeout=5)

    atexit.register(close_bridge)
    deadline = time.monotonic() + 10
    while time.monotonic() < deadline:
        if process.poll() is not None:
            break
        try:
            with socket.create_connection((listen_host, listen_port), timeout=0.25):
                metadata = {
                    "name": name,
                    "listen_host": listen_host,
                    "listen_port": listen_port,
                    "target_host": target_host,
                    "target_port": target_port,
                    "pid": process.pid,
                    "log": portable_path(log_path),
                }
                return process, metadata, close_bridge
        except OSError:
            time.sleep(0.1)
    close_bridge()
    raise RuntimeError(
        f"{name} bridge did not become ready; inspect {portable_path(log_path)}"
    )


def agent_api_probe_url(base_url: str, protocol: str) -> str:
    base = base_url.rstrip("/")
    if protocol == "anthropic":
        return base + ("/messages" if base.endswith("/v1") else "/v1/messages")
    if protocol == "openai":
        return base + "/models"
    raise ValueError(f"unsupported agent API protocol: {protocol!r}")


def authenticated_api_probe(
    base_url: str,
    api_key: str,
    *,
    protocol: str = "openai",
    model: str | None = None,
    thinking: dict[str, str] | None = None,
    reasoning_effort: str | None = None,
) -> dict[str, Any]:
    url = agent_api_probe_url(base_url, protocol)
    headers = {"Authorization": f"Bearer {api_key}"}
    data: bytes | None = None
    if protocol == "anthropic":
        if not model:
            raise ValueError("Anthropic API probes require a model")
        headers.update(
            {
                "anthropic-version": "2023-06-01",
                "content-type": "application/json",
            }
        )
        payload: dict[str, Any] = {
            "model": model,
            "max_tokens": 1,
            "messages": [{"role": "user", "content": "Reply OK."}],
        }
        if thinking is not None:
            payload["thinking"] = thinking
        if reasoning_effort is not None:
            payload["reasoning_effort"] = reasoning_effort
        data = json.dumps(payload).encode("utf-8")
    request = urllib.request.Request(
        url,
        headers=headers,
        data=data,
    )
    opener = urllib.request.build_opener(urllib.request.ProxyHandler({}))
    try:
        with opener.open(request, timeout=10) as response:
            return {"passed": response.status == 200, "status": response.status}
    except urllib.error.HTTPError as exc:
        return {"passed": False, "status": exc.code, "error": str(exc)}
    except (OSError, urllib.error.URLError) as exc:
        return {"passed": False, "status": None, "error": str(exc)}


def append_no_proxy(env: dict[str, str], host: str) -> None:
    current = env.get("NO_PROXY") or env.get("no_proxy") or ""
    entries = [item.strip() for item in current.split(",") if item.strip()]
    if host not in entries:
        entries.insert(0, host)
    value = ",".join(entries)
    env["NO_PROXY"] = value
    env["no_proxy"] = value
    env["SFORGE_NO_PROXY"] = value


def judge_server_environment(
    *,
    api_key: str | None,
    api_base_url: str | None,
    bridge_host: str | None,
    model: str = "gpt-5.5",
) -> dict[str, str]:
    env = dict(os.environ)
    configure_temp_environment(env)
    entries: dict[str, str] = {}
    for item in env.get("SFORGE_JUDGE_EXTRA_ENV", "").split(","):
        if "=" in item:
            key, value = item.split("=", 1)
            entries[key.strip()] = value.strip()
    if api_key:
        entries.setdefault("SFORGE_JUDGE_API_KEY", api_key)
    if api_base_url:
        entries.setdefault("SFORGE_JUDGE_API_BASE_URL", api_base_url)
    entries.setdefault("SFORGE_JUDGE_MODEL", model)
    if entries:
        env["SFORGE_JUDGE_EXTRA_ENV"] = ",".join(
            f"{key}={value}" for key, value in sorted(entries.items())
        )
    if bridge_host:
        append_no_proxy(env, bridge_host)
    return env


def docker_http_probe(
    image: str,
    url: str,
    *,
    api_key: str | None = None,
    protocol: str | None = None,
    model: str | None = None,
    thinking_type: str | None = None,
    reasoning_effort: str | None = None,
) -> dict[str, Any]:
    env = dict(os.environ)
    configure_temp_environment(env)
    env["SFORGE_PROBE_URL"] = (
        agent_api_probe_url(url, protocol) if protocol else url
    )
    command = [
        "docker",
        "run",
        "--rm",
        "--entrypoint",
        "/bin/sh",
        "-e",
        "SFORGE_PROBE_URL",
    ]
    if api_key:
        env["SFORGE_PROBE_API_KEY"] = api_key
        command.extend(["-e", "SFORGE_PROBE_API_KEY"])
    if protocol:
        env["SFORGE_PROBE_PROTOCOL"] = protocol
        command.extend(["-e", "SFORGE_PROBE_PROTOCOL"])
    if model:
        env["SFORGE_PROBE_MODEL"] = model
        command.extend(["-e", "SFORGE_PROBE_MODEL"])
    if thinking_type:
        env["SFORGE_PROBE_THINKING_TYPE"] = thinking_type
        command.extend(["-e", "SFORGE_PROBE_THINKING_TYPE"])
    if reasoning_effort:
        env["SFORGE_PROBE_REASONING_EFFORT"] = reasoning_effort
        command.extend(["-e", "SFORGE_PROBE_REASONING_EFFORT"])
    command.extend(
        [
            image,
            "-c",
            (
                "if [ \"${SFORGE_PROBE_PROTOCOL:-}\" = anthropic ]; then "
                "if [ -n \"${SFORGE_PROBE_REASONING_EFFORT:-}\" ]; then "
                "payload='{\"model\":\"'\"$SFORGE_PROBE_MODEL\"'\","
                "\"max_tokens\":1,\"messages\":[{\"role\":\"user\","
                "\"content\":\"Reply OK.\"}],\"thinking\":{\"type\":\"'"
                "\"$SFORGE_PROBE_THINKING_TYPE\"'\"},\"reasoning_effort\":\"'"
                "\"$SFORGE_PROBE_REASONING_EFFORT\"'\"}'; "
                "else payload='{\"model\":\"'\"$SFORGE_PROBE_MODEL\"'\","
                "\"max_tokens\":1,\"messages\":[{\"role\":\"user\","
                "\"content\":\"Reply OK.\"}],\"thinking\":{\"type\":\"'"
                "\"$SFORGE_PROBE_THINKING_TYPE\"'\"}}'; fi; "
                "auth=\"Authorization: Bearer $SFORGE_PROBE_API_KEY\"; "
                "code=$(curl --noproxy '*' -sS -o /dev/null -w '%{http_code}' "
                "--max-time 30 -X POST -H \"$auth\" "
                "-H 'anthropic-version: 2023-06-01' "
                "-H 'content-type: application/json' "
                "--data \"$payload\" \"$SFORGE_PROBE_URL\"); "
                "elif [ -n \"${SFORGE_PROBE_API_KEY:-}\" ]; then "
                "auth=\"Authorization: Bearer $SFORGE_PROBE_API_KEY\"; "
                "code=$(curl --noproxy '*' -sS -o /dev/null -w '%{http_code}' "
                "--max-time 15 -H \"$auth\" \"$SFORGE_PROBE_URL\"); "
                "else code=$(curl --noproxy '*' -sS -o /dev/null -w '%{http_code}' "
                "--max-time 15 \"$SFORGE_PROBE_URL\"); fi; "
                "printf '%s\\n' \"$code\"; test \"$code\" = 200"
            ),
        ]
    )
    result = run_capture(command, env=env)
    return {
        "passed": result["returncode"] == 0,
        "status": result["stdout"].splitlines()[-1] if result["stdout"] else None,
        "stderr": result["stderr"][-400:] or None,
    }


def _docker_memory_bytes(value: str) -> int:
    match = re.fullmatch(r"([0-9]+)([kmgt]?)(?:i?b)?", value.strip().lower())
    if not match:
        raise ValueError(f"unsupported Docker memory limit: {value}")
    amount = int(match.group(1))
    exponent = {"": 0, "k": 1, "m": 2, "g": 3, "t": 4}[match.group(2)]
    return amount * (1024**exponent)


def docker_resource_limit_probe(
    image: str, *, cpu_limit: int, mem_limit: str
) -> dict[str, Any]:
    name = f"edgebench-resource-probe-{os.getpid()}-{time.time_ns()}"
    expected_nano_cpus = int(cpu_limit * 1_000_000_000)
    expected_memory = _docker_memory_bytes(mem_limit)
    started = run_capture(
        [
            "docker",
            "run",
            "--detach",
            "--name",
            name,
            "--cpus",
            str(cpu_limit),
            "--memory",
            mem_limit,
            "--entrypoint",
            "/bin/sh",
            image,
            "-c",
            "while :; do sleep 60; done",
        ]
    )
    inspected: dict[str, Any] | None = None
    inspect_result: dict[str, Any] | None = None
    try:
        if started["returncode"] == 0:
            inspect_result = run_capture(
                [
                    "docker",
                    "inspect",
                    "--format",
                    "{{json .HostConfig}}",
                    name,
                ]
            )
            if inspect_result["returncode"] == 0:
                try:
                    inspected = json.loads(inspect_result["stdout"])
                except json.JSONDecodeError:
                    inspected = None
    finally:
        run_capture(["docker", "rm", "--force", name])

    actual_nano_cpus = inspected.get("NanoCpus") if inspected else None
    actual_memory = inspected.get("Memory") if inspected else None
    return {
        "passed": (
            started["returncode"] == 0
            and inspect_result is not None
            and inspect_result["returncode"] == 0
            and actual_nano_cpus == expected_nano_cpus
            and actual_memory == expected_memory
        ),
        "image": image,
        "cpu_limit": cpu_limit,
        "mem_limit": mem_limit,
        "expected_nano_cpus": expected_nano_cpus,
        "actual_nano_cpus": actual_nano_cpus,
        "expected_memory_bytes": expected_memory,
        "actual_memory_bytes": actual_memory,
        "stderr": (
            started["stderr"][-400:]
            or ((inspect_result or {}).get("stderr") or "")[-400:]
            or None
        ),
    }


def sforge_iptables_permission_probe() -> dict[str, Any]:
    if not VENV_PYTHON.is_file():
        return {"passed": False, "stderr": "benchmark virtualenv is missing"}
    result = run_capture(
        [
            str(VENV_PYTHON),
            "-c",
            (
                "from sforge.harness.network_isolation import "
                "check_iptables_permission; "
                "raise SystemExit(0 if check_iptables_permission() else 1)"
            ),
        ]
    )
    return {
        "passed": result["returncode"] == 0,
        "stderr": result["stderr"][-400:] or None,
    }


def task_config(task_id: str) -> dict[str, Any]:
    path = TASKS_DIR / f"{task_id}.json"
    if not path.is_file():
        raise FileNotFoundError(
            f"task definition missing: {path}; run provision first"
        )
    return read_json(path)


def task_images(task_id: str) -> tuple[str, str]:
    config = task_config(task_id)
    return (
        f"edgebench.work.{task_id}:{config['work']['image_tag']}",
        f"edgebench.judge.{task_id}:{config['judge']['image_tag']}",
    )


def rust_runtime_asset() -> dict[str, str]:
    path = EDGE_ROOT / "sforge" / "harness" / "runtime_assets.json"
    payload = read_json(path)
    asset = payload.get("rust") if payload.get("schema_version") == 1 else None
    required = {"version", "target", "archive_name", "url", "sha256"}
    if not isinstance(asset, dict) or required - set(asset):
        raise RuntimeError(f"invalid EdgeBench runtime asset manifest: {path}")
    return {key: str(asset[key]) for key in required}


def rust_runtime_archive_status() -> dict[str, Any]:
    try:
        asset = rust_runtime_asset()
        archive = (
            Path.home()
            / ".cache"
            / "sforge"
            / "rust"
            / asset["archive_name"]
        )
        actual_sha256 = sha256_file(archive) if archive.is_file() else None
        return {
            "passed": actual_sha256 == asset["sha256"],
            "path": str(archive),
            "version": asset["version"],
            "expected_sha256": asset["sha256"],
            "actual_sha256": actual_sha256,
        }
    except (OSError, RuntimeError, json.JSONDecodeError) as exc:
        return {"passed": False, "error": str(exc)}


def rust_image_runtime_probe(image: str, version: str) -> dict[str, Any]:
    command = (
        "set -e; command -v cargo; command -v rustc; "
        f"cargo --version | grep -F 'cargo {version} '; "
        f"rustc --version | grep -F 'rustc {version} '"
    )
    return run_capture(
        [
            "docker",
            "run",
            "--rm",
            "--entrypoint",
            "/bin/bash",
            image,
            "-c",
            command,
        ]
    )


def dataset_revision(task_id: str) -> str | None:
    metadata = TASKS_DIR / ".cache" / "huggingface" / "download" / f"{task_id}.json.metadata"
    if not metadata.is_file():
        return None
    lines = metadata.read_text(encoding="utf-8").splitlines()
    return lines[0].strip() if lines else None


def ensure_local_task_exclude() -> None:
    """Keep fetched task data out of managed-source dirty-state checks."""

    exclude = EDGE_ROOT / ".git" / "info" / "exclude"
    if not exclude.is_file():
        return
    lines = exclude.read_text(encoding="utf-8").splitlines()
    if "tasks/" not in lines:
        exclude.write_text(
            "\n".join([*lines, "tasks/"]).rstrip() + "\n",
            encoding="utf-8",
        )


def provision(profile: dict[str, Any]) -> int:
    if not SFORGE.is_file():
        raise FileNotFoundError("SForge is not installed; run repro_env.py bootstrap --only edgebench")
    ensure_local_task_exclude()
    env = dict(os.environ)
    configure_temp_environment(env)
    fetch = [
        str(SFORGE),
        "--tasks-dir",
        str(TASKS_DIR),
        "fetch-tasks",
        "--repo",
        str(profile["dataset_repository"]),
        "--revision",
        str(profile["dataset_revision"]),
    ]
    completed = subprocess.run(fetch, cwd=ROOT, env=env, check=False)
    if completed.returncode != 0:
        return completed.returncode
    pull = [
        str(SFORGE),
        "--tasks-dir",
        str(TASKS_DIR),
        "pull",
        "--task",
        *[str(task) for task in profile["task_ids"]],
        "--registry",
        str(profile["registry"]),
    ]
    return subprocess.run(pull, cwd=ROOT, env=env, check=False).returncode


def doctor_payload(profile: dict[str, Any]) -> dict[str, Any]:
    expected_edge = upstream_entry("edgebench")["tracking_branch"]
    expected_goal = upstream_entry("goal_plus")["tracking_branch"]
    api_protocol = api_protocol_for_methods(profile["methods"])
    agents = {str(METHODS[method]["agent"]) for method in profile["methods"]}
    checks: list[dict[str, Any]] = []

    def add(name: str, passed: bool, **details: Any) -> None:
        checks.append({"name": name, "passed": bool(passed), **details})

    official_protocol: dict[str, Any] | None = None
    try:
        official_protocol = load_official_codex_protocol()
        add(
            "protocol:official-source",
            True,
            path=official_protocol["source"],
            sha256=official_protocol["source_sha256"],
            task_count=len(official_protocol["tasks"]),
        )
        missing_protocol_tasks = sorted(
            set(profile["task_ids"]) - set(official_protocol["tasks"])
        )
        add(
            "protocol:task-coverage",
            not missing_protocol_tasks,
            profile_task_count=len(profile["task_ids"]),
            official_task_count=len(official_protocol["tasks"]),
            missing_tasks=missing_protocol_tasks,
        )
    except (OSError, ValueError, yaml.YAMLError) as exc:
        add(
            "protocol:official-source",
            False,
            path=portable_path(OFFICIAL_CODEX_PROTOCOL_PATH),
            error=str(exc),
        )

    add(
        "checkout:edgebench",
        git_branch(EDGE_ROOT) == expected_edge and git_dirty(EDGE_ROOT) is False,
        expected_branch=expected_edge,
        actual_branch=git_branch(EDGE_ROOT),
        actual_commit=git_head(EDGE_ROOT),
        dirty=git_dirty(EDGE_ROOT),
    )
    add(
        "checkout:goal-plus",
        git_branch(GOAL_PLUS_ROOT) == expected_goal
        and git_dirty(GOAL_PLUS_ROOT) is False,
        expected_branch=expected_goal,
        actual_branch=git_branch(GOAL_PLUS_ROOT),
        actual_commit=git_head(GOAL_PLUS_ROOT),
        dirty=git_dirty(GOAL_PLUS_ROOT),
    )
    add("entrypoint:sforge", SFORGE.is_file(), path=" .bench-env/venv/bin/sforge".strip())
    imports = run_capture(
        [str(VENV_PYTHON), "-c", "import fastapi, sforge"]
    ) if VENV_PYTHON.is_file() else {"returncode": 127, "stderr": "venv missing"}
    add(
        "runtime:sforge-server-dependencies",
        imports["returncode"] == 0,
        stderr=imports["stderr"][-400:] or None,
    )
    add("runtime:repository-local-temp", ensure_temp_root().is_dir(), path=".tmp")

    api_config = resolve_agent_api_config(protocol=api_protocol)
    auth_override = os.environ.get("SFORGE_CODEX_AUTH_FILE")
    codex_home = Path(os.environ.get("CODEX_HOME", Path.home() / ".codex"))
    auth = Path(auth_override).expanduser() if auth_override else codex_home / "auth.json"
    api_key = api_config["api_key"]
    api_base_url = api_config["api_base_url"]
    add(
        "auth:agent",
        bool(api_key) or (api_protocol == "openai" and auth.is_file()),
        mode="api_key" if api_key else "oauth",
        protocol=api_protocol,
        api_key_source=api_config["api_key_source"],
        api_base_url_source=api_config["api_base_url_source"],
        policy=(
            "SFORGE_AGENT_* > protocol-native environment; Codex may "
            "otherwise use SFORGE_CODEX_AUTH_FILE or CODEX_HOME/auth.json"
        ),
    )
    if api_key and api_base_url:
        api_probe = authenticated_api_probe(
            str(api_base_url),
            str(api_key),
            protocol=api_protocol,
            model=str(profile["model"]),
            thinking=profile.get("thinking"),
            reasoning_effort=profile.get("reasoning_effort"),
        )
        add(
            "auth:agent-api-host",
            bool(api_probe["passed"]),
            base_url=str(api_base_url),
            status=api_probe.get("status"),
            error=api_probe.get("error"),
        )
        try:
            loopback = loopback_api_target(str(api_base_url)) is not None
        except ValueError as exc:
            loopback = False
            add("auth:agent-api-url", False, error=str(exc))
        if loopback:
            add(
                "runtime:rootless-loopback-bridge",
                Path("/usr/bin/systemd-socket-activate").is_file()
                and Path("/lib/systemd/systemd-socket-proxyd").is_file(),
                mechanism="systemd-socket-proxyd",
            )

    if any(agent.startswith("codex") for agent in agents):
        codex_runtime = (
            Path.home()
            / ".cache"
            / "sforge"
            / "codex"
            / "codex-0.144.1-linux-x64.tgz"
        )
        add(
            "runtime:codex-host-cache",
            codex_runtime.is_file() and codex_runtime.stat().st_size > 0,
            path=str(codex_runtime),
            size=(codex_runtime.stat().st_size if codex_runtime.is_file() else None),
        )

    docker_info = run_capture(["docker", "info", "--format", "{{json .}}"])
    docker_details: dict[str, Any] = {}
    if docker_info["returncode"] == 0:
        try:
            docker_details = json.loads(docker_info["stdout"])
        except json.JSONDecodeError:
            docker_details = {}
    architecture = str(docker_details.get("Architecture") or "").lower()
    add(
        "docker:engine",
        docker_info["returncode"] == 0 and bool(docker_details),
        architecture=architecture or None,
        stderr=docker_info["stderr"][-400:] or None,
    )
    add(
        "docker:linux-amd64",
        architecture in {"amd64", "x86_64"},
        required="linux/amd64",
        actual=architecture or None,
    )

    rust_archive: dict[str, Any] | None = None
    resource_probe_image: str | None = None
    effective_protocols: list[dict[str, Any]] = []
    offline_task_ids: list[str] = []

    for task_id in profile["task_ids"]:
        task_path = TASKS_DIR / f"{task_id}.json"
        add(f"task:{task_id}", task_path.is_file(), path=portable_path(task_path))
        actual_revision = dataset_revision(task_id)
        add(
            f"dataset-revision:{task_id}",
            actual_revision == profile["dataset_revision"],
            expected=profile["dataset_revision"],
            actual=actual_revision,
        )
        if not task_path.is_file():
            continue
        config = task_config(task_id)
        if official_protocol is not None:
            try:
                effective = profile_task_protocol(
                    profile, official_protocol, task_id, config
                )
                effective_protocols.append(effective)
                if effective["internet"] is False:
                    offline_task_ids.append(task_id)
                add(
                    f"protocol-effective:{task_id}",
                    True,
                    internet=effective["internet"],
                    internet_source=(
                        f"profiles/{profile['id']}.protocol_overrides.internet"
                        if "internet" in profile.get("protocol_overrides", {})
                        else f"tasks/{task_id}.json"
                    ),
                    submission_cooldown=effective["submission_cooldown"],
                )
            except ValueError as exc:
                add(f"protocol-effective:{task_id}", False, error=str(exc))
        if config.get("base_image") == "rust" and rust_archive is None:
            rust_archive = rust_runtime_archive_status()
            rust_details = {
                key: value
                for key, value in rust_archive.items()
                if key != "passed"
            }
            add(
                "runtime:rust-host-cache",
                bool(rust_archive["passed"]),
                **rust_details,
            )
        for image_index, image in enumerate(task_images(task_id)):
            inspected = run_capture(["docker", "image", "inspect", image])
            add(
                f"image:{image}",
                inspected["returncode"] == 0,
                image=image,
            )
            if (
                image_index == 0
                and inspected["returncode"] == 0
                and resource_probe_image is None
            ):
                resource_probe_image = image
            if config.get("base_image") == "rust" and inspected["returncode"] == 0:
                assert rust_archive is not None
                version = str(rust_archive.get("version") or "")
                probe = (
                    rust_image_runtime_probe(image, version)
                    if version
                    else {
                        "returncode": 1,
                        "stdout": "",
                        "stderr": "Rust runtime manifest has no version",
                    }
                )
                native = probe["returncode"] == 0
                add(
                    f"runtime:rust:{image}",
                    native or bool(rust_archive["passed"]),
                    image=image,
                    expected_version=version,
                    native=native,
                    fallback_archive_ready=bool(rust_archive["passed"]),
                    stdout=probe["stdout"][-400:] or None,
                    stderr=probe["stderr"][-400:] or None,
                )

    if effective_protocols:
        work_cpu_limit = max(
            max(int(item["work_cpu_limit"]), int(item["judge_cpu_limit"]))
            for item in effective_protocols
        )
        work_mem_limit = max(
            (
                str(limit)
                for item in effective_protocols
                for limit in (item["work_mem_limit"], item["judge_mem_limit"])
            ),
            key=_docker_memory_bytes,
        )
        daemon_cpu_support = docker_details.get("CpuCfsQuota")
        daemon_memory_support = docker_details.get("MemoryLimit")
        daemon_resource_support = (
            daemon_cpu_support is not False and daemon_memory_support is not False
        )
        if resource_probe_image:
            resource_probe = docker_resource_limit_probe(
                resource_probe_image,
                cpu_limit=work_cpu_limit,
                mem_limit=work_mem_limit,
            )
        else:
            resource_probe = {
                "passed": False,
                "error": "no prepared Work image is available for the resource probe",
            }
        add(
            "docker:official-resource-limits",
            daemon_resource_support and bool(resource_probe["passed"]),
            daemon_cpu_cfs_quota=daemon_cpu_support,
            daemon_memory_limit=daemon_memory_support,
            **{key: value for key, value in resource_probe.items() if key != "passed"},
        )

    if offline_task_ids:
        isolation_probe = sforge_iptables_permission_probe()
        add(
            "network:offline-task-isolation",
            bool(isolation_probe["passed"]),
            mechanism="SForge passwordless sudo iptables allowlist",
            offline_task_count=len(offline_task_ids),
            sample_task=offline_task_ids[0],
            stderr=isolation_probe.get("stderr"),
        )

    return {
        "schema_version": 1,
        "checked_at": utc_now(),
        "profile": profile["id"],
        "ok": all(check["passed"] for check in checks),
        "checks": checks,
    }


def doctor(profile: dict[str, Any], *, output: Path | None = None) -> int:
    payload = doctor_payload(profile)
    if output:
        write_json(output, payload)
    print(json.dumps(payload, indent=2, ensure_ascii=False))
    return 0 if payload["ok"] else 1


def sanitize_id(value: str) -> str:
    clean = re.sub(r"[^A-Za-z0-9_.-]+", "-", value).strip("-")
    if not clean:
        raise ValueError("campaign id must contain a letter or digit")
    return clean


def prepare(args: argparse.Namespace, profile: dict[str, Any]) -> Path:
    official_protocol = load_official_codex_protocol()
    methods = args.method or list(profile["methods"])
    unknown = set(methods) - set(METHODS)
    if unknown:
        raise ValueError("unknown EdgeBench method(s): " + ", ".join(sorted(unknown)))
    api_protocol = api_protocol_for_methods(methods)
    wall_time = int(args.wall_time_seconds or profile["wall_time_seconds"])
    concurrency = int(args.concurrency or profile["concurrency"])
    cell_concurrency = int(
        getattr(args, "cell_concurrency", None)
        or profile.get("cell_concurrency", 1)
    )
    model = args.model or profile["model"]
    requested_reasoning = getattr(args, "reasoning_effort", None)
    if requested_reasoning is not None:
        reasoning = requested_reasoning
    elif "reasoning_effort" in profile:
        reasoning = profile["reasoning_effort"]
    elif api_protocol == "anthropic":
        reasoning = None
    else:
        reasoning = "high"
    thinking = profile.get("thinking") if api_protocol == "anthropic" else None
    if api_protocol == "anthropic":
        validate_claude_thinking_contract(thinking, reasoning)
    backend = str(profile.get("backend") or "docker")
    judge_concurrency = int(profile.get("judge_concurrency", 1))
    override_reasons = dict(profile["protocol_override_reasons"])
    profile_protocol_overrides = dict(profile.get("protocol_overrides") or {})
    allowed_protocol_override_fields = (
        ALLOWED_PROTOCOL_OVERRIDE_FIELDS | set(profile_protocol_overrides)
    )
    if wall_time < 1 or concurrency < 1 or cell_concurrency < 1:
        raise ValueError(
            "wall time, concurrency, and cell concurrency must be positive"
        )

    campaign_id = sanitize_id(
        args.campaign_id or f"{profile['id']}-{campaign_stamp()}"
    )
    destination = RUNS_ROOT / campaign_id
    if destination.exists():
        raise FileExistsError(
            f"campaign already exists and will not be overwritten: {destination}"
        )
    destination.mkdir(parents=True)

    cells: list[dict[str, Any]] = []
    for task_id in profile["task_ids"]:
        config = task_config(task_id)
        official_effective = official_task_protocol(
            official_protocol, task_id, config
        )
        profile_effective = profile_task_protocol(
            profile, official_protocol, task_id, config
        )
        official_contract = {
            **official_effective,
            "attempts_per_task": OFFICIAL_SCHEDULED_RUNS,
            "cell_concurrency": None,
            "judge_concurrency": None,
            "model": official_protocol["official_model"],
            "reasoning_effort": None,
        }
        prompt = str(config["work"]["agent_query"])
        for method in methods:
            method_config = METHODS[method]
            cell_id = sanitize_id(f"{task_id}--{method}")
            outer_replicas = (
                concurrency
                if method_config["outer_replicas"] == "concurrency"
                else int(method_config["outer_replicas"])
            )
            effective_contract = {
                **profile_effective,
                "agent": method_config["agent"],
                "attempts_per_task": outer_replicas,
                "backend": backend,
                "cell_concurrency": cell_concurrency,
                "judge_concurrency": judge_concurrency,
                "model": model,
                "reasoning_effort": reasoning,
                "timeout": wall_time,
            }
            protocol_diff = _protocol_diff(
                official=official_contract,
                effective=effective_contract,
                reasons=override_reasons,
                allowed_fields=allowed_protocol_override_fields,
            )
            cell = {
                "schema_version": 1,
                "cell_id": cell_id,
                "task_id": task_id,
                "method": method,
                "sforge_agent": method_config["agent"],
                "api_protocol": method_config["api_protocol"],
                "backend": backend,
                "model": model,
                "reasoning_effort": reasoning,
                "thinking": thinking,
                "claude_context_window_tokens": profile.get(
                    "claude_context_window_tokens"
                ),
                "claude_autocompact_percent": profile.get(
                    "claude_autocompact_percent"
                ),
                "wall_time_seconds": wall_time,
                "live_search_concurrency": concurrency,
                "outer_replicas": outer_replicas,
                "outer_replica_concurrency": concurrency if outer_replicas > 1 else 1,
                "inner_search_concurrency": concurrency
                if method_config["inner_search"]
                else 0,
                "worker_runtime_seconds": min(
                    wall_time,
                    int(profile.get("worker_runtime_seconds", wall_time)),
                ),
                "eval_interval_seconds": int(effective_contract["eval_interval"]),
                "judge_concurrency": judge_concurrency,
                "judge_port": int(profile.get("judge_port", 8080)),
                "work_cpu_limit": effective_contract["work_cpu_limit"],
                "work_mem_limit": effective_contract["work_mem_limit"],
                "judge_cpu_limit": effective_contract["judge_cpu_limit"],
                "judge_mem_limit": effective_contract["judge_mem_limit"],
                "submission_cooldown": effective_contract[
                    "submission_cooldown"
                ],
                "max_submissions": effective_contract["max_submissions"],
                "auto_eval_enabled": not effective_contract[
                    "disable_auto_eval"
                ],
                "auto_resume_enabled": not effective_contract[
                    "disable_auto_resume"
                ],
                "stop_hook_enabled": not effective_contract[
                    "disable_stop_hook"
                ],
                "internet": effective_contract["internet"],
                "internet_source": (
                    f"profiles/{profile['id']}.protocol_overrides.internet"
                    if "internet" in profile_protocol_overrides
                    else f"tasks/{task_id}.json"
                ),
                "protocol_source": {
                    "path": official_protocol["source"],
                    "sha256": official_protocol["source_sha256"],
                },
                "official_defaults": official_protocol["defaults"],
                "official_task_overrides": official_protocol["tasks"][task_id],
                "official_effective_protocol": official_contract,
                "effective_protocol": effective_contract,
                "intentional_overrides": {
                    entry["field"]: {
                        "value": entry["effective"],
                        "reason": entry["reason"],
                    }
                    for entry in protocol_diff
                },
                "protocol_diff": protocol_diff,
                "protocol_classification": (
                    "official_protocol"
                    if not protocol_diff
                    else "official_protocol_with_intentional_overrides"
                ),
                "official_edgebench_comparable": not protocol_diff,
                "prompt_sha256": sha256_text(prompt),
                "metric_direction": config["judge"].get("score_direction", "maximize"),
                "sforge_run_id": sanitize_id(
                    f"{campaign_id}-{task_id}-{method}"
                ),
                "state": "prepared",
                "created_at": utc_now(),
            }
            cell_path = destination / "cells" / cell_id
            cell_path.mkdir(parents=True)
            write_json(cell_path / "cell.json", cell)
            cells.append(
                {
                    "cell_id": cell_id,
                    "task_id": task_id,
                    "method": method,
                    "state": "prepared",
                    "official_edgebench_comparable": not protocol_diff,
                }
            )

    snapshot = {
        **profile,
        "methods": methods,
        "model": model,
        "reasoning_effort": reasoning,
        "api_protocol": api_protocol,
        "thinking": thinking,
        "wall_time_seconds": wall_time,
        "concurrency": concurrency,
        "cell_concurrency": cell_concurrency,
        "protocol_source": {
            "path": official_protocol["source"],
            "sha256": official_protocol["source_sha256"],
        },
    }
    write_json(destination / "profile.json", snapshot)
    campaign_official_comparable = all(
        item["official_edgebench_comparable"] for item in cells
    )
    write_json(
        destination / "campaign.json",
        {
            "schema_version": 1,
            "campaign_id": campaign_id,
            "profile": profile["id"],
            "state": "prepared",
            "created_at": utc_now(),
            "edgebench_tracking_branch": upstream_entry("edgebench")[
                "tracking_branch"
            ],
            "edgebench_branch": git_branch(EDGE_ROOT),
            "edgebench_commit": git_head(EDGE_ROOT),
            "goal_plus_tracking_branch": upstream_entry("goal_plus")[
                "tracking_branch"
            ],
            "goal_plus_branch": git_branch(GOAL_PLUS_ROOT),
            "goal_plus_commit": git_head(GOAL_PLUS_ROOT),
            "dataset_revision": profile["dataset_revision"],
            "task_ids": list(profile["task_ids"]),
            "methods": methods,
            "model": model,
            "reasoning_effort": reasoning,
            "api_protocol": api_protocol,
            "thinking": thinking,
            "wall_time_seconds": wall_time,
            "concurrency": concurrency,
            "cell_concurrency": cell_concurrency,
            "protocol_source": {
                "path": official_protocol["source"],
                "sha256": official_protocol["source_sha256"],
                "official_model": official_protocol["official_model"],
                "stagger_seconds": official_protocol["stagger_seconds"],
            },
            "protocol_classification": (
                "official_protocol"
                if campaign_official_comparable
                else "official_protocol_with_intentional_overrides"
            ),
            "official_edgebench_comparable": campaign_official_comparable,
            "cells": cells,
        },
    )
    write_json(
        destination / "controller.json",
        {
            "schema_version": 1,
            "state": "prepared",
            "created_at": utc_now(),
            "pid": None,
            "pgid": None,
        },
    )
    print(portable_path(destination))
    return destination


def build_sforge_command(destination: Path, cell: dict[str, Any]) -> list[str]:
    cell_path = destination / "cells" / cell["cell_id"]
    command = [
        str(SFORGE),
        "--log-dir",
        str(cell_path / "sforge"),
        "--tasks-dir",
        str(TASKS_DIR),
        "run",
        "--backend",
        str(cell.get("backend") or "docker"),
        "--task",
        str(cell["task_id"]),
        "--agent",
        str(cell["sforge_agent"]),
        "--model",
        str(cell["model"]),
        "--timeout",
        str(cell["wall_time_seconds"]),
        "--eval-interval",
        str(cell["eval_interval_seconds"]),
        "--run-id",
        str(cell["sforge_run_id"]),
        "--replicas",
        str(cell["outer_replicas"]),
        "--replica-concurrency",
        str(cell["outer_replica_concurrency"]),
        "--judge-concurrency",
        str(cell["judge_concurrency"]),
        "--judge-url",
        str(
            cell.get("judge_url")
            or f"http://host.docker.internal:{cell.get('judge_port', 8080)}"
        ),
    ]
    if cell.get("work_cpu_limit") is not None:
        command.extend(["--work-cpu-limit", str(cell["work_cpu_limit"])])
    if cell.get("work_mem_limit") is not None:
        command.extend(["--work-mem-limit", str(cell["work_mem_limit"])])
    if cell.get("judge_cpu_limit") is not None:
        command.extend(["--judge-cpu-limit", str(cell["judge_cpu_limit"])])
    if cell.get("judge_mem_limit") is not None:
        command.extend(["--judge-mem-limit", str(cell["judge_mem_limit"])])
    if cell.get("submission_cooldown") is not None:
        command.extend(
            ["--submission-cooldown", str(cell["submission_cooldown"])]
        )
    if cell.get("max_submissions") is not None:
        command.extend(["--max-submissions", str(cell["max_submissions"])])
    if not cell.get("auto_eval_enabled", True):
        command.append("--disable-auto-eval")
    if not cell.get("auto_resume_enabled", True):
        command.append("--disable-auto-resume")
    if not cell.get("stop_hook_enabled", True):
        command.append("--disable-stop-hook")
    command.append(
        "--enable-internet" if cell["internet"] else "--disable-internet"
    )
    return command


def merge_agent_extra_env(
    env: dict[str, str],
    additions: dict[str, str],
    *,
    removals: Iterable[str] = (),
) -> None:
    entries: dict[str, str] = {}
    for item in env.get("SFORGE_AGENT_EXTRA_ENV", "").split(","):
        if "=" in item:
            key, value = item.split("=", 1)
            if key.strip():
                entries[key.strip()] = value.strip()
    for key in removals:
        entries.pop(key, None)
    entries.update(additions)
    if entries:
        env["SFORGE_AGENT_EXTRA_ENV"] = ",".join(
            f"{key}={value}" for key, value in entries.items()
        )
    else:
        env.pop("SFORGE_AGENT_EXTRA_ENV", None)


def cell_environment(
    cell: dict[str, Any],
    *,
    api_key: str | None = None,
    api_base_url: str | None = None,
    bridge_host: str | None = None,
) -> dict[str, str]:
    env = dict(os.environ)
    configure_temp_environment(env)
    internet = bool(cell.get("internet", True))
    if internet:
        for sforge_key, candidates in (
            (
                "SFORGE_HTTP_PROXY",
                ("SFORGE_HTTP_PROXY", "HTTP_PROXY", "http_proxy"),
            ),
            (
                "SFORGE_HTTPS_PROXY",
                ("SFORGE_HTTPS_PROXY", "HTTPS_PROXY", "https_proxy"),
            ),
        ):
            value = next((env[key] for key in candidates if env.get(key)), None)
            if value:
                env[sforge_key] = value.replace(
                    "127.0.0.1", "host.docker.internal"
                ).replace("localhost", "host.docker.internal")
    else:
        for key in (
            "ALL_PROXY",
            "HTTP_PROXY",
            "HTTPS_PROXY",
            "SFORGE_HTTP_PROXY",
            "SFORGE_HTTPS_PROXY",
            "all_proxy",
            "http_proxy",
            "https_proxy",
        ):
            env.pop(key, None)
    env.setdefault("SFORGE_NODEJS_MIRROR_URL", "https://npmmirror.com/mirrors/node")
    env.setdefault("SFORGE_NPM_REGISTRY_URL", "https://registry.npmmirror.com")
    if api_key:
        env["SFORGE_AGENT_API_KEY"] = api_key
    if api_base_url:
        env["SFORGE_AGENT_API_BASE_URL"] = api_base_url
    if bridge_host:
        append_no_proxy(env, bridge_host)
    agent = str(cell.get("sforge_agent") or METHODS[cell["method"]]["agent"])
    if agent.startswith("codex"):
        env["SFORGE_CODEX_REASONING_EFFORT"] = str(cell["reasoning_effort"])
    elif agent == "claude-code":
        env["SFORGE_CLAUDE_CACHE_OPT"] = "1"
        model = str(cell.get("model") or "")
        claude_env = {}
        if model:
            claude_env.update(
                {
                    "ANTHROPIC_MODEL": model,
                    "ANTHROPIC_DEFAULT_OPUS_MODEL": model,
                    "ANTHROPIC_DEFAULT_SONNET_MODEL": model,
                    "ANTHROPIC_DEFAULT_HAIKU_MODEL": model,
                    "CLAUDE_CODE_SUBAGENT_MODEL": model,
                }
            )
        context_window = cell.get("claude_context_window_tokens")
        compact_percent = cell.get("claude_autocompact_percent")
        if context_window is not None:
            claude_env["CLAUDE_CODE_AUTO_COMPACT_WINDOW"] = str(context_window)
            claude_env["CLAUDE_AUTOCOMPACT_PCT_OVERRIDE"] = str(compact_percent)
        thinking_type = str((cell.get("thinking") or {}).get("type") or "")
        reasoning_value = cell.get("reasoning_effort")
        reasoning_effort = str(reasoning_value or "")
        thinking_controls = (
            "MAX_THINKING_TOKENS",
            "CLAUDE_CODE_DISABLE_THINKING",
            "CLAUDE_CODE_DISABLE_ADAPTIVE_THINKING",
            "CLAUDE_CODE_EFFORT_LEVEL",
            "CLAUDE_CODE_ALWAYS_ENABLE_EFFORT",
        )
        for key in thinking_controls:
            env.pop(key, None)
        env.update(claude_env)
        if thinking_type == "adaptive" and reasoning_value is None:
            merge_agent_extra_env(
                env,
                claude_env,
                removals=thinking_controls,
            )
        elif reasoning_effort in {"none", "minimal"}:
            merge_agent_extra_env(
                env,
                {
                    **claude_env,
                    "MAX_THINKING_TOKENS": "0",
                    "CLAUDE_CODE_DISABLE_THINKING": "1",
                },
                removals=thinking_controls,
            )
        else:
            merge_agent_extra_env(
                env,
                {
                    **claude_env,
                    "CLAUDE_CODE_EFFORT_LEVEL": reasoning_effort,
                    "CLAUDE_CODE_ALWAYS_ENABLE_EFFORT": "1",
                },
                removals=thinking_controls,
            )
    if cell["method"] == "goal-plus-codex":
        env["SFORGE_GOAL_PLUS_SOURCE_DIR"] = str(GOAL_PLUS_ROOT)
        extra_env = {
            "SFORGE_GOAL_PLUS_MAX_PARALLEL": str(
                cell["inner_search_concurrency"]
            ),
            "SFORGE_GOAL_PLUS_WORKER_RUNTIME_SECONDS": str(
                cell["worker_runtime_seconds"]
            ),
            "GOAL_PLUS_EVIDENCE_ANNOTATOR_MODEL": str(cell["model"]),
            "GOAL_PLUS_EVIDENCE_ANNOTATOR_REASONING_EFFORT": str(
                cell["reasoning_effort"]
            ),
        }
        if api_base_url:
            extra_env.update(
                {
                    "GOAL_PLUS_EVIDENCE_ANNOTATOR_BASE_URL": api_base_url,
                    "GOAL_PLUS_EVIDENCE_ANNOTATOR_PROVIDER_ID": (
                        EVIDENCE_ANNOTATOR_PROVIDER_ID
                    ),
                    "GOAL_PLUS_EVIDENCE_ANNOTATOR_PROVIDER_NAME": (
                        "EdgeBench Evidence provider"
                    ),
                    "GOAL_PLUS_EVIDENCE_ANNOTATOR_API_KEY_ENV": (
                        "SFORGE_AGENT_API_KEY"
                    ),
                    "GOAL_PLUS_EVIDENCE_ANNOTATOR_WIRE_API": "responses",
                }
            )
        merge_agent_extra_env(env, extra_env)
    return env


def process_alive(pid: int | None) -> bool:
    if not pid:
        return False
    try:
        os.kill(pid, 0)
    except (OSError, ProcessLookupError):
        return False
    return True


def judge_ready(port: int) -> bool:
    try:
        with urllib.request.urlopen(
            f"http://127.0.0.1:{port}/openapi.json",
            timeout=1.0,
        ) as response:
            return response.status == 200
    except (OSError, urllib.error.URLError):
        return False


def start_or_reuse_judge(
    destination: Path,
    port: int,
    controller: dict[str, Any],
    *,
    env: dict[str, str] | None = None,
) -> tuple[subprocess.Popen[str] | None, Any]:
    controller_path = destination / "controller.json"
    if judge_ready(port):
        controller.update(
            {
                "judge_owned": False,
                "judge_pid": None,
                "judge_host_url": f"http://127.0.0.1:{port}",
                "judge_container_url": f"http://host.docker.internal:{port}",
            }
        )
        write_json(controller_path, controller)
        return None, lambda: None

    command = [
        str(SFORGE),
        "--log-dir",
        str(destination / "judge"),
        "--tasks-dir",
        str(TASKS_DIR),
        "serve",
        "--host",
        "127.0.0.1",
        "--port",
        str(port),
    ]
    log = (destination / "judge.log").open("a", encoding="utf-8")
    process = subprocess.Popen(
        command,
        cwd=ROOT,
        env=env or dict(configure_temp_environment(dict(os.environ))),
        stdout=log,
        stderr=subprocess.STDOUT,
        text=True,
        start_new_session=True,
    )
    log.close()
    controller.update(
        {
            "judge_owned": True,
            "judge_pid": process.pid,
            "judge_command": portable_command(command),
            "judge_host_url": f"http://127.0.0.1:{port}",
            "judge_container_url": f"http://host.docker.internal:{port}",
        }
    )
    write_json(controller_path, controller)

    def close_judge() -> None:
        if process.poll() is not None:
            return
        process.send_signal(signal.SIGINT)
        try:
            process.wait(timeout=10)
        except subprocess.TimeoutExpired:
            controller["judge_closeout_incomplete"] = True
            write_json(controller_path, controller)

    atexit.register(close_judge)
    deadline = time.monotonic() + 30
    while time.monotonic() < deadline:
        if judge_ready(port):
            return process, close_judge
        if process.poll() is not None:
            break
        time.sleep(0.25)
    close_judge()
    raise RuntimeError(
        f"SForge judge did not become ready; inspect {portable_path(destination / 'judge.log')}"
    )


def cell_has_scored_results(destination: Path, cell: dict[str, Any]) -> bool:
    cell_path = destination / "cells" / cell["cell_id"]
    task_runs = sorted(
        (cell_path / "sforge" / "runs").glob(f"*/{cell['task_id']}")
    )
    if len(task_runs) < int(cell["outer_replicas"]):
        return False
    for task_run in task_runs:
        final_path = task_run / "final_result.json"
        if not final_path.is_file():
            return False
        final = read_json(final_path)
        scored_reports = list((task_run / "submissions").glob("*/report.json"))
        if final.get("best_score") is None and not scored_reports:
            return False
    return True


def update_campaign_cell(
    destination: Path,
    cell_id: str,
    state: str,
) -> None:
    campaign = read_json(destination / "campaign.json")
    for item in campaign["cells"]:
        if item["cell_id"] == cell_id:
            item["state"] = state
            break
    campaign["updated_at"] = utc_now()
    write_json(destination / "campaign.json", campaign)


def start_campaign_cell(
    destination: Path,
    cell_summary: dict[str, Any],
    *,
    judge_container_url: str,
    api_config: dict[str, str | None],
    api_key: str | None,
    runtime_api_base_url: str | None,
    bridge_host: str | None,
) -> dict[str, Any] | None:
    cell_id = str(cell_summary["cell_id"])
    cell_path = destination / "cells" / cell_id
    cell_file = cell_path / "cell.json"
    cell = read_json(cell_file)
    if cell.get("state") == "completed":
        return None
    cell["judge_url"] = judge_container_url
    command = build_sforge_command(destination, cell)
    write_json(
        cell_path / "command.json",
        {
            "command": portable_command(command),
            "environment_policy": {
                "credentials": (
                    "host API environment mapped to SForge; values are never persisted"
                    if api_key
                    else "host Codex OAuth; auth contents are never persisted"
                ),
                "api_key_source": api_config["api_key_source"],
                "api_base_url_source": api_config["api_base_url_source"],
                "temp": ".tmp",
                "goal_plus_source": "third_party/goal-plus"
                if cell["method"] == "goal-plus-codex"
                else None,
            },
        },
    )
    log = (cell_path / "controller.log").open("a", encoding="utf-8")
    try:
        process = subprocess.Popen(
            command,
            cwd=ROOT,
            env=cell_environment(
                cell,
                api_key=api_key,
                api_base_url=runtime_api_base_url,
                bridge_host=bridge_host,
            ),
            stdout=log,
            stderr=subprocess.STDOUT,
            text=True,
        )
    except Exception:
        log.close()
        raise
    cell.update(
        {
            "state": "running",
            "started_at": utc_now(),
            "pid": process.pid,
        }
    )
    write_json(cell_file, cell)
    update_campaign_cell(destination, cell_id, "running")
    return {
        "cell": cell,
        "cell_file": cell_file,
        "log": log,
        "process": process,
    }


def finish_campaign_cell(
    destination: Path,
    active: dict[str, Any],
    *,
    stop_requested: bool,
) -> int:
    process: subprocess.Popen[str] = active["process"]
    returncode = process.poll()
    if returncode is None:
        returncode = process.wait()
    active["log"].close()
    cell = active["cell"]
    scored = cell_has_scored_results(destination, cell)
    if returncode == 0 and not stop_requested and not scored:
        returncode = 1
        cell["result_validation_error"] = (
            "SForge exited without the expected scored final result"
        )
    cell.update(
        {
            "state": "interrupted"
            if stop_requested
            else "completed"
            if returncode == 0
            else "failed",
            "returncode": returncode,
            "finished_at": utc_now(),
        }
    )
    write_json(active["cell_file"], cell)
    update_campaign_cell(destination, str(cell["cell_id"]), str(cell["state"]))
    return int(returncode)


def execute_cell_queue(
    destination: Path,
    campaign: dict[str, Any],
    controller: dict[str, Any],
    *,
    cell_concurrency: int,
    judge_container_url: str,
    api_config: dict[str, str | None],
    api_key: str | None,
    runtime_api_base_url: str | None,
    bridge_host: str | None,
    stop_requested: Any,
) -> int:
    controller_path = destination / "controller.json"
    pending = deque(campaign["cells"])
    active: dict[str, dict[str, Any]] = {}
    overall_returncode = 0
    stop_forwarded = False

    def record_active() -> None:
        controller["active_children"] = {
            cell_id: {
                "pid": running["process"].pid,
                "task_id": running["cell"]["task_id"],
                "started_at": running["cell"]["started_at"],
            }
            for cell_id, running in sorted(active.items())
        }
        write_json(controller_path, controller)

    controller["cell_concurrency"] = cell_concurrency
    record_active()
    while pending or active:
        while pending and len(active) < cell_concurrency and not stop_requested():
            cell_summary = pending.popleft()
            cell_id = str(cell_summary["cell_id"])
            try:
                running = start_campaign_cell(
                    destination,
                    cell_summary,
                    judge_container_url=judge_container_url,
                    api_config=api_config,
                    api_key=api_key,
                    runtime_api_base_url=runtime_api_base_url,
                    bridge_host=bridge_host,
                )
            except Exception as exc:
                cell_path = destination / "cells" / cell_id
                cell_file = cell_path / "cell.json"
                cell = read_json(cell_file)
                cell.update(
                    {
                        "state": "failed",
                        "returncode": 1,
                        "finished_at": utc_now(),
                        "launch_error": str(exc),
                    }
                )
                write_json(cell_file, cell)
                update_campaign_cell(destination, cell_id, "failed")
                overall_returncode = overall_returncode or 1
                continue
            if running is not None:
                active[cell_id] = running
                record_active()

        if stop_requested() and not stop_forwarded:
            for running in active.values():
                process = running["process"]
                if process.poll() is None:
                    try:
                        process.send_signal(signal.SIGINT)
                    except ProcessLookupError:
                        pass
            stop_forwarded = True

        completed = [
            cell_id
            for cell_id, running in active.items()
            if running["process"].poll() is not None
        ]
        if not completed:
            if not active:
                break
            time.sleep(0.25)
            continue

        for cell_id in completed:
            running = active.pop(cell_id)
            returncode = finish_campaign_cell(
                destination,
                running,
                stop_requested=bool(stop_requested()),
            )
            if returncode != 0:
                overall_returncode = overall_returncode or returncode
        record_active()

    return overall_returncode


def execute_campaign(destination: Path) -> int:
    controller_path = destination / "controller.json"
    controller = read_json(controller_path)
    controller.update(
        {
            "state": "running",
            "started_at": utc_now(),
            "pid": os.getpid(),
            "pgid": os.getpgrp(),
        }
    )
    write_json(controller_path, controller)
    campaign = read_json(destination / "campaign.json")
    campaign["state"] = "running"
    campaign["started_at"] = utc_now()
    write_json(destination / "campaign.json", campaign)

    stop_requested = False

    def handle_stop(signum: int, _frame: Any) -> None:
        nonlocal stop_requested
        stop_requested = True
        controller["stop_requested_at"] = utc_now()
        controller["stop_signal"] = signal.Signals(signum).name
        write_json(controller_path, controller)

    signal.signal(signal.SIGINT, handle_stop)
    signal.signal(signal.SIGTERM, handle_stop)

    profile = read_json(destination / "profile.json")
    judge_port = int(profile.get("judge_port", 8080))
    api_protocol = str(
        profile.get("api_protocol")
        or api_protocol_for_methods(profile["methods"])
    )
    judge_process: subprocess.Popen[str] | None = None
    close_judge = lambda: None
    bridge_processes: list[subprocess.Popen[str]] = []
    bridge_closers: list[Any] = []
    controller["bridges"] = []
    api_config = resolve_agent_api_config(protocol=api_protocol)
    api_key = api_config["api_key"]
    api_base_url = api_config["api_base_url"]
    runtime_api_base_url = str(api_base_url) if api_base_url else None
    bridge_host: str | None = None
    judge_container_url = f"http://host.docker.internal:{judge_port}"
    try:
        if api_protocol == "anthropic" and (not api_key or not runtime_api_base_url):
            raise RuntimeError(
                "Claude Code campaigns require an API key and Anthropic base URL"
            )
        if runtime_api_base_url and loopback_api_target(runtime_api_base_url):
            bridge_host = default_route_ipv4()
            target_host, target_port = loopback_api_target(runtime_api_base_url) or ("", 0)
            api_bridge, metadata, close_api_bridge = start_socket_bridge(
                destination,
                name="agent-api",
                listen_host=bridge_host,
                target_host=target_host,
                target_port=target_port,
            )
            bridge_processes.append(api_bridge)
            bridge_closers.append(close_api_bridge)
            controller.setdefault("bridges", []).append(metadata)
            runtime_api_base_url = bridged_base_url(
                runtime_api_base_url,
                bridge_host,
                int(metadata["listen_port"]),
            )
            api_probe = authenticated_api_probe(
                runtime_api_base_url,
                str(api_key or ""),
                protocol=api_protocol,
                model=str(profile["model"]),
                thinking=profile.get("thinking"),
                reasoning_effort=profile.get("reasoning_effort"),
            )
            if not api_key or not api_probe["passed"]:
                raise RuntimeError(
                    "authenticated agent API bridge probe failed "
                    f"(HTTP {api_probe.get('status')})"
                )

        if api_key and runtime_api_base_url:
            probe_image = task_images(str(profile["task_ids"][0]))[0]
            container_probe = docker_http_probe(
                probe_image,
                runtime_api_base_url,
                api_key=str(api_key),
                protocol=api_protocol,
                model=str(profile["model"]),
                thinking_type=str((profile.get("thinking") or {}).get("type") or ""),
                reasoning_effort=str(profile.get("reasoning_effort") or ""),
            )
            if not container_probe["passed"]:
                raise RuntimeError(
                    "agent API is not reachable from an EdgeBench Work container "
                    f"(HTTP {container_probe.get('status')}; "
                    f"{container_probe.get('stderr') or 'no stderr'})"
                )

        judge_env = judge_server_environment(
            api_key=str(api_key) if api_key else None,
            api_base_url=runtime_api_base_url,
            bridge_host=bridge_host,
        )
        judge_process, close_judge = start_or_reuse_judge(
            destination,
            judge_port,
            controller,
            env=judge_env,
        )
        if sys.platform.startswith("linux"):
            bridge_host = bridge_host or default_route_ipv4()
            judge_bridge, metadata, close_judge_bridge = start_socket_bridge(
                destination,
                name="judge",
                listen_host=bridge_host,
                target_host="127.0.0.1",
                target_port=judge_port,
            )
            bridge_processes.append(judge_bridge)
            bridge_closers.append(close_judge_bridge)
            controller.setdefault("bridges", []).append(metadata)
            judge_container_url = (
                f"http://{bridge_host}:{int(metadata['listen_port'])}"
            )
            judge_probe = docker_http_probe(
                task_images(str(profile["task_ids"][0]))[0],
                judge_container_url + "/openapi.json",
            )
            if not judge_probe["passed"]:
                raise RuntimeError(
                    "Judge is not reachable from an EdgeBench Work container "
                    f"(HTTP {judge_probe.get('status')}; "
                    f"{judge_probe.get('stderr') or 'no stderr'})"
                )
        controller.update(
            {
                "agent_auth_mode": "api_key" if api_key else "oauth",
                "agent_api_protocol": api_protocol,
                "agent_api_key_source": api_config["api_key_source"],
                "agent_api_base_url_source": api_config["api_base_url_source"],
                "agent_container_api_base_url": runtime_api_base_url,
                "judge_container_url": judge_container_url,
            }
        )
        write_json(controller_path, controller)
    except Exception as exc:
        close_judge()
        for close_bridge in reversed(bridge_closers):
            close_bridge()
        campaign = read_json(destination / "campaign.json")
        campaign.update(
            {
                "state": "failed",
                "finished_at": utc_now(),
                "controller_error": str(exc),
            }
        )
        write_json(destination / "campaign.json", campaign)
        controller.update(
            {
                "state": "failed",
                "finished_at": utc_now(),
                "returncode": 1,
                "error": str(exc),
            }
        )
        write_json(controller_path, controller)
        return 1
    cell_concurrency = int(profile.get("cell_concurrency", 1))
    overall_returncode = execute_cell_queue(
        destination,
        campaign,
        controller,
        cell_concurrency=cell_concurrency,
        judge_container_url=judge_container_url,
        api_config=api_config,
        api_key=str(api_key) if api_key else None,
        runtime_api_base_url=runtime_api_base_url,
        bridge_host=bridge_host,
        stop_requested=lambda: stop_requested,
    )

    campaign = read_json(destination / "campaign.json")
    states = {cell["state"] for cell in campaign["cells"]}
    if stop_requested:
        final_state = "interrupted"
        overall_returncode = overall_returncode or 130
    elif states == {"completed"}:
        final_state = "completed"
    elif "failed" in states:
        final_state = "failed"
        overall_returncode = overall_returncode or 1
    else:
        final_state = "partial"
    campaign.update({"state": final_state, "finished_at": utc_now()})
    write_json(destination / "campaign.json", campaign)
    if judge_process is not None:
        close_judge()
    for close_bridge in reversed(bridge_closers):
        close_bridge()
    controller.update(
        {
            "state": final_state,
            "finished_at": utc_now(),
            "returncode": overall_returncode,
            "active_children": {},
            "judge_alive_after_closeout": process_alive(
                judge_process.pid if judge_process is not None else None
            ),
            "bridges_alive_after_closeout": [
                process_alive(process.pid) for process in bridge_processes
            ],
        }
    )
    write_json(controller_path, controller)
    finalize_campaign(destination)
    return overall_returncode


def launch(destination: Path, *, detach: bool) -> int:
    controller = read_json(destination / "controller.json")
    if process_alive(controller.get("pid")):
        raise RuntimeError(f"campaign controller is already running: {controller['pid']}")
    if not detach:
        return execute_campaign(destination)

    command = [
        str(VENV_PYTHON if VENV_PYTHON.is_file() else Path(sys.executable)),
        str(Path(__file__).resolve()),
        "_execute",
        "--campaign",
        portable_path(destination),
    ]
    log = (destination / "controller.log").open("a", encoding="utf-8")
    process = subprocess.Popen(
        command,
        cwd=ROOT,
        env=dict(configure_temp_environment(dict(os.environ))),
        stdout=log,
        stderr=subprocess.STDOUT,
        text=True,
        start_new_session=True,
    )
    log.close()
    controller = read_json(destination / "controller.json")
    controller.update(
        {
            "schema_version": 1,
            "launched_at": controller.get("launched_at") or utc_now(),
            "pid": process.pid,
            "pgid": process.pid,
            "command": portable_command(command),
        }
    )
    if controller.get("state") in {"prepared", "launching"}:
        controller["state"] = "launching"
    write_json(destination / "controller.json", controller)
    print(json.dumps({"pid": process.pid, "campaign": portable_path(destination)}))
    return 0


def status_payload(destination: Path) -> dict[str, Any]:
    campaign = read_json(destination / "campaign.json")
    controller = read_json(destination / "controller.json")
    cells: list[dict[str, Any]] = []
    for item in campaign["cells"]:
        cell_path = destination / "cells" / item["cell_id"]
        cell = read_json(cell_path / "cell.json")
        task_runs = sorted(
            (cell_path / "sforge" / "runs").glob(f"*/{cell['task_id']}")
        )
        final_results = [
            run / "final_result.json" for run in task_runs if (run / "final_result.json").is_file()
        ]
        cells.append(
            {
                "cell_id": item["cell_id"],
                "task_id": item["task_id"],
                "method": item["method"],
                "state": cell["state"],
                "pid": cell.get("pid"),
                "pid_alive": process_alive(cell.get("pid")),
                "completed_trajectories": len(final_results),
                "expected_trajectories": cell["outer_replicas"],
                "summary": portable_path(cell_path / "summary.json")
                if (cell_path / "summary.json").is_file()
                else None,
            }
        )
    return {
        "campaign": campaign["campaign_id"],
        "state": campaign["state"],
        "controller": {
            "state": controller["state"],
            "pid": controller.get("pid"),
            "pgid": controller.get("pgid"),
            "alive": process_alive(controller.get("pid")),
            "judge_owned": controller.get("judge_owned"),
            "judge_pid": controller.get("judge_pid"),
            "judge_alive": process_alive(controller.get("judge_pid")),
        },
        "cells": cells,
    }


def print_status(destination: Path, *, as_json: bool) -> int:
    payload = status_payload(destination)
    if as_json:
        print(json.dumps(payload, indent=2, ensure_ascii=False))
        return 0
    print(
        f"{payload['campaign']}: {payload['state']} "
        f"(controller alive={payload['controller']['alive']})"
    )
    for cell in payload["cells"]:
        print(
            f"- {cell['cell_id']}: {cell['state']}; "
            f"{cell['completed_trajectories']}/{cell['expected_trajectories']} trajectories"
        )
    return 0


def stop_campaign(destination: Path, *, wait_seconds: int) -> int:
    controller_path = destination / "controller.json"
    controller = read_json(controller_path)
    pid = controller.get("pid")
    pgid = controller.get("pgid")
    if not process_alive(pid):
        print("controller is not running; no signal sent")
        return 0
    if not pgid:
        raise RuntimeError("running controller has no recorded process group")
    os.kill(int(pid), signal.SIGINT)
    controller["state"] = "stopping"
    controller["stop_requested_at"] = utc_now()
    write_json(controller_path, controller)
    deadline = time.monotonic() + max(0, wait_seconds)
    while process_alive(pid) and time.monotonic() < deadline:
        time.sleep(0.25)
    payload = {
        "signal": "SIGINT",
        "pid": pid,
        "pgid": pgid,
        "alive_after_wait": process_alive(pid),
        "artifacts_preserved": True,
    }
    print(json.dumps(payload, indent=2))
    return 0 if not payload["alive_after_wait"] else 2


def iter_json_lines(text: str) -> Iterable[dict[str, Any]]:
    for line in text.splitlines():
        try:
            item = json.loads(line)
        except json.JSONDecodeError:
            continue
        if isinstance(item, dict):
            yield item


def add_usage(total: dict[str, int], event: dict[str, Any]) -> None:
    if event.get("type") != "turn.completed":
        return
    usage = event.get("usage")
    if not isinstance(usage, dict):
        return
    for key, value in usage.items():
        if isinstance(value, int) and not isinstance(value, bool):
            total[key] += value


def codex_usage(task_run: Path) -> dict[str, Any]:
    totals: dict[str, int] = defaultdict(int)
    session_ids: set[str] = set()
    archive_path = task_run / "codex-sessions.tar"
    coverage = "agent_output_only"
    if archive_path.is_file():
        coverage = "all_collected_codex_sessions"
        try:
            with tarfile.open(archive_path) as archive:
                for member in archive:
                    if not member.isfile() or not member.name.endswith(".jsonl"):
                        continue
                    extracted = archive.extractfile(member)
                    if extracted is None:
                        continue
                    text = extracted.read().decode("utf-8", errors="replace")
                    rollout_total: dict[str, int] = {}
                    for event in iter_json_lines(text):
                        if event.get("type") == "thread.started" and event.get("thread_id"):
                            session_ids.add(str(event["thread_id"]))
                        if event.get("type") == "session_meta":
                            payload = event.get("payload", {})
                            if isinstance(payload, dict):
                                session_id = payload.get("id") or payload.get("session_id")
                                if session_id:
                                    session_ids.add(str(session_id))
                        if event.get("type") == "event_msg":
                            payload = event.get("payload", {})
                            if (
                                isinstance(payload, dict)
                                and payload.get("type") == "token_count"
                            ):
                                info = payload.get("info", {})
                                usage = (
                                    info.get("total_token_usage", {})
                                    if isinstance(info, dict)
                                    else {}
                                )
                                if isinstance(usage, dict):
                                    rollout_total = {
                                        key: value
                                        for key, value in usage.items()
                                        if isinstance(value, int)
                                        and not isinstance(value, bool)
                                    }
                        add_usage(totals, event)
                    for key, value in rollout_total.items():
                        totals[key] += value
        except tarfile.TarError:
            coverage = "invalid_codex_sessions_archive"
    else:
        output = task_run / "agent_output.txt"
        if output.is_file():
            for event in iter_json_lines(output.read_text(encoding="utf-8", errors="replace")):
                if event.get("type") == "thread.started" and event.get("thread_id"):
                    session_ids.add(str(event["thread_id"]))
                add_usage(totals, event)
    return {
        "coverage": coverage,
        "session_count": len(session_ids),
        "tokens": dict(sorted(totals.items())),
    }


def goal_plus_stats(task_run: Path) -> dict[str, Any] | None:
    archive_path = task_run / "goal-plus-state.tar"
    if not archive_path.is_file():
        return None
    candidates: set[tuple[str, str]] = set()
    sessions = 0
    verifier_runs = 0
    search_runs: set[str] = set()
    search_run_states: dict[str, int] = defaultdict(int)
    annotation_usage: dict[str, int | float] = {}
    annotation_tasks = 0
    annotation_attempts = 0
    annotation_states: dict[str, int] = defaultdict(int)
    try:
        with tarfile.open(archive_path) as archive:
            for member in archive:
                run_match = re.search(r"/runs/([^/]+)/run\.json$", member.name)
                if run_match:
                    search_runs.add(run_match.group(1))
                    extracted = archive.extractfile(member)
                    if extracted:
                        try:
                            payload = json.loads(
                                extracted.read().decode("utf-8", errors="replace")
                            )
                            state = payload.get("state")
                            if state:
                                search_run_states[str(state)] += 1
                        except (json.JSONDecodeError, TypeError):
                            pass
                match = re.search(r"/runs/([^/]+)/candidates/([^/]+)/candidate\.json$", member.name)
                if match:
                    search_runs.add(match.group(1))
                    candidates.add((match.group(1), match.group(2)))
                if "/agent_sessions/" in member.name and member.name.endswith(".json"):
                    sessions += 1
                    extracted = archive.extractfile(member)
                    if extracted:
                        try:
                            payload = json.loads(
                                extracted.read().decode("utf-8", errors="replace")
                            )
                            verifier_runs += int(
                                payload.get("counters", {}).get("verifier_runs", 0)
                            )
                        except (json.JSONDecodeError, TypeError, ValueError):
                            pass
                if (
                    "/evidence-annotations/" in member.name
                    and member.name.endswith(".json")
                ):
                    extracted = archive.extractfile(member)
                    if extracted:
                        try:
                            payload = json.loads(
                                extracted.read().decode("utf-8", errors="replace")
                            )
                            annotation_tasks += 1
                            annotation_attempts += int(payload.get("attempts") or 0)
                            state = str(payload.get("state") or "unknown")
                            annotation_states[state] += 1
                            task_usage = payload.get("usage")
                            if not isinstance(task_usage, dict):
                                task_usage = {}
                            for key, value in task_usage.items():
                                if isinstance(value, (int, float)) and not isinstance(
                                    value, bool
                                ):
                                    annotation_usage[key] = (
                                        annotation_usage.get(key, 0) + value
                                    )
                        except (json.JSONDecodeError, TypeError, ValueError):
                            pass
    except tarfile.TarError:
        return {"archive": "invalid"}
    return {
        "search_runs": len(search_runs),
        "candidates": len(candidates),
        "agent_sessions": sessions,
        "worker_verifier_runs": verifier_runs,
        "search_run_states": dict(sorted(search_run_states.items())),
        "evidence_annotator_usage": {
            **annotation_usage,
            "tasks": annotation_tasks,
            "attempts": annotation_attempts,
            "states": dict(sorted(annotation_states.items())),
            "coverage": "persisted Goal Plus Evidence annotator turns",
        },
    }


def score_task_run(task_run: Path, cell: dict[str, Any]) -> dict[str, Any]:
    reporter = EDGE_ROOT / "scripts" / "report_edgebench_scores.py"
    command = [
        str(VENV_PYTHON if VENV_PYTHON.is_file() else Path(sys.executable)),
        str(reporter),
        "--run-dir",
        str(task_run),
        "--model",
        str(cell["model"]),
        "--budget-seconds",
        str(cell["wall_time_seconds"]),
        "--json",
    ]
    scored = run_capture(command, env=dict(configure_temp_environment(dict(os.environ))))
    if scored["returncode"] != 0:
        return {
            "task_run": portable_path(task_run),
            "error": scored["stderr"] or scored["stdout"],
        }
    observation = json.loads(scored["stdout"])
    observation["source"] = portable_path(Path(observation["source"]))
    observation["task_run"] = portable_path(task_run)
    final = read_json(task_run / "final_result.json")
    observation["runtime_seconds"] = final.get("runtime_seconds")
    observation["total_rounds"] = final.get("total_rounds")
    observation["agent_submissions"] = final.get("agent_submissions")
    observation["auto_submissions"] = final.get("auto_submissions")
    observation["resume_count"] = final.get("resume_count")
    observation["timed_out"] = final.get("timed_out")
    evaluator_calls = 0
    for history_name, entry_type in (
        ("run_history.json", "submission"),
        ("game_history.json", "game"),
    ):
        history_path = task_run / history_name
        if not history_path.is_file():
            continue
        entries = read_json(history_path).get("entries", [])
        if isinstance(entries, list):
            evaluator_calls += sum(
                1
                for entry in entries
                if isinstance(entry, dict) and entry.get("type") == entry_type
            )
    if not evaluator_calls:
        evaluator_calls = len(list((task_run / "submissions").glob("*/report.json")))
    observation["evaluator_calls"] = evaluator_calls
    observation["codex_usage"] = codex_usage(task_run)
    observation["goal_plus"] = goal_plus_stats(task_run)
    return observation


def summarize_cell(destination: Path, cell: dict[str, Any]) -> dict[str, Any]:
    cell_path = destination / "cells" / cell["cell_id"]
    task_runs = sorted(
        (cell_path / "sforge" / "runs").glob(f"*/{cell['task_id']}")
    )
    observations = [
        score_task_run(task_run, cell)
        for task_run in task_runs
        if (task_run / "final_result.json").is_file()
    ]
    valid = [item for item in observations if "edgebench_score" in item]
    best = max(valid, key=lambda item: float(item["edgebench_score"])) if valid else None
    summary = {
        "schema_version": 1,
        "cell_id": cell["cell_id"],
        "task_id": cell["task_id"],
        "method": cell["method"],
        "model": cell["model"],
        "reasoning_effort": cell["reasoning_effort"],
        "metric_direction": cell["metric_direction"],
        "wall_time_seconds": cell["wall_time_seconds"],
        "live_search_concurrency": cell["live_search_concurrency"],
        "outer_replicas": cell["outer_replicas"],
        "inner_search_concurrency": cell["inner_search_concurrency"],
        "expected_trajectories": cell["outer_replicas"],
        "completed_trajectories": len(observations),
        "valid_trajectories": len(valid),
        "observations": observations,
        "best": best,
        "protocol_classification": cell.get("protocol_classification"),
        "official_edgebench_comparable": cell.get(
            "official_edgebench_comparable", False
        ),
        "protocol_diff": cell.get("protocol_diff", []),
        "known_protocol_issue": paper_protocol_issue(cell),
        "finalized_at": utc_now(),
    }
    write_json(cell_path / "summary.json", summary)
    return summary


def paper_protocol_issue(cell: dict[str, Any]) -> str | None:
    task_id = str(cell["task_id"])
    if task_id == "borden_source_inversion":
        if cell.get("submission_cooldown") != 120:
            return LEGACY_PAPER_PROTOCOL_ISSUES[task_id]
    elif task_id == "exchange_core_throughput":
        resources = (
            cell.get("work_cpu_limit"),
            cell.get("work_mem_limit"),
            cell.get("judge_cpu_limit"),
            cell.get("judge_mem_limit"),
        )
        if cell.get("internet") is not False or any(value is None for value in resources):
            return LEGACY_PAPER_PROTOCOL_ISSUES[task_id]
    elif task_id.startswith("schemathesis_"):
        if cell.get("internet") is not False or cell.get("submission_cooldown") != 216:
            return LEGACY_PAPER_PROTOCOL_ISSUES.get(task_id)
    return None


def comparison_record(
    cell: dict[str, Any],
    paper_tasks: dict[str, Any],
    local_fast_checkpoints: dict[str, Any] | None = None,
) -> dict[str, Any]:
    best = cell.get("best") or {}
    observations = cell.get("observations", [])
    evaluator_calls = sum(int(item.get("evaluator_calls") or 0) for item in observations)
    runtime = sum(float(item.get("runtime_seconds") or 0) for item in observations)
    input_tokens = 0
    output_tokens = 0
    coverage: set[str] = set()
    for item in observations:
        usage = item.get("codex_usage") or {}
        tokens = usage.get("tokens") or {}
        input_tokens += int(tokens.get("input_tokens") or 0)
        output_tokens += int(tokens.get("output_tokens") or 0)
        if usage.get("coverage"):
            coverage.add(str(usage["coverage"]))
        annotator = (item.get("goal_plus") or {}).get(
            "evidence_annotator_usage"
        ) or {}
        input_tokens += int(annotator.get("input_tokens") or 0)
        output_tokens += int(annotator.get("output_tokens") or 0)
        if annotator.get("tasks"):
            coverage.add(str(annotator.get("coverage")))

    normalized = best.get("edgebench_score")
    official = best.get("official_comparison") or {}
    checkpoint_hours = official.get("checkpoint_hours")
    same_budget_score = (official.get("references") or {}).get("GPT-5.5")
    same_budget_delta = (
        float(normalized) - float(same_budget_score)
        if normalized is not None and same_budget_score is not None
        else None
    )
    paper_score = paper_tasks[cell["task_id"]]
    paper_delta = (
        float(normalized) - float(paper_score["mean"])
        if normalized is not None
        else None
    )
    local_fast_checkpoints = local_fast_checkpoints or {}
    local_half = (
        local_fast_checkpoints.get("0.5h", {}).get("tasks", {}).get(cell["task_id"], {})
    )
    local_one = (
        local_fast_checkpoints.get("1h", {}).get("tasks", {}).get(cell["task_id"], {})
    )
    local_half_score = local_half.get("edgebench_score")
    local_one_score = local_one.get("edgebench_score")
    local_half_delta = (
        float(normalized) - float(local_half_score)
        if normalized is not None and local_half_score is not None
        else None
    )
    local_one_delta = (
        float(normalized) - float(local_one_score)
        if normalized is not None and local_one_score is not None
        else None
    )

    if normalized is None:
        issue_marker = "MISSING_CURRENT"
    else:
        known_issue = cell.get("known_protocol_issue")
        if known_issue is None and cell.get("protocol_classification") is None:
            known_issue = LEGACY_PAPER_PROTOCOL_ISSUES.get(cell["task_id"])
        if known_issue:
            issue_marker = f"KNOWN_PROTOCOL: {known_issue}"
        elif same_budget_delta is None:
            issue_marker = "MISSING_SAME_BUDGET_REFERENCE"
        elif same_budget_delta >= PAPER_LARGE_GAP_THRESHOLD_PP:
            issue_marker = "REVIEW_HIGH"
        elif same_budget_delta <= -PAPER_LARGE_GAP_THRESHOLD_PP:
            issue_marker = "REVIEW_LOW"
        else:
            issue_marker = None

    return {
        "Task": cell["task_id"],
        "Method": cell["method"],
        "Model": cell.get("model"),
        "Reasoning": cell.get("reasoning_effort"),
        "Current budget (h)": float(cell["wall_time_seconds"]) / 3600.0,
        "T (s)": cell["wall_time_seconds"],
        "K": cell["live_search_concurrency"],
        "Outer trajectories": cell["completed_trajectories"],
        "Valid trajectories": cell["valid_trajectories"],
        "Best raw": best.get("raw_score"),
        "Current EdgeBench 0-100": normalized,
        "Local <=0.5h best": local_half_score,
        "Delta vs local <=0.5h (pp)": local_half_delta,
        "Local <=1h best": local_one_score,
        "Delta vs local <=1h (pp)": local_one_delta,
        "GPT-5.5 checkpoint (h)": checkpoint_hours,
        "GPT-5.5 same-budget": same_budget_score,
        "Delta vs same-budget (pp)": same_budget_delta,
        "Paper Codex + GPT-5.5 @12h mean": float(paper_score["mean"]),
        "Paper sample stddev": (
            float(paper_score["sample_stddev"])
            if paper_score.get("sample_stddev") is not None
            else None
        ),
        "Delta vs paper 12h (pp)": paper_delta,
        "Evaluator calls": evaluator_calls,
        "Runtime (s)": runtime,
        "Input tokens": input_tokens,
        "Output tokens": output_tokens,
        "Usage coverage": ", ".join(sorted(coverage)) or "unavailable",
        "Protocol classification": cell.get("protocol_classification"),
        "Official comparable": cell.get("official_edgebench_comparable", False),
        "Issue marker": issue_marker,
    }


def style_header(row: Iterable[Any]) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in row:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def write_comparison_workbook(payload: dict[str, Any], destination: Path) -> None:
    paper = payload["paper_reference"]
    paper_source = paper["source"]
    paper_contract = paper["reference"]
    paper_tasks = paper["tasks"]
    local_fast_reference = payload.get("local_fast_reference") or {}
    local_fast_contract = local_fast_reference.get("reference") or {}
    local_fast_checkpoints = local_fast_reference.get("checkpoints") or {}
    cells = payload["cells"]
    models = sorted({str(cell.get("model") or "unknown") for cell in cells})
    reasoning_levels = sorted(
        {str(cell.get("reasoning_effort") or "unspecified") for cell in cells}
    )
    wall_times = sorted({int(cell["wall_time_seconds"]) for cell in cells})
    valid_cells = sum(1 for cell in cells if int(cell.get("valid_trajectories") or 0) > 0)
    protocol_evidence = sum(
        1 for cell in cells if cell.get("protocol_classification") is not None
    )
    records = [
        comparison_record(cell, paper_tasks, local_fast_checkpoints) for cell in cells
    ]
    if not records:
        raise ValueError("comparison workbook requires at least one cell")
    workbook = Workbook()
    overview = workbook.active
    overview.title = "Overview"
    overview_rows = [
        ("Field", "Value"),
        ("Campaign", payload["campaign_id"]),
        ("Finalized at", payload.get("finalized_at")),
        ("Matched protocol", payload["matched_protocol"]),
        ("Models", ", ".join(models)),
        ("Reasoning", ", ".join(reasoning_levels)),
        ("Wall budgets (s)", ", ".join(str(value) for value in wall_times)),
        ("Cells with valid score", f"{valid_cells}/{len(cells)}"),
        ("Cells with protocol evidence", f"{protocol_evidence}/{len(cells)}"),
        ("EdgeBench commit", payload.get("edgebench_commit")),
        ("Goal Plus commit", payload.get("goal_plus_commit")),
        ("Dataset revision", payload.get("dataset_revision")),
        ("Same-budget reference", "Official EdgeBench GPT-5.5 checkpoint from each result"),
        ("Issue rule", f"Known protocol issue first; otherwise |same-budget delta| >= {PAPER_LARGE_GAP_THRESHOLD_PP:g} pp"),
        (
            "Local fast reference",
            local_fast_contract.get("label") or "Not included",
        ),
        (
            "Local fast coverage",
            (
                "; ".join(
                    f"<={checkpoint.get('boundary_hours'):g}h: "
                    f"{checkpoint.get('available_count')}/"
                    f"{local_fast_reference.get('task_count')}"
                    for checkpoint in local_fast_checkpoints.values()
                )
                if local_fast_reference
                else "Not included"
            ),
        ),
        (
            "Local fast selection",
            local_fast_contract.get("selection") or "Not included",
        ),
        ("Paper reference role", "12h diagnostic reference; not an apples-to-apples leaderboard comparison"),
        ("Paper agent + model", f"{paper_contract['agent']} + {paper_contract['model']}"),
        ("Paper budget (h)", paper_contract["budget_hours"]),
        ("Paper scheduled runs", paper_contract["scheduled_runs"]),
        ("Paper arXiv", paper_source["arxiv_id"]),
        ("Paper TeX source", paper_source["source_file"]),
        ("Paper TeX SHA256", paper_source["source_file_sha256"]),
        ("Token note", "Zero with non-complete coverage means unavailable telemetry, not free usage"),
    ]
    for row in overview_rows:
        overview.append(row)
    style_header(overview[1])
    overview.freeze_panes = "A2"
    overview.column_dimensions["A"].width = 30
    overview.column_dimensions["B"].width = 100
    for row in overview.iter_rows(min_row=2):
        row[0].font = Font(bold=True)
        row[1].alignment = Alignment(vertical="top", wrap_text=True)

    results = workbook.create_sheet("Results")
    headers = list(records[0])
    results.append(headers)
    for record in records:
        results.append([record[header] for header in headers])
    style_header(results[1])
    results.freeze_panes = "A2"
    if records:
        table = Table(displayName="EdgeBenchResults", ref=f"A1:{get_column_letter(len(headers))}{len(records) + 1}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False,
        )
        results.add_table(table)
    widths = {
        "Task": 38, "Method": 18, "Model": 20, "Reasoning": 12,
        "Usage coverage": 20, "Protocol classification": 24, "Issue marker": 58,
    }
    for index, header in enumerate(headers, start=1):
        results.column_dimensions[get_column_letter(index)].width = widths.get(header, 18)
    numeric_headers = {
        "Best raw", "Current EdgeBench 0-100", "Local <=0.5h best",
        "Delta vs local <=0.5h (pp)", "Local <=1h best",
        "Delta vs local <=1h (pp)", "GPT-5.5 same-budget",
        "Delta vs same-budget (pp)", "Paper Codex + GPT-5.5 @12h mean",
        "Paper sample stddev", "Delta vs paper 12h (pp)", "Runtime (s)",
    }
    for header in numeric_headers:
        column = headers.index(header) + 1
        for row in range(2, len(records) + 2):
            results.cell(row=row, column=column).number_format = "0.00"
    issue_column = headers.index("Issue marker") + 1
    issue_letter = get_column_letter(issue_column)
    issue_range = f"{issue_letter}2:{issue_letter}{len(records) + 1}"
    if records:
        results.conditional_formatting.add(
            issue_range,
            FormulaRule(formula=[f'LEFT({issue_letter}2,6)="REVIEW"'], fill=PatternFill("solid", fgColor="FFF2CC")),
        )
        results.conditional_formatting.add(
            issue_range,
            FormulaRule(formula=[f'LEFT({issue_letter}2,5)="KNOWN"'], fill=PatternFill("solid", fgColor="F4CCCC")),
        )
        results.conditional_formatting.add(
            issue_range,
            FormulaRule(formula=[f'LEFT({issue_letter}2,7)="MISSING"'], fill=PatternFill("solid", fgColor="F4CCCC")),
        )

    if local_fast_reference:
        local_fast = workbook.create_sheet("Local Fast")
        local_fast_headers = [
            "Boundary <= (h)", "Task", "Status", "Selected checkpoint (h)",
            "Raw score", "EdgeBench 0-100", "Model", "Reasoning", "Campaign",
            "Protocol classification", "Normalization", "Best round",
            "Evidence source",
        ]
        local_fast.append(local_fast_headers)
        for checkpoint in local_fast_checkpoints.values():
            boundary = checkpoint.get("boundary_hours")
            for task_id, record in sorted((checkpoint.get("tasks") or {}).items()):
                local_fast.append(
                    [
                        boundary, task_id, "available", record.get("checkpoint_hours"),
                        record.get("raw_score"), record.get("edgebench_score"),
                        record.get("model"), record.get("reasoning_effort"),
                        record.get("campaign_id"),
                        record.get("protocol_classification")
                        or "legacy development evidence",
                        record.get("normalization_source"), record.get("best_round"),
                        record.get("source"),
                    ]
                )
            for task_id, attempts in sorted(
                (checkpoint.get("missing_tasks") or {}).items()
            ):
                reasons = sorted(
                    {
                        str(item.get("reason") or item.get("status") or "unavailable")
                        for item in attempts
                    }
                )
                local_fast.append(
                    [boundary, task_id, "missing", None, None, None, None, None,
                     None, None, None, None, "; ".join(reasons)]
                )
        style_header(local_fast[1])
        local_fast.freeze_panes = "A2"
        local_fast.auto_filter.ref = f"A1:M{local_fast.max_row}"
        local_fast_widths = (16, 38, 12, 22, 18, 18, 20, 12, 65, 38, 24, 18, 100)
        for column, width in enumerate(local_fast_widths, start=1):
            local_fast.column_dimensions[get_column_letter(column)].width = width
        for row in local_fast.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    protocol = workbook.create_sheet("Protocol")
    protocol_headers = [
        "Task", "Method", "Classification", "Official comparable",
        "Known protocol issue", "Protocol diff",
    ]
    protocol.append(protocol_headers)
    for cell in cells:
        protocol.append(
            [
                cell["task_id"], cell["method"], cell.get("protocol_classification"),
                cell.get("official_edgebench_comparable", False),
                cell.get("known_protocol_issue"),
                json.dumps(cell.get("protocol_diff") or [], ensure_ascii=False, sort_keys=True),
            ]
        )
    style_header(protocol[1])
    protocol.freeze_panes = "A2"
    protocol.auto_filter.ref = f"A1:F{len(cells) + 1}"
    for column, width in enumerate((38, 18, 24, 20, 58, 100), start=1):
        protocol.column_dimensions[get_column_letter(column)].width = width
    for row in protocol.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)


def finalize_campaign(
    destination: Path, local_fast_reference_path: Path | None = None
) -> dict[str, Any]:
    campaign = read_json(destination / "campaign.json")
    paper_reference = load_paper_reference()
    missing_paper_tasks = sorted(
        set(campaign["task_ids"]) - set(paper_reference["tasks"])
    )
    if missing_paper_tasks:
        raise ValueError(
            "paper GPT-5.5 reference is missing campaign tasks: "
            + ", ".join(missing_paper_tasks)
        )
    summaries: list[dict[str, Any]] = []
    for item in campaign["cells"]:
        cell_path = destination / "cells" / item["cell_id"]
        cell = read_json(cell_path / "cell.json")
        summaries.append(summarize_cell(destination, cell))
    protocol_fields = {
        (
            summary["task_id"],
            summary["model"],
            summary["reasoning_effort"],
            summary["wall_time_seconds"],
            summary["live_search_concurrency"],
        )
        for summary in summaries
    }
    payload = {
        "schema_version": 1,
        "campaign_id": campaign["campaign_id"],
        "matched_protocol": len(protocol_fields) == len(set(campaign["task_ids"])),
        "edgebench_commit": campaign.get("edgebench_commit"),
        "goal_plus_commit": campaign.get("goal_plus_commit"),
        "dataset_revision": campaign.get("dataset_revision"),
        "wall_time_seconds": campaign.get("wall_time_seconds"),
        "live_search_concurrency": campaign.get("concurrency"),
        "cell_concurrency": campaign.get("cell_concurrency"),
        "paper_reference": paper_reference,
        "cells": summaries,
        "finalized_at": utc_now(),
    }
    if local_fast_reference_path is not None:
        payload["local_fast_reference"] = load_local_fast_reference(
            local_fast_reference_path
        )
    write_json(destination / "comparison.json", payload)
    write_comparison_workbook(
        payload, destination / f"{payload['campaign_id']}.xlsx"
    )
    return payload


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser()
    subparsers = parser.add_subparsers(dest="command", required=True)

    for name in ("provision", "doctor"):
        child = subparsers.add_parser(name)
        child.add_argument("--profile", default="vliw-smoke")
        if name == "doctor":
            child.add_argument("--output", type=Path)

    prepare_parser = subparsers.add_parser("prepare")
    prepare_parser.add_argument("--profile", default="vliw-smoke")
    prepare_parser.add_argument("--campaign-id")
    prepare_parser.add_argument("--method", action="append", choices=sorted(METHODS))
    prepare_parser.add_argument("--model")
    prepare_parser.add_argument("--reasoning-effort")
    prepare_parser.add_argument("--wall-time-seconds", type=int)
    prepare_parser.add_argument("--concurrency", type=int)
    prepare_parser.add_argument("--cell-concurrency", type=int)

    for name in ("run", "status", "stop", "finalize", "_execute"):
        child = subparsers.add_parser(name)
        child.add_argument("--campaign", required=True)
        if name == "run":
            child.add_argument("--detach", action="store_true")
        elif name == "status":
            child.add_argument("--json", action="store_true")
        elif name == "stop":
            child.add_argument("--wait-seconds", type=int, default=10)
        elif name == "finalize":
            child.add_argument("--local-fast-reference", type=Path)
    return parser


def main() -> int:
    configure_temp_environment()
    args = build_parser().parse_args()
    if args.command in {"provision", "doctor", "prepare"}:
        _, profile = load_profile(args.profile)
        if args.command == "provision":
            return provision(profile)
        if args.command == "doctor":
            return doctor(profile, output=args.output)
        prepare(args, profile)
        return 0

    destination = campaign_dir(args.campaign)
    if args.command == "run":
        return launch(destination, detach=args.detach)
    if args.command == "_execute":
        return execute_campaign(destination)
    if args.command == "status":
        return print_status(destination, as_json=args.json)
    if args.command == "stop":
        return stop_campaign(destination, wait_seconds=args.wait_seconds)
    if args.command == "finalize":
        payload = finalize_campaign(destination, args.local_fast_reference)
        print(json.dumps(payload, indent=2, ensure_ascii=False))
        return 0
    raise AssertionError(args.command)


if __name__ == "__main__":
    raise SystemExit(main())
