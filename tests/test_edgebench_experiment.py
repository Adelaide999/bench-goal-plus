from __future__ import annotations

import importlib.util
import io
import json
import sys
import tarfile
import time
import unittest
from dataclasses import replace
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SPEC = importlib.util.spec_from_file_location(
    "edgebench_experiment",
    ROOT / "experiments" / "edgebench" / "experiment.py",
)
assert SPEC and SPEC.loader
EDGE = importlib.util.module_from_spec(SPEC)
sys.modules[SPEC.name] = EDGE
SPEC.loader.exec_module(EDGE)

from experiments.edgebench.controller import io as EDGE_IO
from experiments.edgebench.controller import evidence as EDGE_EVIDENCE
from experiments.edgebench.controller import reporting as EDGE_REPORTING
from experiments.edgebench.controller import runtime as EDGE_RUNTIME


class EdgeBenchExperimentTest(unittest.TestCase):
    def test_compatibility_entrypoint_stays_thin(self) -> None:
        entrypoint = ROOT / "experiments" / "edgebench" / "experiment.py"
        source = entrypoint.read_text(encoding="utf-8")

        self.assertLessEqual(len(source.splitlines()), 200)
        for implementation in (
            "def doctor_payload(",
            "def execute_campaign(",
            "def finalize_campaign(",
            "def prepare(",
        ):
            self.assertNotIn(implementation, source)
        for module in (
            "cli.py",
            "context.py",
            "environment.py",
            "evidence.py",
            "io.py",
            "preparation.py",
            "profiles.py",
            "reporting.py",
            "runtime.py",
        ):
            self.assertTrue(
                (ROOT / "experiments" / "edgebench" / "controller" / module).is_file()
            )

    def test_live_goal_plus_status_uses_pi_runtime_snapshot(self) -> None:
        task_run = self.temp / "task-run"
        task_run.mkdir(parents=True)
        (task_run / "goal-plus-live-status.json").write_text(
            json.dumps(
                {
                    "captured_at": "2026-07-31T10:40:00Z",
                    "candidate_ids": ["c001", "c002"],
                    "candidate_count": 2,
                    "worker_sessions": [
                        {
                            "agent_session_id": "agent_001",
                            "candidate_id": "c001",
                            "host": "pi",
                            "verifier_runs": 4,
                        },
                        {
                            "agent_session_id": "agent_002",
                            "candidate_id": "c002",
                            "host": "pi",
                            "verifier_runs": 7,
                        },
                    ],
                    "agent_session_count": 2,
                    "bound_worker_handles": [
                        {
                            "agent_session_id": "agent_001",
                            "host": "pi-rpc",
                            "external_id": "agent_001",
                        },
                        {
                            "agent_session_id": "agent_002",
                            "host": "pi-rpc",
                            "external_id": "agent_002",
                        },
                    ],
                    "actual_worker_launch_count": 2,
                    "verifier_ledger": [
                        {"candidate_id": "c001", "iteration": index}
                        for index in range(1, 5)
                    ]
                    + [
                        {"candidate_id": "c002", "iteration": index}
                        for index in range(1, 8)
                    ],
                    "worker_verifier_runs": 11,
                    "verifier_candidate_ids": ["c001", "c002"],
                    "selected_candidate_ids": ["c001"],
                    "promoted_candidate_ids": ["c001"],
                    "goal_statuses": [
                        {"goal_plus_id": "gp_0001", "status": "complete"}
                    ],
                    "terminal_ready": True,
                }
            ),
            encoding="utf-8",
        )

        status = EDGE_EVIDENCE.live_goal_plus_status(
            self.temp,
            {
                "state": "running",
                "started_at": "2026-07-31T10:35:00Z",
                "wall_time_seconds": 600,
                "goal_plus_finalization_grace_seconds": 120,
                "task_id": "vliw_kernel_optimization",
                "sforge_run_id": "pi-live-status-test",
            },
            task_run,
        )

        self.assertEqual(status["candidate_count"], 2)
        self.assertEqual(status["agent_session_count"], 2)
        self.assertEqual(status["actual_worker_launch_count"], 2)
        self.assertEqual(status["worker_verifier_runs"], 11)
        self.assertEqual(status["promoted_candidate_ids"], ["c001"])
        self.assertTrue(status["terminal_ready"])
        self.assertEqual(
            status["state_sources"], ["goal-plus-live-status.json"]
        )

    def setUp(self) -> None:
        self.temp = (
            EDGE.ensure_temp_root("test-edgebench-experiment")
            / f"{self._testMethodName}-{time.time_ns()}"
        )
        self.temp.mkdir(parents=True)
        self.original_paths = EDGE.current_paths()
        self.test_paths = replace(
            self.original_paths,
            edge_root=self.temp / "edgebench",
            goal_plus_root=self.temp / "goal-plus",
            tasks_dir=self.temp / "edgebench" / "tasks",
            runs_root=self.temp / "runs",
        )
        EDGE.set_paths(self.test_paths)
        self.test_paths.tasks_dir.mkdir(parents=True)
        self.test_paths.goal_plus_root.mkdir(parents=True)
        (self.test_paths.tasks_dir / "vliw_kernel_optimization.json").write_text(
            json.dumps(
                {
                    "task_id": "vliw_kernel_optimization",
                    "internet": False,
                    "work": {
                        "agent_query": "Optimize solution.py.",
                        "image_tag": "work123",
                    },
                    "judge": {
                        "image_tag": "judge123",
                        "score_direction": "minimize",
                    },
                }
            ),
            encoding="utf-8",
        )

    def tearDown(self) -> None:
        EDGE.set_paths(self.original_paths)

    def profile(self) -> dict:
        _, profile = EDGE.load_profile("vliw-smoke")
        return profile

    def test_full_codex_profile_covers_all_public_tasks(self) -> None:
        _, profile = EDGE.load_profile("full-codex-2h")

        self.assertEqual(len(profile["task_ids"]), 51)
        self.assertEqual(len(set(profile["task_ids"])), 51)
        self.assertEqual(profile["methods"], ["plain-codex"])
        self.assertEqual(profile["model"], "gpt-5.6-sol")
        self.assertEqual(profile["reasoning_effort"], "medium")
        self.assertEqual(profile["wall_time_seconds"], 7200)
        self.assertEqual(profile["concurrency"], 1)
        self.assertEqual(profile["cell_concurrency"], 2)
        self.assertNotIn("work_cpu_limit", profile)
        self.assertNotIn("judge_cpu_limit", profile)

    def test_vliw_codex_local_smoke_is_explicit_and_reproducible(self) -> None:
        _, profile = EDGE.load_profile("vliw-codex-sol-medium-local-smoke")

        self.assertEqual(profile["task_ids"], ["vliw_kernel_optimization"])
        self.assertEqual(profile["methods"], ["plain-codex"])
        self.assertEqual(profile["model"], "gpt-5.6-sol")
        self.assertEqual(profile["reasoning_effort"], "medium")
        self.assertEqual(profile["wall_time_seconds"], 300)
        self.assertEqual(profile["concurrency"], 1)
        self.assertEqual(profile["cell_concurrency"], 1)
        self.assertEqual(
            profile["protocol_overrides"],
            {"eval_interval": 60, "internet": True},
        )
        self.assertTrue(profile["protocol_override_reasons"]["eval_interval"])
        self.assertTrue(profile["protocol_override_reasons"]["internet"])

    def test_pi_profiles_use_canonical_method_names_and_explicit_budgets(self) -> None:
        _, plain = EDGE.load_profile("vliw-pi-sol-medium-local-smoke")
        _, goal_plus = EDGE.load_profile(
            "vliw-goal-plus-pi-sol-medium-local-smoke"
        )
        _, api_provider = EDGE.load_profile(
            "vliw-goal-plus-pi-glm-5-2-provider-1h-k2-c1"
        )
        _, zai_provider = EDGE.load_profile(
            "vliw-goal-plus-pi-zai-glm-5-2-1h-k2-c1"
        )

        self.assertEqual(EDGE.METHODS["plain-pi"]["agent"], "pi")
        self.assertEqual(EDGE.METHODS["goal-plus-pi"]["agent"], "pi-goal-plus")
        self.assertEqual(
            EDGE.METHODS["goal-plus-pi-provider"]["agent"],
            "pi-goal-plus-provider",
        )
        self.assertEqual(plain["methods"], ["plain-pi"])
        self.assertEqual(goal_plus["methods"], ["goal-plus-pi"])
        self.assertEqual(goal_plus["concurrency"], 2)
        self.assertEqual(goal_plus["worker_runtime_seconds"], 240)
        self.assertEqual(goal_plus["goal_plus_finalization_grace_seconds"], 120)
        self.assertEqual(
            api_provider["methods"], ["goal-plus-pi-provider"]
        )
        self.assertEqual(api_provider["model"], "glm-proxy/GLM-5.2")
        self.assertEqual(api_provider["pi_package_version"], "0.83.0")
        self.assertEqual(api_provider["wall_time_seconds"], 3600)
        self.assertEqual(api_provider["concurrency"], 2)
        self.assertEqual(api_provider["cell_concurrency"], 1)
        self.assertEqual(zai_provider["methods"], ["goal-plus-pi-provider"])
        self.assertEqual(zai_provider["model"], "zai/glm-5.2")
        self.assertEqual(zai_provider["pi_package_version"], "0.83.0")
        self.assertEqual(zai_provider["wall_time_seconds"], 3600)
        self.assertEqual(zai_provider["concurrency"], 2)
        self.assertEqual(zai_provider["cell_concurrency"], 1)

    def test_pi_provider_profile_requires_qualified_provider_model(self) -> None:
        _, profile = EDGE.load_profile(
            "vliw-goal-plus-pi-glm-5-2-provider-1h-k2-c1"
        )
        profile["model"] = "GLM-5.2"
        path = self.temp / "invalid-pi-provider-model.json"
        path.write_text(json.dumps(profile), encoding="utf-8")

        with self.assertRaisesRegex(ValueError, "PROVIDER/MODEL"):
            EDGE.load_profile(path)

    def test_profile_rejects_invalid_eval_interval_override(self) -> None:
        _, profile = EDGE.load_profile("vliw-codex-sol-medium-local-smoke")
        for value in (0, "60"):
            with self.subTest(value=value):
                profile["protocol_overrides"]["eval_interval"] = value
                path = self.temp / f"invalid-eval-interval-{value}.json"
                path.write_text(json.dumps(profile), encoding="utf-8")

                with self.assertRaisesRegex(
                    ValueError, "eval_interval override must be a positive integer"
                ):
                    EDGE.load_profile(path)

    def test_vliw_glm_profile_pins_claude_thinking_effort_and_budget(self) -> None:
        _, profile = EDGE.load_profile("vliw-glm-5-2-high-20m-k1")

        self.assertEqual(profile["task_ids"], ["vliw_kernel_optimization"])
        self.assertEqual(profile["methods"], ["plain-claude"])
        self.assertEqual(profile["model"], "glm-5.2")
        self.assertEqual(profile["thinking"], {"type": "enabled"})
        self.assertEqual(profile["reasoning_effort"], "high")
        self.assertEqual(profile["wall_time_seconds"], 1200)
        self.assertEqual(profile["concurrency"], 1)
        self.assertEqual(profile["cell_concurrency"], 1)

        _, none_profile = EDGE.load_profile("vliw-glm-5-2-none-20m-k1")
        self.assertEqual(none_profile["task_ids"], ["vliw_kernel_optimization"])
        self.assertEqual(none_profile["methods"], ["plain-claude"])
        self.assertEqual(none_profile["model"], "glm-5.2")
        self.assertEqual(none_profile["thinking"], {"type": "disabled"})
        self.assertEqual(none_profile["reasoning_effort"], "none")
        self.assertEqual(none_profile["wall_time_seconds"], 1200)
        self.assertEqual(none_profile["concurrency"], 1)

    def test_vliw_glm_51_profile_uses_official_adaptive_defaults(self) -> None:
        _, profile = EDGE.load_profile("vliw-glm-5-1-adaptive-2h-k1")

        self.assertEqual(profile["task_ids"], ["vliw_kernel_optimization"])
        self.assertEqual(profile["methods"], ["plain-claude"])
        self.assertEqual(profile["model"], "glm-5.1")
        self.assertEqual(profile["thinking"], {"type": "adaptive"})
        self.assertNotIn("reasoning_effort", profile)
        self.assertEqual(profile["claude_context_window_tokens"], 200000)
        self.assertEqual(profile["claude_autocompact_percent"], 80)
        self.assertEqual(profile["wall_time_seconds"], 7200)
        self.assertEqual(profile["concurrency"], 1)

    def test_protocol_regression_profile_targets_prior_failures(self) -> None:
        _, profile = EDGE.load_profile("protocol-regression-codex-2h")

        self.assertEqual(
            profile["task_ids"],
            [
                "borden_source_inversion",
                "exchange_core_throughput",
                "schemathesis_config_modernization",
                "schemathesis_datagen_pipeline",
                "schemathesis_reporting_observability",
                "anchorhead_text_adventure",
                "trinity_text_adventure",
                "tryst_text_adventure",
            ],
        )
        self.assertEqual(profile["methods"], ["plain-codex"])
        self.assertEqual(profile["model"], "gpt-5.6-sol")
        self.assertEqual(profile["reasoning_effort"], "medium")
        self.assertEqual(profile["wall_time_seconds"], 7200)
        self.assertEqual(profile["concurrency"], 1)
        self.assertEqual(profile["cell_concurrency"], 2)

    def test_validation_regression_profile_targets_suspicious_legacy_cells(self) -> None:
        _, profile = EDGE.load_profile("validation-regression-codex-2h-c4")

        self.assertEqual(len(profile["task_ids"]), 17)
        self.assertEqual(len(set(profile["task_ids"])), 17)
        self.assertEqual(profile["task_ids"][-1], "integer_compression_codec")
        self.assertEqual(profile["methods"], ["plain-codex"])
        self.assertEqual(profile["model"], "gpt-5.6-sol")
        self.assertEqual(profile["reasoning_effort"], "medium")
        self.assertEqual(profile["wall_time_seconds"], 7200)
        self.assertEqual(profile["concurrency"], 1)
        self.assertEqual(profile["cell_concurrency"], 4)
        self.assertEqual(profile["judge_concurrency"], 1)

    def test_official_codex_protocol_covers_all_tasks_and_overrides(self) -> None:
        protocol = EDGE.load_official_codex_protocol()
        _, full_profile = EDGE.load_profile("full-codex-2h")

        self.assertEqual(len(protocol["tasks"]), 51)
        self.assertEqual(set(protocol["tasks"]), set(full_profile["task_ids"]))
        self.assertEqual(protocol["official_model"], "gpt-5.5")
        self.assertEqual(protocol["stagger_seconds"], 600)
        self.assertEqual(
            protocol["defaults"],
            {
                "backend": "k8s",
                "agent": "codex",
                "timeout": 43200,
                "eval_interval": 1800,
                "submission_cooldown": 120,
                "work_cpu_limit": 4,
                "work_mem_limit": "16g",
                "judge_cpu_limit": 4,
                "judge_mem_limit": "8g",
            },
        )
        self.assertEqual(
            protocol["tasks"]["dabic_gravity_inversion"],
            {"submission_cooldown": 2160},
        )
        self.assertEqual(
            protocol["tasks"]["schemathesis_config_modernization"],
            {"submission_cooldown": 216},
        )
        self.assertEqual(
            protocol["tasks"]["anchorhead_text_adventure"],
            {"submission_cooldown": 0},
        )
        self.assertEqual(
            protocol["tasks"]["graph_node_classification"],
            {"judge_mem_limit": "16g"},
        )
        self.assertEqual(
            protocol["tasks"]["lean_analysis_proofs"],
            {
                "work_cpu_limit": 8,
                "work_mem_limit": "16g",
                "judge_cpu_limit": 8,
                "judge_mem_limit": "16g",
            },
        )
        self.assertEqual(protocol["tasks"]["smt_solver"]["work_cpu_limit"], 16)
        self.assertEqual(protocol["tasks"]["smt_solver"]["judge_cpu_limit"], 16)
        serialized = json.dumps(protocol)
        self.assertNotIn("sk-xxxx", serialized)
        self.assertNotIn("SFORGE_K8S_IMAGE_REGISTRY", serialized)
        self.assertNotIn("api_key", serialized)

    def test_official_task_protocol_uses_task_owned_internet(self) -> None:
        protocol = EDGE.load_official_codex_protocol()

        isolated = EDGE.official_task_protocol(
            protocol,
            "vliw_kernel_optimization",
            {"internet": False},
        )
        connected = EDGE.official_task_protocol(
            protocol,
            "college_english_exam_bank",
            {"internet": True},
        )

        self.assertFalse(isolated["internet"])
        self.assertTrue(connected["internet"])
        self.assertEqual(isolated["eval_interval"], 1800)
        self.assertEqual(isolated["submission_cooldown"], 120)
        self.assertEqual(isolated["work_cpu_limit"], 4)
        self.assertEqual(isolated["work_mem_limit"], "16g")
        self.assertEqual(isolated["judge_cpu_limit"], 4)
        self.assertEqual(isolated["judge_mem_limit"], "8g")
        self.assertFalse(isolated["disable_auto_eval"])
        self.assertFalse(isolated["disable_auto_resume"])
        self.assertFalse(isolated["disable_stop_hook"])

        explicit_lifecycle = {
            **protocol,
            "defaults": {
                **protocol["defaults"],
                "disable_auto_eval": True,
                "max_submissions": 5,
            },
        }
        explicit = EDGE.official_task_protocol(
            explicit_lifecycle,
            "vliw_kernel_optimization",
            {"internet": False},
        )
        self.assertTrue(explicit["disable_auto_eval"])
        self.assertEqual(explicit["max_submissions"], 5)

    def test_known_protocol_marker_depends_on_effective_cell_config(self) -> None:
        aligned = {
            "task_id": "schemathesis_config_modernization",
            "internet": False,
            "submission_cooldown": 216,
            "work_cpu_limit": 4,
            "work_mem_limit": "16g",
            "judge_cpu_limit": 4,
            "judge_mem_limit": "8g",
        }
        legacy = {**aligned, "internet": True, "submission_cooldown": None}

        self.assertIsNone(EDGE.paper_protocol_issue(aligned))
        self.assertIn("Internet access used", EDGE.paper_protocol_issue(legacy))

    def test_protocol_diff_rejects_resource_or_network_overrides(self) -> None:
        for field, official, effective in (
            ("work_cpu_limit", 4, 8),
            ("internet", False, True),
            ("submission_cooldown", 120, 0),
        ):
            with self.subTest(field=field):
                with self.assertRaisesRegex(
                    ValueError, "unsupported EdgeBench protocol override"
                ):
                    EDGE._protocol_diff(
                        official={field: official},
                        effective={field: effective},
                        reasons={field: "not permitted"},
                    )

    def test_protocol_diff_accepts_only_explicitly_allowed_network_override(self) -> None:
        diff = EDGE._protocol_diff(
            official={"internet": False},
            effective={"internet": True},
            reasons={"internet": "local development smoke"},
            allowed_fields=(
                EDGE.ALLOWED_PROTOCOL_OVERRIDE_FIELDS | {"internet"}
            ),
        )

        self.assertEqual(
            diff,
            [
                {
                    "field": "internet",
                    "official": False,
                    "effective": True,
                    "reason": "local development smoke",
                }
            ],
        )

    def test_paper_gpt55_reference_covers_profile_and_records_provenance(self) -> None:
        _, profile = EDGE.load_profile("full-codex-2h")
        paper = EDGE.load_paper_reference()

        self.assertEqual(set(paper["tasks"]), set(profile["task_ids"]))
        self.assertEqual(paper["reference"]["agent"], "Codex")
        self.assertEqual(paper["reference"]["model"], "GPT-5.5")
        self.assertEqual(paper["reference"]["budget_hours"], 12)
        self.assertEqual(paper["reference"]["scheduled_runs"], 3)
        self.assertEqual(paper["tasks"]["borden_source_inversion"]["mean"], 38.5)
        self.assertEqual(
            paper["tasks"]["vliw_kernel_optimization"]["sample_stddev"],
            1.9,
        )
        self.assertEqual(
            paper["source"]["source_archive_sha256"],
            "8193aeb41a3474690a40fac82e2ecbd53e651ab6b4759984b4c6845c04fbfd29",
        )

    def test_comparison_workbook_uses_same_budget_gap_for_issue_marker(self) -> None:
        paper = EDGE.load_paper_reference()
        payload = {
            "campaign_id": "comparison-test",
            "matched_protocol": True,
            "paper_reference": paper,
            "finalized_at": "2026-07-29T00:00:00+00:00",
            "local_fast_reference": {
                "schema_version": 2,
                "reference": {
                    "label": "Local Codex + gpt-5.6-sol inclusive checkpoints",
                    "selection": "strict local checkpoint",
                    "official_comparison": False,
                },
                "task_count": 2,
                "checkpoints": {
                    "0.5h": {
                        "boundary_hours": 0.5,
                        "boundary_seconds": 1800,
                        "available_count": 1,
                        "tasks": {
                            "portfolio_risk_calibration": {
                                "task_id": "portfolio_risk_calibration",
                                "checkpoint_hours": 0.5,
                                "checkpoint_seconds": 1800,
                                "raw_score": 19.83,
                                "edgebench_score": 19.83,
                                "model": "gpt-5.6-sol",
                                "reasoning_effort": "medium",
                                "campaign_id": "fast-campaign",
                                "source": "runs/evidence.json",
                            }
                        },
                        "missing_tasks": {
                            "borden_source_inversion": [
                                {"status": "no_scored_submission"}
                            ]
                        },
                    },
                    "1h": {
                        "boundary_hours": 1,
                        "boundary_seconds": 3600,
                        "available_count": 1,
                        "tasks": {
                            "portfolio_risk_calibration": {
                                "task_id": "portfolio_risk_calibration",
                                "checkpoint_hours": 1,
                                "checkpoint_seconds": 3600,
                                "raw_score": 30,
                                "edgebench_score": 30,
                                "model": "gpt-5.6-sol",
                                "reasoning_effort": "medium",
                                "campaign_id": "fast-campaign",
                                "source": "runs/evidence-1h.json",
                            }
                        },
                        "missing_tasks": {
                            "borden_source_inversion": [
                                {"status": "no_scored_submission"}
                            ]
                        },
                    }
                },
            },
            "cells": [
                {
                    "task_id": "borden_source_inversion",
                    "method": "plain-codex",
                    "model": "gpt-5.6-sol",
                    "reasoning_effort": "medium",
                    "wall_time_seconds": 7200,
                    "live_search_concurrency": 1,
                    "completed_trajectories": 1,
                    "valid_trajectories": 1,
                    "observations": [],
                    "best": {
                        "raw_score": 78.502,
                        "edgebench_score": 78.502,
                        "official_comparison": {
                            "checkpoint_hours": 2,
                            "references": {"GPT-5.5": 38.5},
                        },
                    },
                },
                {
                    "task_id": "portfolio_risk_calibration",
                    "method": "plain-codex",
                    "model": "gpt-5.6-sol",
                    "reasoning_effort": "medium",
                    "wall_time_seconds": 7200,
                    "live_search_concurrency": 1,
                    "completed_trajectories": 1,
                    "valid_trajectories": 1,
                    "observations": [],
                    "best": {
                        "raw_score": 44.97,
                        "edgebench_score": 44.97,
                        "official_comparison": {
                            "checkpoint_hours": 2,
                            "references": {"GPT-5.5": 17.3},
                        },
                    },
                },
            ],
        }
        destination = self.temp / "comparison.xlsx"

        EDGE.write_comparison_workbook(payload, destination)

        workbook = load_workbook(destination, data_only=True)
        self.assertEqual(
            workbook.sheetnames,
            ["Overview", "Results", "Local Fast", "Protocol"],
        )
        results = workbook["Results"]
        headers = [cell.value for cell in results[1]]
        rows = {
            row[headers.index("Task")].value: {
                header: row[index].value for index, header in enumerate(headers)
            }
            for row in results.iter_rows(min_row=2)
        }
        portfolio = rows["portfolio_risk_calibration"]
        self.assertEqual(portfolio["Current budget (h)"], 2)
        self.assertEqual(portfolio["T (s)"], 7200)
        self.assertEqual(portfolio["Current EdgeBench 0-100"], 44.97)
        self.assertEqual(portfolio["Local <=0.5h best"], 19.83)
        self.assertAlmostEqual(portfolio["Delta vs local <=0.5h (pp)"], 25.14)
        self.assertEqual(portfolio["Local <=1h best"], 30)
        self.assertAlmostEqual(portfolio["Delta vs local <=1h (pp)"], 14.97)
        self.assertEqual(portfolio["GPT-5.5 checkpoint (h)"], 2)
        self.assertEqual(portfolio["GPT-5.5 same-budget"], 17.3)
        self.assertAlmostEqual(portfolio["Delta vs same-budget (pp)"], 27.67)
        self.assertEqual(portfolio["Paper Codex + GPT-5.5 @12h mean"], 25.0)
        self.assertEqual(portfolio["Paper sample stddev"], 6.5)
        self.assertAlmostEqual(portfolio["Delta vs paper 12h (pp)"], 19.97)
        self.assertEqual(portfolio["Issue marker"], "REVIEW_HIGH")
        self.assertIn("KNOWN_PROTOCOL", rows["borden_source_inversion"]["Issue marker"])
        self.assertEqual(results.freeze_panes, "A2")
        self.assertEqual(len(results.tables), 1)
        overview_values = {row[0].value: row[1].value for row in workbook["Overview"]}
        self.assertIn("not an apples-to-apples", overview_values["Paper reference role"])
        self.assertEqual(
            overview_values["Local fast coverage"],
            "<=0.5h: 1/2; <=1h: 1/2",
        )
        local_fast_rows = list(workbook["Local Fast"].iter_rows(values_only=True))
        self.assertEqual(
            local_fast_rows[1][0:3],
            (0.5, "portfolio_risk_calibration", "available"),
        )
        self.assertEqual(
            local_fast_rows[2][0:3],
            (0.5, "borden_source_inversion", "missing"),
        )

    def test_score_task_run_counts_game_sessions_without_run_history(self) -> None:
        task_run = self.temp / "game-run"
        task_run.mkdir()
        EDGE.write_json(
            task_run / "final_result.json",
            {
                "runtime_seconds": 120,
                "total_rounds": 2,
                "agent_submissions": 2,
                "auto_submissions": 0,
                "resume_count": 0,
                "timed_out": True,
            },
        )
        EDGE.write_json(
            task_run / "game_history.json",
            {
                "entries": [
                    {"type": "game", "round": "game-1"},
                    {"type": "game", "round": "game-2"},
                ]
            },
        )
        original_run_capture = EDGE_IO.run_capture
        EDGE_IO.run_capture = lambda *_args, **_kwargs: {
            "returncode": 0,
            "stdout": json.dumps(
                {
                    "source": str(task_run / "final_result.json"),
                    "edgebench_score": 25,
                }
            ),
            "stderr": "",
        }
        try:
            observation = EDGE.score_task_run(
                task_run,
                {"model": "gpt-test", "wall_time_seconds": 120},
            )
        finally:
            EDGE_IO.run_capture = original_run_capture

        self.assertEqual(observation["evaluator_calls"], 2)

    def test_prepare_encodes_plain_outer_and_goal_plus_inner_concurrency(self) -> None:
        args = SimpleNamespace(
            method=None,
            wall_time_seconds=180,
            concurrency=2,
            model="gpt-test",
            reasoning_effort="high",
            campaign_id="unit-campaign",
        )

        destination = EDGE.prepare(args, self.profile())

        plain = json.loads(
            (
                destination
                / "cells"
                / "vliw_kernel_optimization--plain-codex"
                / "cell.json"
            ).read_text()
        )
        goal_plus = json.loads(
            (
                destination
                / "cells"
                / "vliw_kernel_optimization--goal-plus-codex"
                / "cell.json"
            ).read_text()
        )
        self.assertEqual(plain["outer_replicas"], 2)
        self.assertEqual(plain["inner_search_concurrency"], 0)
        self.assertEqual(goal_plus["outer_replicas"], 1)
        self.assertEqual(goal_plus["inner_search_concurrency"], 2)
        self.assertFalse(plain["internet"])
        self.assertEqual(plain["eval_interval_seconds"], 1800)
        self.assertEqual(plain["submission_cooldown"], 120)
        self.assertEqual(plain["work_cpu_limit"], 4)
        self.assertEqual(plain["work_mem_limit"], "16g")
        self.assertEqual(plain["judge_cpu_limit"], 4)
        self.assertEqual(plain["judge_mem_limit"], "8g")
        self.assertTrue(plain["auto_eval_enabled"])
        self.assertTrue(plain["auto_resume_enabled"])
        self.assertTrue(plain["stop_hook_enabled"])
        self.assertEqual(
            {item["field"] for item in plain["protocol_diff"]},
            {
                "attempts_per_task",
                "backend",
                "cell_concurrency",
                "judge_concurrency",
                "model",
                "reasoning_effort",
                "timeout",
            },
        )
        self.assertEqual(
            {item["field"] for item in goal_plus["protocol_diff"]},
            {
                "agent",
                "attempts_per_task",
                "backend",
                "cell_concurrency",
                "judge_concurrency",
                "model",
                "reasoning_effort",
                "timeout",
            },
        )
        self.assertEqual(
            plain["protocol_source"]["sha256"],
            EDGE.sha256_file(EDGE.OFFICIAL_CODEX_PROTOCOL_PATH),
        )
        self.assertFalse(plain["official_edgebench_comparable"])
        self.assertNotIn("sk-xxxx", json.dumps(plain))
        self.assertEqual(
            json.loads((destination / "profile.json").read_text())["cell_concurrency"],
            1,
        )
        self.assertFalse(any(path.name in {".gp", ".goal-plus"} for path in destination.rglob("*")))

    def test_prepare_applies_local_smoke_network_override_with_provenance(self) -> None:
        _, profile = EDGE.load_profile("vliw-codex-sol-medium-local-smoke")
        args = SimpleNamespace(
            method=None,
            wall_time_seconds=None,
            concurrency=None,
            cell_concurrency=None,
            model=None,
            reasoning_effort=None,
            campaign_id="unit-local-codex-smoke",
        )

        destination = EDGE.prepare(args, profile)
        cell = json.loads(
            (
                destination
                / "cells"
                / "vliw_kernel_optimization--plain-codex"
                / "cell.json"
            ).read_text()
        )
        command = EDGE.build_sforge_command(destination, cell)

        self.assertTrue(cell["internet"])
        self.assertEqual(cell["eval_interval_seconds"], 60)
        self.assertEqual(
            cell["internet_source"],
            "profiles/vliw-codex-sol-medium-local-smoke.protocol_overrides.internet",
        )
        self.assertTrue(
            {"eval_interval", "internet"}
            <= {item["field"] for item in cell["protocol_diff"]}
        )
        self.assertFalse(cell["official_edgebench_comparable"])
        self.assertEqual(
            command[command.index("--eval-interval") + 1],
            "60",
        )
        self.assertIn("--enable-internet", command)
        self.assertNotIn("--disable-internet", command)

    def test_prepare_encodes_plain_claude_api_and_thinking_contract(self) -> None:
        _, profile = EDGE.load_profile("vliw-glm-5-2-high-20m-k1")
        args = SimpleNamespace(
            method=None,
            wall_time_seconds=None,
            concurrency=None,
            cell_concurrency=None,
            model=None,
            reasoning_effort=None,
            campaign_id="unit-claude-campaign",
        )

        destination = EDGE.prepare(args, profile)
        cell = json.loads(
            (
                destination
                / "cells"
                / "vliw_kernel_optimization--plain-claude"
                / "cell.json"
            ).read_text()
        )

        self.assertEqual(cell["sforge_agent"], "claude-code")
        self.assertEqual(cell["api_protocol"], "anthropic")
        self.assertEqual(cell["thinking"], {"type": "enabled"})
        self.assertEqual(cell["reasoning_effort"], "high")
        self.assertEqual(cell["outer_replicas"], 1)
        self.assertEqual(cell["inner_search_concurrency"], 0)

    def test_prepare_preserves_adaptive_claude_without_effort(self) -> None:
        _, profile = EDGE.load_profile("vliw-glm-5-1-adaptive-2h-k1")
        args = SimpleNamespace(
            method=None,
            wall_time_seconds=None,
            concurrency=None,
            cell_concurrency=None,
            model=None,
            reasoning_effort=None,
            campaign_id="unit-claude-adaptive-campaign",
        )

        destination = EDGE.prepare(args, profile)
        cell = json.loads(
            (
                destination
                / "cells"
                / "vliw_kernel_optimization--plain-claude"
                / "cell.json"
            ).read_text()
        )

        self.assertEqual(cell["thinking"], {"type": "adaptive"})
        self.assertIsNone(cell["reasoning_effort"])
        self.assertEqual(cell["claude_context_window_tokens"], 200000)
        self.assertEqual(cell["claude_autocompact_percent"], 80)
        self.assertNotIn(
            "reasoning_effort",
            {item["field"] for item in cell["protocol_diff"]},
        )

    def test_cell_queue_limits_parallel_cells_and_continues_after_failure(self) -> None:
        destination = self.temp / "campaign"
        destination.mkdir()
        EDGE.write_json(destination / "controller.json", {"state": "running"})
        campaign = {
            "cells": [
                {"cell_id": cell_id, "task_id": cell_id}
                for cell_id in ("a", "b", "c")
            ]
        }
        started = []
        processes = []
        live = 0
        max_live = 0

        class FakeProcess:
            def __init__(self, pid):
                self.pid = pid
                self.done = False

            def poll(self):
                return 0 if self.done else None

        def fake_start(_destination, summary, **_kwargs):
            nonlocal live, max_live
            process = FakeProcess(100 + len(processes))
            processes.append(process)
            started.append(summary["cell_id"])
            live += 1
            max_live = max(max_live, live)
            return {
                "cell": {
                    "cell_id": summary["cell_id"],
                    "task_id": summary["task_id"],
                    "started_at": "now",
                },
                "process": process,
            }

        def fake_finish(_destination, running, *, stop_requested):
            nonlocal live
            self.assertFalse(stop_requested)
            live -= 1
            return 1 if running["cell"]["cell_id"] == "b" else 0

        sleeps = 0

        def fake_sleep(_seconds):
            nonlocal sleeps
            sleeps += 1
            if sleeps == 1:
                processes[0].done = True
            else:
                for process in processes:
                    process.done = True

        original_start = EDGE_RUNTIME.start_campaign_cell
        original_finish = EDGE_RUNTIME.finish_campaign_cell
        original_sleep = EDGE.time.sleep
        EDGE_RUNTIME.start_campaign_cell = fake_start
        EDGE_RUNTIME.finish_campaign_cell = fake_finish
        EDGE.time.sleep = fake_sleep
        try:
            returncode = EDGE.execute_cell_queue(
                destination,
                campaign,
                {"state": "running"},
                cell_concurrency=2,
                judge_container_url="http://judge",
                api_config={"api_key_source": None, "api_base_url_source": None},
                api_key=None,
                runtime_api_base_url=None,
                bridge_host=None,
                stop_requested=lambda: False,
            )
        finally:
            EDGE_RUNTIME.start_campaign_cell = original_start
            EDGE_RUNTIME.finish_campaign_cell = original_finish
            EDGE.time.sleep = original_sleep

        self.assertEqual(returncode, 1)
        self.assertEqual(started, ["a", "b", "c"])
        self.assertEqual(max_live, 2)
        self.assertEqual(
            json.loads((destination / "controller.json").read_text())[
                "active_children"
            ],
            {},
        )

    def test_cell_queue_stop_interrupts_active_cells_without_starting_more(self) -> None:
        destination = self.temp / "campaign"
        destination.mkdir()
        EDGE.write_json(destination / "controller.json", {"state": "running"})
        campaign = {
            "cells": [
                {"cell_id": cell_id, "task_id": cell_id}
                for cell_id in ("a", "b", "c")
            ]
        }
        started = []
        processes = []
        requested = False

        class FakeProcess:
            def __init__(self, pid):
                self.pid = pid
                self.done = False
                self.signals = []

            def poll(self):
                return 130 if self.done else None

            def send_signal(self, value):
                self.signals.append(value)
                self.done = True

        def fake_start(_destination, summary, **_kwargs):
            process = FakeProcess(200 + len(processes))
            processes.append(process)
            started.append(summary["cell_id"])
            return {
                "cell": {
                    "cell_id": summary["cell_id"],
                    "task_id": summary["task_id"],
                    "started_at": "now",
                },
                "process": process,
            }

        def fake_finish(_destination, _running, *, stop_requested):
            self.assertTrue(stop_requested)
            return 130

        def fake_sleep(_seconds):
            nonlocal requested
            requested = True

        original_start = EDGE_RUNTIME.start_campaign_cell
        original_finish = EDGE_RUNTIME.finish_campaign_cell
        original_sleep = EDGE.time.sleep
        EDGE_RUNTIME.start_campaign_cell = fake_start
        EDGE_RUNTIME.finish_campaign_cell = fake_finish
        EDGE.time.sleep = fake_sleep
        try:
            returncode = EDGE.execute_cell_queue(
                destination,
                campaign,
                {"state": "running"},
                cell_concurrency=2,
                judge_container_url="http://judge",
                api_config={"api_key_source": None, "api_base_url_source": None},
                api_key=None,
                runtime_api_base_url=None,
                bridge_host=None,
                stop_requested=lambda: requested,
            )
        finally:
            EDGE_RUNTIME.start_campaign_cell = original_start
            EDGE_RUNTIME.finish_campaign_cell = original_finish
            EDGE.time.sleep = original_sleep

        self.assertEqual(returncode, 130)
        self.assertEqual(started, ["a", "b"])
        self.assertEqual(
            [process.signals for process in processes],
            [[EDGE.signal.SIGINT], [EDGE.signal.SIGINT]],
        )

    def test_goal_plus_environment_uses_pinned_source_and_configured_k(self) -> None:
        cell = {
            "method": "goal-plus-codex",
            "model": "gpt-5.6-sol",
            "reasoning_effort": "xhigh",
            "inner_search_concurrency": 4,
            "worker_runtime_seconds": 600,
            "goal_plus_finalization_grace_seconds": 90,
        }

        env = EDGE.cell_environment(
            cell,
            api_key="runtime-key",
            api_base_url="http://192.0.2.10:45678/v1",
        )

        self.assertEqual(
            env["SFORGE_GOAL_PLUS_SOURCE_DIR"],
            str(EDGE.current_paths().goal_plus_root),
        )
        extra = dict(
            item.split("=", 1) for item in env["SFORGE_AGENT_EXTRA_ENV"].split(",")
        )
        self.assertEqual(extra["SFORGE_GOAL_PLUS_PARALLEL_NUM"], "4")
        self.assertEqual(extra["SFORGE_GOAL_PLUS_WORKER_RUNTIME_SECONDS"], "600")
        self.assertEqual(
            extra["SFORGE_GOAL_PLUS_FINALIZATION_GRACE_SECONDS"], "90"
        )
        self.assertEqual(
            extra["GOAL_PLUS_EVIDENCE_ANNOTATOR_MODEL"], "gpt-5.6-sol"
        )
        self.assertEqual(
            extra["GOAL_PLUS_EVIDENCE_ANNOTATOR_REASONING_EFFORT"], "xhigh"
        )
        self.assertEqual(
            extra["GOAL_PLUS_EVIDENCE_ANNOTATOR_BASE_URL"],
            "http://192.0.2.10:45678/v1",
        )
        self.assertEqual(
            extra["GOAL_PLUS_EVIDENCE_ANNOTATOR_API_KEY_ENV"],
            "SFORGE_AGENT_API_KEY",
        )
        for key in ("TMPDIR", "TMP", "TEMP"):
            self.assertTrue(Path(env[key]).is_relative_to(ROOT))

    def test_goal_plus_pi_environment_uses_the_same_runtime_contract(self) -> None:
        env = EDGE.cell_environment(
            {
                "method": "goal-plus-pi",
                "sforge_agent": "pi-goal-plus",
                "model": "gpt-5.6-sol",
                "reasoning_effort": "medium",
                "pi_package_version": "0.83.0",
                "internet": True,
                "inner_search_concurrency": 2,
                "worker_runtime_seconds": 240,
                "goal_plus_finalization_grace_seconds": 120,
            }
        )
        extra = dict(
            item.split("=", 1) for item in env["SFORGE_AGENT_EXTRA_ENV"].split(",")
        )

        self.assertEqual(
            env["SFORGE_GOAL_PLUS_SOURCE_DIR"],
            str(EDGE.current_paths().goal_plus_root),
        )
        self.assertEqual(extra["SFORGE_GOAL_PLUS_PARALLEL_NUM"], "2")
        self.assertEqual(extra["SFORGE_GOAL_PLUS_WORKER_RUNTIME_SECONDS"], "240")
        self.assertEqual(extra["SFORGE_PI_REASONING_EFFORT"], "medium")
        self.assertEqual(extra["SFORGE_PI_PACKAGE_VERSION"], "0.83.0")
        self.assertEqual(
            extra["SFORGE_GOAL_PLUS_FINALIZATION_GRACE_SECONDS"], "120"
        )

    def test_goal_plus_pi_provider_environment_uses_goal_plus_contract(self) -> None:
        env = EDGE.cell_environment(
            {
                "method": "goal-plus-pi-provider",
                "sforge_agent": "pi-goal-plus-provider",
                "model": "glm-proxy/GLM-5.2",
                "reasoning_effort": "high",
                "internet": True,
                "inner_search_concurrency": 2,
                "worker_runtime_seconds": 3300,
                "goal_plus_finalization_grace_seconds": 300,
            }
        )
        extra = dict(
            item.split("=", 1) for item in env["SFORGE_AGENT_EXTRA_ENV"].split(",")
        )

        self.assertEqual(extra["SFORGE_GOAL_PLUS_PARALLEL_NUM"], "2")
        self.assertEqual(extra["SFORGE_GOAL_PLUS_WORKER_RUNTIME_SECONDS"], "3300")
        self.assertEqual(extra["SFORGE_PI_REASONING_EFFORT"], "high")
        self.assertEqual(
            extra["GOAL_PLUS_EVIDENCE_ANNOTATOR_MODEL"],
            "glm-proxy/GLM-5.2",
        )

    def test_api_config_prefers_sforge_then_openai_then_codex(self) -> None:
        config = EDGE.resolve_agent_api_config(
            {
                "SFORGE_AGENT_API_KEY": "sforge-key",
                "SFORGE_AGENT_API_BASE_URL": "https://sforge.example/v1",
                "OPENAI_API_KEY": "openai-key",
                "OPENAI_BASE_URL": "https://openai.example/v1",
                "CODEX_API_KEY": "codex-key",
            }
        )

        self.assertEqual(config["api_key"], "sforge-key")
        self.assertEqual(config["api_key_source"], "SFORGE_AGENT_API_KEY")
        self.assertEqual(config["api_base_url"], "https://sforge.example/v1")

        fallback = EDGE.resolve_agent_api_config(
            {
                "OPENAI_API_KEY": "openai-key",
                "OPENAI_BASE_URL": "http://127.0.0.1:3788/v1",
                "CODEX_API_KEY": "codex-key",
            }
        )
        self.assertEqual(fallback["api_key"], "openai-key")
        self.assertEqual(fallback["api_key_source"], "OPENAI_API_KEY")
        self.assertEqual(fallback["api_base_url_source"], "OPENAI_BASE_URL")

        anthropic = EDGE.resolve_agent_api_config(
            {
                "ANTHROPIC_AUTH_TOKEN": "anthropic-token",
                "ANTHROPIC_BASE_URL": "https://anthropic.example",
                "OPENAI_API_KEY": "wrong-protocol-key",
            },
            protocol="anthropic",
        )
        self.assertEqual(anthropic["api_key"], "anthropic-token")
        self.assertEqual(anthropic["api_key_source"], "ANTHROPIC_AUTH_TOKEN")
        self.assertEqual(
            anthropic["api_base_url"], "https://anthropic.example"
        )
        self.assertEqual(
            EDGE.agent_api_probe_url(
                "https://anthropic.example/api/anthropic", "anthropic"
            ),
            "https://anthropic.example/api/anthropic/v1/messages",
        )

        provider = EDGE.resolve_agent_api_config(
            {
                "OPENAI_API_KEY": "must-not-be-used",
                "ANTHROPIC_API_KEY": "must-not-be-used",
            },
            protocol="pi-provider",
        )
        self.assertIsNone(provider["api_key"])
        self.assertIsNone(provider["api_base_url"])

    def test_pi_auth_requires_an_openai_codex_login(self) -> None:
        auth = self.temp / "pi-auth.json"
        auth.write_text(json.dumps({"other-provider": {}}))
        self.assertFalse(
            EDGE.resolve_pi_auth({"SFORGE_PI_AUTH_FILE": str(auth)})["valid"]
        )

        auth.write_text(json.dumps({"openai-codex": {"type": "oauth"}}))
        status = EDGE.resolve_pi_auth({"SFORGE_PI_AUTH_FILE": str(auth)})
        self.assertTrue(status["valid"])
        self.assertEqual(status["path"], auth)

    def test_pi_provider_validates_registry_model_and_credential_env(self) -> None:
        models = self.temp / "models.json"
        models.write_text(
            json.dumps(
                {
                    "providers": {
                        "glm-anthropic": {
                            "api": "anthropic-messages",
                            "apiKey": "${GLM_PROXY_API_KEY}",
                            "models": [{"id": "GLM-5.2"}],
                        },
                        "glm-openai": {
                            "api": "openai-completions",
                            "apiKey": "$GLM_PROXY_API_KEY",
                            "models": [{"id": "GLM-5.2"}],
                        }
                    }
                }
            )
        )
        env = {
            "SFORGE_PI_MODELS_FILE": str(models),
            "GLM_PROXY_API_KEY": "secret-value",
        }

        for provider in ("glm-anthropic", "glm-openai"):
            with self.subTest(provider=provider):
                status = EDGE.resolve_pi_provider(
                    f"{provider}/GLM-5.2", env
                )
                self.assertTrue(status["valid"])
                self.assertEqual(status["provider"], provider)
                self.assertEqual(status["model"], "GLM-5.2")
                self.assertEqual(status["credential_env"], "GLM_PROXY_API_KEY")
                self.assertNotIn("secret-value", json.dumps(status))

        missing = EDGE.resolve_pi_provider(
            "glm-anthropic/GLM-5.2",
            {"SFORGE_PI_MODELS_FILE": str(models)},
        )
        self.assertFalse(missing["valid"])
        self.assertEqual(missing["error"], "missing GLM_PROXY_API_KEY")

    def test_pi_provider_rejects_literal_or_bare_api_key(self) -> None:
        for api_key in ("GLM_PROXY_API_KEY", "literal-secret"):
            with self.subTest(api_key=api_key):
                models = self.temp / f"models-{len(api_key)}.json"
                models.write_text(
                    json.dumps(
                        {
                            "providers": {
                                "custom": {
                                    "apiKey": api_key,
                                    "models": [{"id": "model"}],
                                }
                            }
                        }
                    )
                )
                status = EDGE.resolve_pi_provider(
                    "custom/model", {"SFORGE_PI_MODELS_FILE": str(models)}
                )
                self.assertFalse(status["valid"])
                self.assertIn("$NAME", status["error"])
                self.assertNotIn("literal-secret", json.dumps(status))

    def test_pi_provider_requires_custom_api_key_reference(self) -> None:
        models = self.temp / "models-missing-api-key.json"
        models.write_text(
            json.dumps(
                {
                    "providers": {
                        "custom": {"models": [{"id": "model"}]}
                    }
                }
            )
        )

        status = EDGE.resolve_pi_provider(
            "custom/model", {"SFORGE_PI_MODELS_FILE": str(models)}
        )

        self.assertFalse(status["valid"])
        self.assertIn("apiKey as $NAME", status["error"])

    def test_pi_provider_uses_builtin_deepseek_environment(self) -> None:
        status = EDGE.resolve_pi_provider(
            "deepseek/deepseek-chat",
            {"DEEPSEEK_API_KEY": "secret-value"},
        )

        self.assertTrue(status["valid"])
        self.assertEqual(status["credential_env"], "DEEPSEEK_API_KEY")
        self.assertNotIn("secret-value", json.dumps(status))

    def test_pi_provider_prefers_anthropic_oauth_environment(self) -> None:
        status = EDGE.resolve_pi_provider(
            "anthropic/claude-sonnet-4-20250514",
            {
                "ANTHROPIC_OAUTH_TOKEN": "oauth-secret",
                "ANTHROPIC_API_KEY": "api-secret",
            },
        )

        self.assertTrue(status["valid"])
        self.assertEqual(status["credential_env"], "ANTHROPIC_OAUTH_TOKEN")
        self.assertNotIn("oauth-secret", json.dumps(status))
        self.assertNotIn("api-secret", json.dumps(status))

    def test_claude_environment_pins_effort_and_preserves_extra_env(self) -> None:
        previous = EDGE.os.environ.get("SFORGE_AGENT_EXTRA_ENV")
        EDGE.os.environ["SFORGE_AGENT_EXTRA_ENV"] = "EXISTING=value"
        try:
            env = EDGE.cell_environment(
                {
                    "method": "plain-claude",
                    "sforge_agent": "claude-code",
                    "reasoning_effort": "high",
                    "internet": False,
                }
            )
        finally:
            if previous is None:
                EDGE.os.environ.pop("SFORGE_AGENT_EXTRA_ENV", None)
            else:
                EDGE.os.environ["SFORGE_AGENT_EXTRA_ENV"] = previous

        extra = dict(
            item.split("=", 1)
            for item in env["SFORGE_AGENT_EXTRA_ENV"].split(",")
        )
        self.assertEqual(extra["EXISTING"], "value")
        self.assertEqual(extra["CLAUDE_CODE_EFFORT_LEVEL"], "high")
        self.assertEqual(extra["CLAUDE_CODE_ALWAYS_ENABLE_EFFORT"], "1")
        self.assertEqual(env["SFORGE_CLAUDE_CACHE_OPT"], "1")
        self.assertNotIn("SFORGE_CODEX_REASONING_EFFORT", env)

    def test_claude_none_environment_disables_thinking_and_removes_effort(self) -> None:
        keys = (
            "SFORGE_AGENT_EXTRA_ENV",
            "CLAUDE_CODE_EFFORT_LEVEL",
            "CLAUDE_CODE_ALWAYS_ENABLE_EFFORT",
        )
        previous = {key: EDGE.os.environ.get(key) for key in keys}
        EDGE.os.environ["SFORGE_AGENT_EXTRA_ENV"] = (
            "EXISTING=value,CLAUDE_CODE_EFFORT_LEVEL=high,"
            "CLAUDE_CODE_ALWAYS_ENABLE_EFFORT=1"
        )
        EDGE.os.environ["CLAUDE_CODE_EFFORT_LEVEL"] = "high"
        EDGE.os.environ["CLAUDE_CODE_ALWAYS_ENABLE_EFFORT"] = "1"
        try:
            env = EDGE.cell_environment(
                {
                    "method": "plain-claude",
                    "sforge_agent": "claude-code",
                    "reasoning_effort": "none",
                    "internet": False,
                }
            )
        finally:
            for key, value in previous.items():
                if value is None:
                    EDGE.os.environ.pop(key, None)
                else:
                    EDGE.os.environ[key] = value

        extra = dict(
            item.split("=", 1)
            for item in env["SFORGE_AGENT_EXTRA_ENV"].split(",")
        )
        self.assertEqual(extra["EXISTING"], "value")
        self.assertEqual(extra["MAX_THINKING_TOKENS"], "0")
        self.assertEqual(extra["CLAUDE_CODE_DISABLE_THINKING"], "1")
        self.assertNotIn("CLAUDE_CODE_EFFORT_LEVEL", extra)
        self.assertNotIn("CLAUDE_CODE_ALWAYS_ENABLE_EFFORT", extra)
        self.assertNotIn("CLAUDE_CODE_EFFORT_LEVEL", env)
        self.assertNotIn("CLAUDE_CODE_ALWAYS_ENABLE_EFFORT", env)

    def test_claude_adaptive_environment_removes_fixed_thinking_controls(self) -> None:
        keys = (
            "SFORGE_AGENT_EXTRA_ENV",
            "MAX_THINKING_TOKENS",
            "CLAUDE_CODE_DISABLE_THINKING",
            "CLAUDE_CODE_DISABLE_ADAPTIVE_THINKING",
            "CLAUDE_CODE_EFFORT_LEVEL",
            "CLAUDE_CODE_ALWAYS_ENABLE_EFFORT",
        )
        previous = {key: EDGE.os.environ.get(key) for key in keys}
        EDGE.os.environ["SFORGE_AGENT_EXTRA_ENV"] = (
            "EXISTING=value,MAX_THINKING_TOKENS=0,"
            "CLAUDE_CODE_DISABLE_THINKING=1,CLAUDE_CODE_EFFORT_LEVEL=high,"
            "CLAUDE_CODE_ALWAYS_ENABLE_EFFORT=1"
        )
        for key in keys[1:]:
            EDGE.os.environ[key] = "1"
        try:
            env = EDGE.cell_environment(
                {
                    "method": "plain-claude",
                    "sforge_agent": "claude-code",
                    "model": "glm-5.1",
                    "thinking": {"type": "adaptive"},
                    "reasoning_effort": None,
                    "claude_context_window_tokens": 200000,
                    "claude_autocompact_percent": 80,
                    "internet": False,
                }
            )
        finally:
            for key, value in previous.items():
                if value is None:
                    EDGE.os.environ.pop(key, None)
                else:
                    EDGE.os.environ[key] = value

        extra = dict(
            item.split("=", 1)
            for item in env["SFORGE_AGENT_EXTRA_ENV"].split(",")
        )
        self.assertEqual(extra["EXISTING"], "value")
        for key in (
            "ANTHROPIC_MODEL",
            "ANTHROPIC_DEFAULT_OPUS_MODEL",
            "ANTHROPIC_DEFAULT_SONNET_MODEL",
            "ANTHROPIC_DEFAULT_HAIKU_MODEL",
            "CLAUDE_CODE_SUBAGENT_MODEL",
        ):
            self.assertEqual(extra[key], "glm-5.1")
        self.assertEqual(extra["CLAUDE_CODE_AUTO_COMPACT_WINDOW"], "200000")
        self.assertEqual(extra["CLAUDE_AUTOCOMPACT_PCT_OVERRIDE"], "80")
        for key in keys[1:]:
            self.assertNotIn(key, extra)
            self.assertNotIn(key, env)

    def test_loopback_api_bridge_preserves_base_path(self) -> None:
        self.assertEqual(
            EDGE.loopback_api_target("http://127.0.0.1:3788/v1"),
            ("127.0.0.1", 3788),
        )
        self.assertEqual(
            EDGE.bridged_base_url(
                "http://127.0.0.1:3788/v1", "192.0.2.10", 45678
            ),
            "http://192.0.2.10:45678/v1",
        )
        self.assertIsNone(
            EDGE.loopback_api_target("https://api.example.com/v1")
        )

    def test_cell_environment_maps_api_key_and_bridge(self) -> None:
        env = EDGE.cell_environment(
            {
                "method": "plain-codex",
                "reasoning_effort": "medium",
            },
            api_key="runtime-key",
            api_base_url="http://192.0.2.10:45678/v1",
            bridge_host="192.0.2.10",
        )

        self.assertEqual(env["SFORGE_AGENT_API_KEY"], "runtime-key")
        self.assertEqual(
            env["SFORGE_AGENT_API_BASE_URL"],
            "http://192.0.2.10:45678/v1",
        )
        self.assertIn("192.0.2.10", env["SFORGE_NO_PROXY"].split(","))

    def test_judge_environment_uses_fixed_model_and_runtime_api(self) -> None:
        previous = EDGE.os.environ.pop("SFORGE_JUDGE_EXTRA_ENV", None)
        try:
            env = EDGE.judge_server_environment(
                api_key="judge-key",
                api_base_url="http://192.0.2.10:45678/v1",
                bridge_host="192.0.2.10",
            )
        finally:
            if previous is not None:
                EDGE.os.environ["SFORGE_JUDGE_EXTRA_ENV"] = previous

        values = dict(
            item.split("=", 1)
            for item in env["SFORGE_JUDGE_EXTRA_ENV"].split(",")
        )
        self.assertEqual(values["SFORGE_JUDGE_API_KEY"], "judge-key")
        self.assertEqual(
            values["SFORGE_JUDGE_API_BASE_URL"],
            "http://192.0.2.10:45678/v1",
        )
        self.assertEqual(values["SFORGE_JUDGE_MODEL"], "gpt-5.5")

    def test_child_proxy_is_rewritten_for_container_access(self) -> None:
        previous = {
            key: EDGE.os.environ.get(key)
            for key in ("SFORGE_HTTPS_PROXY", "HTTPS_PROXY")
        }
        EDGE.os.environ.pop("SFORGE_HTTPS_PROXY", None)
        EDGE.os.environ["HTTPS_PROXY"] = "http://127.0.0.1:3128"
        try:
            env = EDGE.cell_environment(
                {
                    "method": "plain-codex",
                    "reasoning_effort": "high",
                }
            )
        finally:
            for key, value in previous.items():
                if value is None:
                    EDGE.os.environ.pop(key, None)
                else:
                    EDGE.os.environ[key] = value

        self.assertEqual(
            env["SFORGE_HTTPS_PROXY"],
            "http://host.docker.internal:3128",
        )

    def test_isolated_cell_environment_removes_all_proxy_variables(self) -> None:
        proxy_keys = (
            "ALL_PROXY",
            "HTTP_PROXY",
            "HTTPS_PROXY",
            "SFORGE_HTTP_PROXY",
            "SFORGE_HTTPS_PROXY",
            "all_proxy",
            "http_proxy",
            "https_proxy",
        )
        previous = {key: EDGE.os.environ.get(key) for key in proxy_keys}
        for key in proxy_keys:
            EDGE.os.environ[key] = "http://proxy.example:3128"
        try:
            env = EDGE.cell_environment(
                {
                    "method": "plain-codex",
                    "reasoning_effort": "high",
                    "internet": False,
                }
            )
        finally:
            for key, value in previous.items():
                if value is None:
                    EDGE.os.environ.pop(key, None)
                else:
                    EDGE.os.environ[key] = value

        for key in proxy_keys:
            self.assertNotIn(key, env)

    def test_docker_resource_probe_verifies_applied_host_config(self) -> None:
        original = EDGE_IO.run_capture
        commands = []

        def fake_run_capture(command, *, env=None):
            commands.append(command)
            if command[1] == "run":
                return {"returncode": 0, "stdout": "container-id\n", "stderr": ""}
            if command[1] == "inspect":
                return {
                    "returncode": 0,
                    "stdout": json.dumps(
                        {"NanoCpus": 4_000_000_000, "Memory": 16 * 1024**3}
                    ),
                    "stderr": "",
                }
            return {"returncode": 0, "stdout": "", "stderr": ""}

        EDGE_IO.run_capture = fake_run_capture
        try:
            result = EDGE.docker_resource_limit_probe(
                "example:work", cpu_limit=4, mem_limit="16g"
            )
        finally:
            EDGE_IO.run_capture = original

        self.assertTrue(result["passed"])
        self.assertIn("--cpus", commands[0])
        self.assertIn("--memory", commands[0])
        self.assertEqual(commands[-1][1:3], ["rm", "--force"])

    def test_rust_runtime_probe_preserves_image_environment(self) -> None:
        original = EDGE_IO.run_capture
        captured = []

        def fake_run_capture(command, *, env=None):
            captured.append(command)
            return {"returncode": 0, "stdout": "rustc 1.88.0", "stderr": ""}

        EDGE_IO.run_capture = fake_run_capture
        try:
            result = EDGE.rust_image_runtime_probe("example:rust", "1.88.0")
        finally:
            EDGE_IO.run_capture = original

        self.assertEqual(result["returncode"], 0)
        self.assertIn("-c", captured[0])
        self.assertNotIn("-lc", captured[0])
        self.assertIn("command -v cargo", captured[0][-1])

    def test_codex_usage_reads_jsonl_agent_output(self) -> None:
        run = self.temp / "task-run"
        run.mkdir()
        (run / "agent_output.txt").write_text(
            "\n".join(
                [
                    '{"type":"thread.started","thread_id":"thread-1"}',
                    '{"type":"turn.completed","usage":{"input_tokens":11,"cached_input_tokens":3,"output_tokens":5}}',
                    "non-json status line",
                    '{"type":"turn.completed","usage":{"input_tokens":7,"output_tokens":2}}',
                ]
            ),
            encoding="utf-8",
        )

        usage = EDGE.codex_usage(run)

        self.assertEqual(usage["coverage"], "agent_output_only")
        self.assertEqual(usage["session_count"], 1)
        self.assertEqual(usage["tokens"]["input_tokens"], 18)
        self.assertEqual(usage["tokens"]["output_tokens"], 7)
        self.assertEqual(usage["tokens"]["cached_input_tokens"], 3)

    def test_codex_usage_reads_cumulative_rollout_tokens_once(self) -> None:
        run = self.temp / "task-run"
        run.mkdir()
        events = "\n".join(
            [
                '{"type":"session_meta","payload":{"id":"session-1"}}',
                '{"type":"event_msg","payload":{"type":"token_count","info":{"total_token_usage":{"input_tokens":11,"cached_input_tokens":3,"output_tokens":5,"total_tokens":16}}}}',
                '{"type":"event_msg","payload":{"type":"token_count","info":{"total_token_usage":{"input_tokens":18,"cached_input_tokens":7,"output_tokens":9,"total_tokens":27}}}}',
            ]
        ).encode()
        member = tarfile.TarInfo("sessions/rollout.jsonl")
        member.size = len(events)
        with tarfile.open(run / "codex-sessions.tar", "w") as archive:
            archive.addfile(member, io.BytesIO(events))

        usage = EDGE.codex_usage(run)

        self.assertEqual(usage["coverage"], "all_collected_codex_sessions")
        self.assertEqual(usage["session_count"], 1)
        self.assertEqual(usage["tokens"]["input_tokens"], 18)
        self.assertEqual(usage["tokens"]["cached_input_tokens"], 7)
        self.assertEqual(usage["tokens"]["output_tokens"], 9)
        self.assertEqual(usage["tokens"]["total_tokens"], 27)

    def test_goal_plus_stats_counts_empty_search_run(self) -> None:
        run = self.temp / "task-run"
        run.mkdir()
        payload = b'{"run_id":"run-1","state":"running"}'
        member = tarfile.TarInfo(".goal-plus/runs/run-1/run.json")
        member.size = len(payload)
        annotation = json.dumps(
            {
                "state": "completed",
                "attempts": 1,
                "usage": {
                    "input_tokens": 13,
                    "output_tokens": 5,
                    "cost_usd": 0.001,
                },
            }
        ).encode()
        annotation_member = tarfile.TarInfo(
            ".goal-plus/runs/run-1/candidates/c001/"
            "evidence-annotations/iteration-0001.json"
        )
        annotation_member.size = len(annotation)
        with tarfile.open(run / "goal-plus-state.tar", "w") as archive:
            archive.addfile(member, io.BytesIO(payload))
            archive.addfile(annotation_member, io.BytesIO(annotation))

        stats = EDGE.goal_plus_stats(run)

        self.assertIsNotNone(stats)
        assert stats is not None
        self.assertEqual(stats["search_runs"], 1)
        self.assertEqual(stats["candidates"], 0)
        self.assertEqual(stats["search_run_states"], {"running": 1})
        self.assertEqual(stats["selected_candidate_ids"], [])
        self.assertEqual(stats["promoted_candidate_ids"], [])
        self.assertEqual(
            stats["evidence_annotator_usage"],
            {
                "input_tokens": 13,
                "output_tokens": 5,
                "cost_usd": 0.001,
                "tasks": 1,
                "attempts": 1,
                "states": {"completed": 1},
                "coverage": "persisted Goal Plus Evidence annotator turns",
            },
        )

    def test_goal_plus_stats_recovers_archived_promotion(self) -> None:
        run = self.temp / "task-run"
        run.mkdir()
        payload = json.dumps(
            {
                "run_id": "run-1",
                "state": "promoted",
                "selected_candidate_id": "c001",
            }
        ).encode()
        member = tarfile.TarInfo(".goal-plus/runs/run-1/run.json")
        member.size = len(payload)
        with tarfile.open(run / "goal-plus-state.tar", "w") as archive:
            archive.addfile(member, io.BytesIO(payload))

        stats = EDGE.goal_plus_stats(run)

        self.assertIsNotNone(stats)
        assert stats is not None
        self.assertEqual(stats["search_run_states"], {"promoted": 1})
        self.assertEqual(stats["selected_candidate_ids"], ["c001"])
        self.assertEqual(stats["promoted_candidate_ids"], ["c001"])

    def test_provision_excludes_downloaded_tasks_from_git_status(self) -> None:
        exclude = EDGE.current_paths().edge_root / ".git" / "info" / "exclude"
        exclude.parent.mkdir(parents=True)
        exclude.write_text("# local excludes\n", encoding="utf-8")

        EDGE.ensure_local_task_exclude()
        EDGE.ensure_local_task_exclude()

        self.assertEqual(
            exclude.read_text(encoding="utf-8").splitlines().count("tasks/"),
            1,
        )

    def test_command_keeps_plain_replicas_distinct_from_goal_plus_workers(self) -> None:
        destination = self.temp / "campaign"
        plain = {
            "cell_id": "plain",
            "task_id": "vliw_kernel_optimization",
            "sforge_agent": "codex",
            "model": "gpt-test",
            "wall_time_seconds": 300,
            "eval_interval_seconds": 120,
            "sforge_run_id": "run-plain",
            "outer_replicas": 4,
            "outer_replica_concurrency": 4,
            "judge_concurrency": 1,
            "judge_port": 8080,
            "work_cpu_limit": 4,
            "work_mem_limit": "16g",
            "judge_cpu_limit": 4,
            "judge_mem_limit": "8g",
            "submission_cooldown": 0,
            "max_submissions": 7,
            "auto_eval_enabled": False,
            "auto_resume_enabled": False,
            "stop_hook_enabled": False,
            "internet": False,
        }

        command = EDGE.build_sforge_command(destination, plain)

        self.assertLess(command.index("--silent"), command.index("run"))
        self.assertEqual(command[command.index("--replicas") + 1], "4")
        self.assertEqual(command[command.index("--replica-concurrency") + 1], "4")
        self.assertEqual(
            command[command.index("--judge-url") + 1],
            "http://host.docker.internal:8080",
        )
        self.assertEqual(command[command.index("--work-cpu-limit") + 1], "4")
        self.assertEqual(command[command.index("--work-mem-limit") + 1], "16g")
        self.assertEqual(command[command.index("--judge-cpu-limit") + 1], "4")
        self.assertEqual(command[command.index("--judge-mem-limit") + 1], "8g")
        self.assertEqual(command[command.index("--submission-cooldown") + 1], "0")
        self.assertEqual(command[command.index("--max-submissions") + 1], "7")
        self.assertIn("--disable-auto-eval", command)
        self.assertIn("--disable-auto-resume", command)
        self.assertIn("--disable-stop-hook", command)
        self.assertIn("--disable-internet", command)

    def test_goal_plus_codex_completion_requires_real_spawn_and_verifier_evidence(
        self,
    ) -> None:
        cell = {
            "method": "goal-plus-codex",
            "outer_replicas": 1,
            "inner_search_concurrency": 2,
        }
        complete = {
            "edgebench_score": 50.0,
            "goal_plus": {
                "candidates": 2,
                "agent_sessions": 2,
                "worker_verifier_runs": 2,
            },
            "agent_events": {
                "spawn_agent_completed_count": 2,
                "spawned_agent_thread_count": 2,
                "goal_plus": {
                    "candidate_ids": ["c001", "c002"],
                    "agent_session_ids": ["a001", "a002"],
                    "verifier_ledger": [
                        {"candidate_id": "c001"},
                        {"candidate_id": "c002"},
                    ],
                    "selected_candidate_ids": ["c001"],
                    "promoted_candidate_ids": ["c001"],
                },
            },
        }

        passed = EDGE.goal_plus_completion_evidence(
            cell, [complete], valid_trajectories=1
        )
        missing_spawn = EDGE.goal_plus_completion_evidence(
            cell,
            [
                {
                    **complete,
                    "agent_events": {
                        **complete["agent_events"],
                        "spawn_agent_completed_count": 0,
                        "spawned_agent_thread_count": 0,
                    },
                }
            ],
            valid_trajectories=1,
        )
        too_many_spawns = EDGE.goal_plus_completion_evidence(
            cell,
            [
                {
                    **complete,
                    "agent_events": {
                        **complete["agent_events"],
                        "spawn_agent_completed_count": 3,
                        "spawned_agent_thread_count": 3,
                    },
                }
            ],
            valid_trajectories=1,
        )

        self.assertTrue(passed["passed"])
        self.assertFalse(missing_spawn["passed"])
        self.assertFalse(too_many_spawns["passed"])
        self.assertEqual(
            missing_spawn["checks"]["actual_worker_launches"],
            {"expected": 2, "actual": 0},
        )
        self.assertEqual(
            too_many_spawns["checks"]["actual_worker_launches"],
            {"expected": 2, "actual": 3},
        )

    def test_goal_plus_pi_completion_uses_persisted_session_evidence(self) -> None:
        cell = {
            "method": "goal-plus-pi",
            "outer_replicas": 1,
            "inner_search_concurrency": 2,
        }
        complete = {
            "edgebench_score": 40.0,
            "goal_plus": {
                "candidates": 2,
                "agent_sessions": 2,
                "worker_verifier_runs": 3,
                "verifier_candidate_ids": ["c001", "c002"],
                "selected_candidate_ids": ["c001"],
                "promoted_candidate_ids": ["c001"],
            },
            "agent_events": {
                "spawn_agent_completed_count": 0,
                "goal_plus": {
                    "candidate_ids": [],
                    "agent_session_ids": [],
                    "verifier_ledger": [],
                    "selected_candidate_ids": [],
                    "promoted_candidate_ids": [],
                },
            },
        }
        evidence = EDGE.goal_plus_completion_evidence(
            cell,
            [complete],
            valid_trajectories=1,
        )
        too_many_sessions = EDGE.goal_plus_completion_evidence(
            cell,
            [
                {
                    **complete,
                    "goal_plus": {
                        **complete["goal_plus"],
                        "agent_sessions": 3,
                    },
                }
            ],
            valid_trajectories=1,
        )

        self.assertTrue(evidence["passed"])
        self.assertFalse(too_many_sessions["passed"])
        self.assertNotIn("actual_worker_launches", evidence["checks"])
        self.assertEqual(
            too_many_sessions["checks"]["agent_sessions"],
            {"expected": 2, "actual": 3},
        )

    def test_finalize_downgrades_missing_goal_plus_evidence_to_partial(self) -> None:
        destination = self.temp / "campaign-finalize"
        cell_dir = destination / "cells" / "vliw--goal-plus-codex"
        cell_dir.mkdir(parents=True)
        campaign = {
            "campaign_id": "campaign-finalize",
            "state": "completed",
            "task_ids": ["vliw_kernel_optimization"],
            "cells": [
                {
                    "cell_id": "vliw--goal-plus-codex",
                    "task_id": "vliw_kernel_optimization",
                    "method": "goal-plus-codex",
                    "state": "completed",
                }
            ],
        }
        cell = {
            "cell_id": "vliw--goal-plus-codex",
            "task_id": "vliw_kernel_optimization",
            "method": "goal-plus-codex",
            "state": "completed",
        }
        (destination / "campaign.json").write_text(json.dumps(campaign))
        (cell_dir / "cell.json").write_text(json.dumps(cell))
        original_summary = EDGE_REPORTING.summarize_cell
        original_reference = EDGE_REPORTING.load_paper_reference
        original_workbook = EDGE_REPORTING.write_comparison_workbook
        EDGE_REPORTING.summarize_cell = lambda *_args, **_kwargs: {
            "cell_id": "vliw--goal-plus-codex",
            "task_id": "vliw_kernel_optimization",
            "model": "gpt-test",
            "reasoning_effort": "medium",
            "wall_time_seconds": 60,
            "live_search_concurrency": 2,
            "completion_evidence": {"passed": False},
            "incomplete_reason": "missing worker evidence",
        }
        EDGE_REPORTING.load_paper_reference = lambda: {
            "tasks": {"vliw_kernel_optimization": {}}
        }
        EDGE_REPORTING.write_comparison_workbook = lambda *_args, **_kwargs: None
        try:
            payload = EDGE.finalize_campaign(destination)
        finally:
            EDGE_REPORTING.summarize_cell = original_summary
            EDGE_REPORTING.load_paper_reference = original_reference
            EDGE_REPORTING.write_comparison_workbook = original_workbook

        self.assertFalse(payload["completion_evidence_passed"])
        self.assertEqual(
            json.loads((destination / "campaign.json").read_text())["state"],
            "partial",
        )
        self.assertEqual(
            json.loads((cell_dir / "cell.json").read_text())["state"],
            "partial",
        )

    def test_finalize_recovers_prior_evidence_downgrade(self) -> None:
        destination = self.temp / "campaign-recover"
        cell_id = "vliw--goal-plus-pi"
        cell_dir = destination / "cells" / cell_id
        cell_dir.mkdir(parents=True)
        incomplete_reason = "missing promotion evidence"
        campaign = {
            "campaign_id": "campaign-recover",
            "state": "partial",
            "completion_evidence_passed": True,
            "incomplete_cells": {cell_id: incomplete_reason},
            "task_ids": ["vliw_kernel_optimization"],
            "cells": [
                {
                    "cell_id": cell_id,
                    "task_id": "vliw_kernel_optimization",
                    "method": "goal-plus-pi",
                    "state": "partial",
                    "incomplete_reason": incomplete_reason,
                }
            ],
        }
        cell = {
            "cell_id": cell_id,
            "task_id": "vliw_kernel_optimization",
            "method": "goal-plus-pi",
            "state": "partial",
            "incomplete_reason": incomplete_reason,
        }
        controller = {
            "state": "partial",
            "returncode": 2,
            "completion_evidence_passed": False,
        }
        (destination / "campaign.json").write_text(json.dumps(campaign))
        (destination / "controller.json").write_text(json.dumps(controller))
        (cell_dir / "cell.json").write_text(json.dumps(cell))
        original_summary = EDGE_REPORTING.summarize_cell
        original_reference = EDGE_REPORTING.load_paper_reference
        original_workbook = EDGE_REPORTING.write_comparison_workbook
        EDGE_REPORTING.summarize_cell = lambda *_args, **_kwargs: {
            "cell_id": cell_id,
            "task_id": "vliw_kernel_optimization",
            "model": "gpt-test",
            "reasoning_effort": "medium",
            "wall_time_seconds": 60,
            "live_search_concurrency": 2,
            "completion_evidence": {"passed": True},
            "incomplete_reason": None,
        }
        EDGE_REPORTING.load_paper_reference = lambda: {
            "tasks": {"vliw_kernel_optimization": {}}
        }
        EDGE_REPORTING.write_comparison_workbook = lambda *_args, **_kwargs: None
        try:
            payload = EDGE.finalize_campaign(destination)
        finally:
            EDGE_REPORTING.summarize_cell = original_summary
            EDGE_REPORTING.load_paper_reference = original_reference
            EDGE_REPORTING.write_comparison_workbook = original_workbook

        recovered_campaign = json.loads(
            (destination / "campaign.json").read_text()
        )
        recovered_cell = json.loads((cell_dir / "cell.json").read_text())
        recovered_controller = json.loads(
            (destination / "controller.json").read_text()
        )
        self.assertTrue(payload["completion_evidence_passed"])
        self.assertEqual(recovered_campaign["state"], "completed")
        self.assertTrue(recovered_campaign["completion_evidence_passed"])
        self.assertNotIn("incomplete_cells", recovered_campaign)
        self.assertEqual(recovered_campaign["cells"][0]["state"], "completed")
        self.assertNotIn("incomplete_reason", recovered_campaign["cells"][0])
        self.assertEqual(recovered_cell["state"], "completed")
        self.assertNotIn("incomplete_reason", recovered_cell)
        self.assertEqual(recovered_controller["state"], "completed")
        self.assertEqual(recovered_controller["returncode"], 0)
        self.assertTrue(recovered_controller["completion_evidence_passed"])


if __name__ == "__main__":
    unittest.main()
